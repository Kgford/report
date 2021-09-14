import xlwt
from xlutils.copy import copy # http://pypi.python.org/pypi/xlutils
from xlrd import open_workbook # http://pypi.python.org/pypi/xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
from django.http import HttpResponse
from report.overhead import TimeCode, Security, StringThings,Conversions
from test_db.models import Specifications,Workstation,Workstation1,Testdata,Testdata3,Trace,Tracepoints,Tracepoints2,Effeciency



class ExcelReports:
    def __init__ (self, job_num,operator,workstation):
        self.job_num = job_num
        self.operator = operator
        self.workstation = workstation
        print('job_num=',self.job_num)
        
        
    def coupler_90_deg(self):
        job_list = Testdata.objects.using('TEST').filter(jobnumber=job_num).order_by('jobnumber').values_list('jobnumber', flat=True).distinct()
		part_list = Testdata.objects.using('TEST').filter(jobnumber=job_num).order_by('partnumber').values_list('partnumber', flat=True).distinct()
		report_data = Testdata.objects.using('TEST').filter(jobnumber=job_num).all()
		spec_data = Specifications.objects.using('TEST').filter(jobnumber=job_num).first()
		conversions = Conversions(spec_data.vswr.'')
		spec_rl = conversions.vswr_to_rl
		print('spec_data=',spec_data)
		spec_list = [spec_data.insertionloss,spec_rl,spec_data.isolation,spec_data.amplitudebalance,spec_data.phasebalance]
		print('spec_list=',spec_list)
		if report_data:
			part_num = report_data[0].partnumber
			spectype = spec_data.spectype
			path = 'C:/ATE Data/excel_templates/'
			print('path=',path)
			file = os.path.join(path, 'TestData.xlsx')
			print('file=',file)

			wb = load_workbook(file)
			print('wb=',wb)
			sheet = wb["Raw Data1"]
			print('sheet=',sheet)
			sheet['F2'] = job_num
			sheet['F3'] = part_num 
			sheet['F4'] = spectype 
			sheet['J6'] = str(spec_list[0]) + ' Max'
			sheet['K6'] = str(spec_list[1]) + ' Max'
			sheet['L6'] = str(spec_list[2]) + ' Max'
			sheet['M6'] = "+/- " + str(spec_list[3]) + ' dB'
			sheet['N6'] = "+/- " + str(spec_list[4]) + ' deg'
			
			#Tabular data
			rownum = 6
			insertion_loss = []
			return_loss = []
			isolation = []
			amplitude_balance = []
			phase_balance = []
			stat_list = []
			il_pass = 0
			rl_pass = 0
			iso_pass = 0
			ab_pass = 0
			pb_pass = 0
			il_fail = 0
			rl_fail = 0
			iso_fail = 0
			ab_fail = 0
			pb_fail = 0
			for data in report_data:
				sheet.cell(row=rownum, column=1).value= data.serialnumber
				sheet.cell(row=rownum, column=2).value= data.workstation
				sheet.cell(row=rownum, column=3).value= round(data.insertionloss,2)
				insertion_loss.append(data.insertionloss)
				if data.insertionloss <= spec_list[0]:
					il_pass+=1
				else:
					il_fail+=1
				
				sheet.cell(row=rownum, column=4).value= round(data.returnloss,2)
				return_loss.append(data.returnloss)
				if data.returnloss <= spec_list[1]:
					rl_pass+=1
				else:
					rl_fail+=1
				
				sheet.cell(row=rownum, column=5).value= round(data.isolation,2)
				isolation.append(data.isolation)
				if data.isolation <= spec_list[2]:
					iso_pass+=1
				else:
					iso_fail+=1
				
				sheet.cell(row=rownum, column=6).value= round(data.amplitudebalance,2)
				amplitude_balance.append(data.amplitudebalance)
				if data.amplitudebalance <= spec_list[3]:
					ab_pass+=1
				else:
					ab_fail+=1
				
				sheet.cell(row=rownum, column=7).value= round(data.phasebalance,2)
				phase_balance.append(data.phasebalance)
				if data.phasebalance <= spec_list[4]:
					pb_pass+=1
				else:
					pb_fail+=1
				rownum +=1
			list_names = ['Min','Max','Avg','Stdev']
			print('list_names=',list_names)
			il_stdev = round(statistics.stdev(insertion_loss),2) #Standard deviation
			il_var = round(statistics.variance(insertion_loss),2) #Variance
			il_avg = round(statistics.mean(insertion_loss),2) #Mean Average
			il_min = round(min(insertion_loss),2) #Min
			il_max = round(max(insertion_loss),2) #Max
			sheet['J7'] = il_avg
			sheet['J8'] = il_min
			sheet['J9'] = il_max
			sheet['J10'] = il_stdev
			sheet['J11'] = il_pass
			sheet['J12'] = il_fail
			sheet['J13'] = il_fail/rownum
			il_list = [il_min,il_max,il_avg,il_stdev]
			print('il_list=',il_list)
			
			rl_stdev = round(statistics.stdev(return_loss),2) #Standard deviation
			rl_var = round(statistics.variance(return_loss),2) #Variance
			rl_avg = round(statistics.mean(return_loss),2) #Mean Average
			rl_min = round(min(return_loss),2) #Min
			rl_max = round(max(return_loss),2) #Max
			rl_list = [rl_min,rl_max,rl_avg,rl_stdev]
			sheet['K7'] = rl_avg
			sheet['K8'] = rl_min
			sheet['K9'] = rl_max
			sheet['K10'] = rl_stdev
			sheet['K11'] = rl_pass
			sheet['K12'] = rl_fail
			sheet['K13'] = rl_fail/rownum
			print('rl_list=',rl_list)
			
			iso_stdev = round(statistics.stdev(isolation),2) #Standard deviation
			iso_var = round(statistics.variance(isolation),2) #Variance
			iso_avg = round(statistics.mean(isolation),2) #Mean Average
			iso_min = round(min(isolation),2) #Min
			iso_max = round(max(isolation),2) #Max
			iso_list = [iso_min,iso_max,iso_avg,iso_stdev]
			sheet['L7'] = iso_avg
			sheet['L8'] = iso_min
			sheet['L9'] = iso_max
			sheet['L10'] = iso_stdev
			sheet['L11'] = iso_pass
			sheet['L12'] = iso_fail
			sheet['L13'] = iso_fail/rownum
			print('iso_list=',iso_list)
		   
			ab_stdev = round(statistics.stdev(amplitude_balance),2) #Standard deviation
			ab_var = round(statistics.variance(amplitude_balance),2) #Variance
			ab_avg = round(statistics.mean(amplitude_balance),2) #Mean Average
			ab_min = round(min(amplitude_balance),2) #Min
			ab_max = round(max(amplitude_balance),2) #Max
			ab_list = [ab_min,ab_max,ab_avg,ab_stdev]
			sheet['M7'] = ab_avg
			sheet['M8'] = ab_min
			sheet['M9'] = ab_max
			sheet['M10'] = ab_stdev
			sheet['M11'] = ab_pass
			sheet['M12'] = ab_fail
			sheet['M13'] = ab_fail/rownum
			print('ab_list=',ab_list)
			
			pb_stdev = round(statistics.stdev(phase_balance),2) #Standard deviation
			pb_var = round(statistics.variance(phase_balance),2) #Variance
			pb_avg = round(statistics.mean(phase_balance),2) #Mean Average
			pb_min = round(min(phase_balance),2) #Min
			pb_max = round(max(phase_balance),2) #Max
			pb_list = [pb_min,pb_max,pb_avg,pb_stdev]
			sheet['N7'] = pb_avg
			sheet['N8'] = pb_min
			sheet['N9'] = pb_max
			sheet['N10'] = pb_stdev
			sheet['N11'] = pb_pass
			sheet['N12'] = pb_fail
			sheet['N13'] = pb_fail/rownum
			print('pb_list=',pb_list)
			
			stat_list = [il_list,rl_list,iso_list,ab_list,pb_list]
			print('stat_list=',stat_list)
			
			#~~~~~~~~~~~~~~~~~~~~~~Summary sheet~~~~~~~~~~~~~~~~~~~~~~~~
			sheet = wb["Summary"]
			print('sheet=',sheet)
			sheet['B1'] = spec_list[0] 
			sheet['C1'] = spec_list[1]
			sheet['D1'] = spec_list[2] 
			sheet['E1'] = spec_list[3] 
			sheet['F1'] = spec_list[4]
			#AVG
			sheet['B4'] = str(spec_list[0]) + ' Max'
			sheet['C4'] = str(spec_list[1]) + ' Max'
			sheet['D4'] = str(spec_list[2]) + ' Max'
			sheet['E4'] = "'+/- " + str(spec_list[3]) + ' dB'
			sheet['F4'] = "'+/- " + str(spec_list[4]) + ' deg'
			sheet['B5'] = il_avg
			sheet['C5'] = rl_avg
			sheet['D5'] = iso_avg
			sheet['E5'] = ab_avg
			sheet['F5'] = pb_avg
			sheet['G5'] = rownum
			#MIN
			sheet['B18'] = str(spec_list[0]) + ' Max'
			sheet['C18'] = str(spec_list[1]) + ' Max'
			sheet['D18'] = str(spec_list[2]) + ' Max'
			sheet['E18'] = "+/- " + str(spec_list[3]) + ' dB'
			sheet['F18'] = "+/- " + str(spec_list[4]) + ' deg'
			sheet['B19'] = il_min
			sheet['C19'] = rl_min
			sheet['D19'] = iso_min
			sheet['E19'] = ab_min
			sheet['F19'] = pb_min
			sheet['G19'] = rownum
			#Max
			sheet['B32'] = str(spec_list[0]) + ' Max'
			sheet['C32'] = str(spec_list[1]) + ' Max'
			sheet['D32'] = str(spec_list[2]) + ' Max'
			sheet['E32'] = "+/- " + str(spec_list[3]) + ' dB'
			sheet['F32'] = "+/- " + str(spec_list[4]) + ' deg'
			sheet['B33'] = il_max
			sheet['C33'] = rl_max
			sheet['D33'] = iso_max
			sheet['E33'] = ab_max
			sheet['F33'] = pb_max
			sheet['G33'] = rownum
			#~~~~~~~~~~~~~~~~~~~~~~Summary sheet~~~~~~~~~~~~~~~~~~~~~~~~
			
			
			
			#~~~~~~~~~~~~~~~~~~~~~~Chart~~~~~~~~~~~~~~~~~~~~~~~~
			trace_num = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Insertion Loss J3').count()
			
			if trace_num==5:
				#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~UUT 1 ~~~~~~~~~~~~~~~~~~~~~~~~
				sheet = wb["UUT 1"]
				#*****************************chart1 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J3~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Insertion Loss J3').filter(serialnumber='UUT 1').values_list('id').first()
				print('trace_id=',trace_id[0])
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').filter(traceid=trace_id[0]).all()
				print('trace_points=',trace_points)
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=1).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=2).value= round(point.ydata,0)
					print('rownum=',rownum,' point.xdata=',point.xdata)
					rownum+=1
				 #~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Insertion Loss J4').filter(serialnumber='UUT 1').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=3).value= round(point.ydata,0)
					rownum+=1    
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				#*****************************chart1 *****************************
				
				#*****************************chart2 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~Return Loss~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Return Loss').filter(serialnumber='UUT 1').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=4).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=5).value= round(point.ydata,0)
					rownum+=1
				#~~~~~~~~~~~~~~~~~~~~~~Return Loss~~~~~~~~~~~~~~~~~~~~~~~~
				#*****************************chart2 *****************************
				
				#*****************************chart3 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~isolation~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Return Loss').filter(serialnumber='UUT 1').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=6).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=7).value= round(point.ydata,0)
					rownum+=1
				#*****************************chart3 *****************************
				#*****************************chart4 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J3~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Phase Balance J3').filter(serialnumber='UUT 1').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=8).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=9).value= round(point.ydata,0)
					rownum+=1
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Phase Balance J4').filter(serialnumber='UUT 1').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=10).value= round(point.ydata,0)
					rownum+=1    
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				#*****************************chart4 *****************************
				
				#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~UUT 1 ~~~~~~~~~~~~~~~~~~~~~~~~
				
				#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~UUT 2 ~~~~~~~~~~~~~~~~~~~~~~~~
				sheet = wb["UUT 2"]
				#*****************************chart1 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J3~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Insertion Loss J3').filter(serialnumber='UUT 2').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=1).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=2).value= round(point.ydata,0)
					rownum+=1
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Insertion Loss J4').filter(serialnumber='UUT 2').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=3).value= round(point.ydata,0)
					rownum+=1    
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				#*****************************chart1 *****************************
				
				#*****************************chart2 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~Return Loss~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Return Loss').filter(serialnumber='UUT 2').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=4).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=5).value= round(point.ydata,0)
					rownum+=1
				#~~~~~~~~~~~~~~~~~~~~~~Return Loss~~~~~~~~~~~~~~~~~~~~~~~~
				#*****************************chart2 *****************************
				
				#*****************************chart3 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~isolation~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Return Loss').filter(serialnumber='UUT 2').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=6).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=7).value= round(point.ydata,0)
					rownum+=1
				#*****************************chart3 *****************************
				#*****************************chart4 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J3~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Phase Balance J3').filter(serialnumber='UUT 2').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=8).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=9).value= round(point.ydata,0)
					rownum+=1
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Phase Balance J4').filter(serialnumber='UUT 2').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=10).value= round(point.ydata,0)
					rownum+=1    
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				#*****************************chart4 *****************************
				
				#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~UUT 2 ~~~~~~~~~~~~~~~~~~~~~~~~
				
				#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~UUT 3 ~~~~~~~~~~~~~~~~~~~~~~~~
				sheet = wb["UUT 3"]
				#*****************************chart1 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J3~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Insertion Loss J3').filter(serialnumber='UUT 3').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=1).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=2).value= round(point.ydata,0)
					rownum+=1
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Insertion Loss J4').filter(serialnumber='UUT 3').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=3).value= round(point.ydata,0)
					rownum+=1    
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				#*****************************chart1 *****************************
				
				#*****************************chart2 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~Return Loss~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Return Loss').filter(serialnumber='UUT 3').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=4).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=5).value= round(point.ydata,0)
					rownum+=1
				#~~~~~~~~~~~~~~~~~~~~~~Return Loss~~~~~~~~~~~~~~~~~~~~~~~~
				#*****************************chart2 *****************************
				
				#*****************************chart3 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~isolation~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Return Loss').filter(serialnumber='UUT 3').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=6).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=7).value= round(point.ydata,0)
					rownum+=1
				#*****************************chart3 *****************************
				#*****************************chart4 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J3~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Phase Balance J3').filter(serialnumber='UUT 3').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=8).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=9).value= round(point.ydata,0)
					rownum+=1
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Phase Balance J4').filter(serialnumber='UUT 3').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=10).value= round(point.ydata,0)
					rownum+=1    
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				#*****************************chart4 *****************************
				
				#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~UUT 3 ~~~~~~~~~~~~~~~~~~~~~~~~
				
				#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~UUT 4 ~~~~~~~~~~~~~~~~~~~~~~~~
				sheet = wb["UUT 4"]
				#*****************************chart1 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J3~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Insertion Loss J3').filter(serialnumber='UUT 4').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=1).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=2).value= round(point.ydata,0)
					rownum+=1
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Insertion Loss J4').filter(serialnumber='UUT 4').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=3).value= round(point.ydata,0)
					rownum+=1    
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				#*****************************chart1 *****************************
				
				#*****************************chart2 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~Return Loss~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Return Loss').filter(serialnumber='UUT 4').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=4).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=5).value= round(point.ydata,0)
					rownum+=1
				#~~~~~~~~~~~~~~~~~~~~~~Return Loss~~~~~~~~~~~~~~~~~~~~~~~~
				#*****************************chart2 *****************************
				
				#*****************************chart3 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~isolation~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Return Loss').filter(serialnumber='UUT 4').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=6).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=7).value= round(point.ydata,0)
					rownum+=1
				#*****************************chart3 *****************************
				#*****************************chart4 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J3~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Phase Balance J3').filter(serialnumber='UUT 4').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=8).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=9).value= round(point.ydata,0)
					rownum+=1
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Phase Balance J4').filter(serialnumber='UUT 4').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=10).value= round(point.ydata,0)
					rownum+=1    
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				#*****************************chart4 *****************************
				
				#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~UUT 4 ~~~~~~~~~~~~~~~~~~~~~~~~
				
				#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~UUT 5 ~~~~~~~~~~~~~~~~~~~~~~~~
				sheet = wb["UUT 5"]
				#*****************************chart1 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J3~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Insertion Loss J3').filter(serialnumber='UUT 5').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=1).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=2).value= round(point.ydata,0)
					rownum+=1
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Insertion Loss J4').filter(serialnumber='UUT 5').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=3).value= round(point.ydata,0)
					rownum+=1    
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				#*****************************chart1 *****************************
				
				#*****************************chart2 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~Return Loss~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Return Loss').filter(serialnumber='UUT 5').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=4).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=5).value= round(point.ydata,0)
					rownum+=1
				#~~~~~~~~~~~~~~~~~~~~~~Return Loss~~~~~~~~~~~~~~~~~~~~~~~~
				#*****************************chart2 *****************************
				
				#*****************************chart3 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~isolation~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Return Loss').filter(serialnumber='UUT 5').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=6).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=7).value= round(point.ydata,0)
					rownum+=1
				#*****************************chart3 *****************************
				#*****************************chart4 *****************************
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J3~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Phase Balance J3').filter(serialnumber='UUT 5').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=8).value= round(point.xdata,0)
					sheet.cell(row=rownum, column=9).value= round(point.ydata,0)
					rownum+=1
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				trace_id = Trace.objects.using('TEST').filter(jobnumber=job_num).filter(title='Phase Balance J4').filter(serialnumber='UUT 5').values_list('id').first()
				if trace_id[0] > 171666:
					trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
				else:
					trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
				rownum=56
				for point in trace_points:
					sheet.cell(row=rownum, column=10).value= round(point.ydata,0)
					rownum+=1    
				#~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
				#*****************************chart4 *****************************
				
				#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~UUT 5 ~~~~~~~~~~~~~~~~~~~~~~~~
				 
				# Clean up the template
				print(wb.sheetnames)
				sheetDelete = wb["Raw Data2"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["Raw Data3"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["Raw Data4"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 6"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 7"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 8"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 9"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 10"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 11"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 12"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 13"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 14"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 15"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 16"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 17"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 18"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 19"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 20"]
				wb.remove(sheetDelete)  #Sheet will be deleted
			elif trace_num==10:
				sheetDelete = wb["Raw Data3"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["Raw Data4"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 11"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 12"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 13"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 14"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 15"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 16"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 17"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 18"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 19"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 20"]
				wb.remove(sheetDelete)  #Sheet will be deleted
			elif trace_num==15:
				sheetDelete = wb["Raw Data4"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 11"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 12"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 13"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 14"]
				wb.remove(sheetDelete)  #Sheet will be deleted
				sheetDelete = wb["UUT 15"]
				wb.remove(sheetDelete)  #Sheet will be deleted
			#~~~~~~~~~~~~~~~~~~~~~~Clean up the Template~~~~~~~~~~~~~~~~~~~~~~~~
			
			#~~~~~~~~~~~~~~~~~~~~~~chart1~~~~~~~~~~~~~~~~~~~~~~~~
				
			
								
			
			wb.save("C:/ATE Data/demo4.xlsx")