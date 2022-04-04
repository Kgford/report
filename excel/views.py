from django.shortcuts import render
import xlwt
from xlutils.copy import copy # http://pypi.python.org/pypi/xlutils
from xlrd import open_workbook # http://pypi.python.org/pypi/xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
from django.http import HttpResponse
from django.contrib.auth.models import User
import os
from django import forms
from django.views import View
from report.overhead import TimeCode, Security, StringThings,Conversions
from report.reports import ExcelReports,Statistics,Histogram_data,XY_Chart,X_Range,SDEV_Dist,CreateSheets
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect
from django.http import JsonResponse
from django.core import serializers
from django.core.files import File
#from .forms import QAForm
from django.urls import reverse, reverse_lazy
from django.db.models import Q
from django.utils.timezone import make_aware
import re
import datetime
import io
from io import BytesIO
import shutil
import getpass
import subprocess
import sys
from base64 import *
import pygal
from pygal.style import Style
from report.pygal_extended import LineHist,LineBar
from test_db.models import Specifications,Workstation,Workstation1,Testdata,Testdata3,Trace,Tracepoints,Tracepoints2,Effeciency,ReportQueue

import numpy as np
import matplotlib.pyplot as plt
import scipy.stats as st
import seaborn as sns
from scipy.stats import norm
import statistics
# Create a class extending Graph


class ReportView(View):
    #~~~~~~~~~~~Load Item database from csv. must put this somewhere else later"
    contSuccess = 0
    template_name = "index.html"
    success_url = reverse_lazy('excel:reports')
    def get(self, request, *args, **kwargs):
        operator = self.request.user
        form = 0
        try:
            status ='In Process'
            job_num=-1
            part_num=-1
            workstation=-1
            operator=-1
            start_date=-1
            search=-1
            end_date = -1
            spectype = -1
            spec1 = -1
            spec2 = -1
            spec3 = -1
            spec4 = -1
            spec5 = -1
            report_data = -1
            analyze =-1
            artwork = -1
            stat1_min = -1
            stat1_max = -1
            stat1_avg = -1
            stat1_std = -1
            stat2_min = -1
            stat2_max = -1
            stat2_avg = -1
            stat2_std = -1
            stat3_min = -1
            stat3_max = -1
            stat3_avg = -1
            stat3_std = -1
            stat4_min = -1
            stat4_max = -1
            stat4_avg = -1
            stat4_std = -1
            stat5_min = -1
            stat5_max = -1
            stat5_avg = -1
            stat5_std = -1
            il_histo_data = -1
            rl_histo_data = -1
            iso_histo_data = -1
            ab_histo_data = -1
            pb_histo_data = -1
            coup_histo_data = -1
            dir_histo_data = -1
            cb_histo_data = -1
            il_histo_data2 = -1
            rl_histo_data2 = -1
            iso_histo_data2 = -1
            ab_histo_data2 = -1
            pb_histo_data2 = -1
            coup_histo_data2 = -1
            dir_histo_data2 = -1
            cb_histo_data2 = -1
            chart1 = -1
            chart2 = -1
            chart3 = -1
            chart4 = -1
            chart5 = -1
            
            job_list = []
            part_list = []
            workstation_list = []
            operator_list = []
            stat_list = []
            spec_list = -1
            artwork_list = ['RawData 1',]
            art_rev_list = []
            test1_list = []
            test2_list = []
            test3_list = []
            test4_list = []
            test5_list = []
            artwork_list = ['RawData 1',]
            bad1_list = []
            bad2_list = []
            bad3_list = []
            bad4_list = []
            bad4_list = []
            bad5_list = []
           
            blank = 0
            total = 0
            bad_data1 = 0
            passed1 = 0
            failed1 = 0
            failed_percent1 = 0
            bad_data2 = 0
            passed2 = 0
            failed2 = 0
            failed_percent2 = 0
            bad_data3 = 0
            passed3 = 0
            failed3 = 0
            failed_percent3 = 0
            bad_data4 = 0
            passed4 = 0
            failed4 = 0
            failed_percent4 = 0
            bad_data5 = 0
            passed5 = 0
            failed5 = 0
            failed_percent5 = 0
            test_status1 = -1
            test_status2 = -1
            test_status3 = -1
            test_status4 = -1
            test_status5 = -1
            test_status6 = -1
            test_status7 = -1
            test_status8 = -1
            test_status9 = -1
            test_status10 = -1
            test_comment1 = 1
            test_comment2 = -1
            test_comment3 = -1
            test_comment4 = -1
            test_comment5 = -1
            test_comment6 = -1
            test_comment7 = -1
            test_comment8 = -1
            test_comment9 = -1
            test_comment10 = -1
           
            #  Equations to get today - days
            #~~~~~~~~~~~~~ Time ~~~~~~~~~~~~~~~~~
            days=30 # start_date is today - days 
            time_code = TimeCode(days)
            friday = time_code.friday()
            print('friday=',friday)
            today = datetime.datetime.today()
            today = make_aware(today)
            this_time = today.time()
            print('today =', today)
            #start_date  = time_code.today_minus() # end_date is today - days 
            #start_date = make_aware(start_date)
            #end_date = today
            print('start_date =',start_date)
            print('end_date =',end_date)
            year = time_code.this_year()
            month_num = time_code.this_month()
            month_string = time_code.month_string()
            day = time_code.this_day()
            hour = time_code.this_hour()
            minute = time_code.this_minute()
            sec = time_code.this_sec()
            print('Today=',day,'/',month_num,'/',year,'/ ',hour,':',minute,':',sec)
            print('Month=',month_string)
            #~~~~~~~~~~~~~ Time ~~~~~~~~~~~~~~~~~
            
            
            workstation_list = Workstation.objects.using('TEST').order_by('computername').values_list('computername', flat=True).distinct()
            operator_list = Effeciency.objects.using('TEST').order_by('operator').values_list('operator', flat=True).distinct()
            job_list = Testdata.objects.using('TEST').order_by('jobnumber').values_list('jobnumber', flat=True).order_by('-jobnumber').distinct()
            part_list = Testdata.objects.using('TEST').order_by('partnumber').values_list('partnumber', flat=True).distinct()
            workstation_status = ReportQueue.objects.using('TEST').filter(reportstatus='test running').values_list('workstation','jobnumber','partnumber','activedate','operator','value','maxvalue','ping','pk').all()
            print('part_list=',part_list)
            x=1
            for station, jobs, parts, activedate, opera, value, maxvalue,ping,test_id in workstation_status:
                daywatch=today-activedate
                if not ping:
                    #print('no ping=')
                    ReportQueue.objects.using('TEST').filter(pk=test_id).update(reportstatus='for review')
                elif 'days' in str(daywatch):
                    #print('today-activedate',today-activedate)
                    ReportQueue.objects.using('TEST').filter(pk=test_id).update(reportstatus='for review')
                else:
                    #print('this_time=',this_time)
                    #print('ping=',ping)
                    t1=datetime.timedelta(hours=this_time.hour, minutes=this_time.minute, seconds=this_time.second)
                    t2=datetime.timedelta(hours=ping.hour, minutes=ping.minute, seconds=ping.second)
                    dt=t1-t2
                    #print('dt=',dt)
                    dt=dt.seconds/3600
                    #print('dt=',dt)
                    if int(dt)>2:
                        percent=100 * float(value)/float(maxvalue)
                        #print('percent=',percent)
                        if percent>=30:
                            ReportQueue.objects.using('TEST').filter(pk=test_id).update(reportstatus='report queue')
                        else:
                            ReportQueue.objects.using('TEST').filter(pk=test_id).update(reportstatus='job closed')
                
                gauge = pygal.SolidGauge(
                show_legend=False, half_pie=True, inner_radius=0.70,
                style=pygal.style.styles['default'](value_font_size=80,plot_background="gray"))
                efficiency = Effeciency.objects.using('TEST').filter(workstation=station).filter(jobnumber=jobs).filter(operator=opera).last()
                #print('efficiency=',efficiency)
                #print('************************ping=',ping)
                if efficiency:
                    comment = 'Workstation: ' + str(station) + '\nOperator: ' + str(opera) + '\nJob: ' + str(jobs) + '\nPart: ' + str(parts) + '\nTotal Parts: ' + str(efficiency.totaluuts) + '\nParts Complete: ' + str(efficiency.completeuuts) + '\nOperator Effeciency: ' + str(efficiency.effeciencystatus)
                else:
                    comment = 'Workstation: ' + str(station) + '\nOperator: ' + str(opera) + '\nJob: ' + str(jobs)
                percent_formatter = lambda x: '{:.10g}%'.format(x)
                dollar_formatter = lambda x: '{:.10g}$'.format(x)
                gauge.value_formatter = percent_formatter
                if value:
                    print('value234234=',value,' maxvalue=',maxvalue)
                    new_val = ((value/maxvalue) * 100)
                    gauge.add('', [{'value': int(new_val), 'max_value': 100}])
                    print('value234234=',value,' maxvalue=',maxvalue)
                if x == 1:
                    test_status1=gauge.render_data_uri()
                    test_comment1 = comment
                elif x == 2:
                    test_status2=gauge.render_data_uri()
                    test_comment2 = comment
                elif x == 3:
                    test_status3=gauge.render_data_uri()
                    test_comment3=comment
                elif x == 4:
                    test_status4=gauge.render_data_uri()
                    test_comment4 = comment
                elif x == 5:
                    test_status5=gauge.render_data_uri()
                    test_comment5 = comment
                elif x == 6:
                    test_status6=gauge.render_data_uri()
                    test_comment6 = comment
                elif x == 7:
                    test_status7=gauge.render_data_uri()
                    test_comment7 = comment
                elif x == 8:
                    test_status8=gauge.render_data_uri()
                    test_comment8 = comment
                elif x == 9:
                    test_status9=gauge.render_data_uri()
                    test_comment9 = comment
                elif x == 10:
                    test_status10=gauge.render_data_uri()
                    test_comment10 = comment
                x+=1
                #print('test_status1',test_status1)
                print('comment=',comment)
              
            
        except IOError as e:
            print ("Lists load Failure ", e)
            print('error = ',e)     
        return render (self.request,"excel/index.html",{'job_num':job_num,'part_num':part_num,'workstation':workstation,'operator':operator,'start_date':start_date,'end_date':end_date,'artwork_list':artwork_list,'artwork':artwork,
                                                        'job_list':job_list,'part_list':part_list,'workstation_list':workstation_list,'operator_list':operator_list,'spec1':spec1,'spec2':spec1,'spec3':spec3,'spectype':spectype,
                                                        'spec4':spec4,'spec5':spec5,'report_data':report_data,'test1_list':test1_list,'test2_list':test2_list,'test3_list':test3_list,'test4_list':test4_list,'test5_list':test5_list,
                                                        'stat1_min':stat1_min,'stat1_max':stat1_max,'stat1_avg':stat1_avg,'stat1_std':stat1_std,'stat2_min':stat2_min,'stat2_max':stat2_max,'stat2_avg':stat2_avg,'stat2_std':stat2_std,
                                                        'stat3_min':stat3_min,'stat3_max':stat3_max,'stat3_avg':stat3_avg,'stat3_std':stat3_std,'stat4_min':stat4_min,'stat4_max':stat4_max,'stat4_avg':stat4_avg,'stat4_std':stat4_std,
                                                        'stat5_min':stat3_min,'stat5_max':stat5_max,'stat5_avg':stat5_avg,'stat5_std':stat5_std,'analyze':analyze,'il_histo_data':il_histo_data,'rl_histo_data':rl_histo_data,
                                                        'iso_histo_data':iso_histo_data,'ab_histo_data':ab_histo_data,'pb_histo_data':pb_histo_data,'coup_histo_data':coup_histo_data,'iso_histo_data':iso_histo_data,
                                                        'passed1':passed1,'failed1':failed1,'failed_percent1':failed_percent1,'passed2':passed2,'failed2':failed2,'failed_percent2':failed_percent2,'passed3':passed3,'failed3':failed3,   
                                                        'cb_histo_data':cb_histo_data,'failed_percent3':failed_percent3,'passed4':passed4,'failed4':failed4,'failed_percent4':failed_percent4,'passed5':passed5,'failed5':failed5,
                                                        'failed_percent5':failed_percent5,'bad_data1':bad_data1,'bad_data2':bad_data2,'bad_data3':bad_data3,'bad_data4':bad_data4,'bad_data5':bad_data5,'il_histo_data2':il_histo_data2,
                                                        'chart1':chart1,'chart2':chart2,'chart3':chart3,'chart4':chart4,'chart5':chart5,'blank':blank,'total':total,'art_rev_list':art_rev_list,'blank':blank,'test_status1':test_status1,
                                                        'test_status2':test_status2,'test_status3':test_status3,'test_status4':test_status4,'test_status5':test_status5,'test_status6':test_status6,'test_status7':test_status7,
                                                        'test_status8':test_status8,'test_status9':test_status9,'test_status10':test_status10,'test_comment1':test_comment1,'test_comment2':test_comment2,'test_comment3':test_comment3,
                                                        'test_comment4':test_comment4,'test_comment5':test_comment5,'test_comment6':test_comment6,'test_comment7':test_comment7,'test_comment8':test_comment8,
                                                        'test_comment9':test_comment9,'test_comment10':test_comment10})
    def post(self, request, *args, **kwargs):
        operator = self.request.user
        form = 0
        try:
            status ='In Process'
            job_num=-1
            part_num=-1
            workstation=-1
            operator=-1
            start_date=-1
            search=-1
            end_date = -1
            spectype = -1
            spec1 = -1
            spec2 = -1
            spec3 = -1
            spec4 = -1
            spec5 = -1
            artwork = -1
            report_data = -1
            stat1_min = -1
            stat1_max = -1
            stat1_avg = -1
            stat1_std = -1
            stat2_min = -1
            stat2_max = -1
            stat2_avg = -1
            stat2_std = -1
            stat3_min = -1
            stat3_max = -1
            stat3_avg = -1
            stat3_std = -1
            stat4_min = -1
            stat4_max = -1
            stat4_avg = -1
            stat4_std = -1
            stat5_min = -1
            stat5_max = -1
            stat5_avg = -1
            stat5_std = -1
            il_histo_data = -1
            rl_histo_data = -1
            iso_histo_data = -1
            ab_histo_data = -1
            pb_histo_data = -1
            coup_histo_data = -1
            dir_histo_data = -1
            cb_histo_data = -1
            il_histo_data2 = -1
            rl_histo_data2 = -1
            iso_histo_data2 = -1
            ab_histo_data2 = -1
            pb_histo_data2 = -1
            coup_histo_data2 = -1
            dir_histo_data2 = -1
            cb_histo_data2 = -1
            chart1 = -1
            chart2 = -1
            chart3 = -1
            chart4 = -1
            chart5 = -1
            
            job_list = []
            part_list = []
            workstation_list = []
            operator_list = []
            stat_list = []
            spec_list = -1
            artwork_list = ['RawData 1',]
            art_rev_list = []
            test1_list = []
            test2_list = []
            test3_list = []
            test4_list = []
            test4_list = []
            test5_list = []
            
            bad1_list = []
            bad2_list = []
            bad3_list = []
            bad4_list = []
            bad4_list = []
            bad5_list = []
           
            blank = 0
            total = 0
            bad_data1 = 'X'
            bad1=0
            passed1 = 0
            failed1 = 0
            failed_percent1 = 0
            bad_data2 = 'X'
            bad2=0
            passed2 = 0
            failed2 = 0
            failed_percent2 = 0
            bad_data3 = 'X'
            bad3=0
            passed3 = 0
            failed3 = 0
            failed_percent3 = 0
            bad_data4 = 'X'
            bad4=0
            passed4 = 0
            failed4 = 0
            failed_percent4 = 0
            bad_data5 = 'X'
            bad5=0
            passed5 = 0
            failed5 = 0
            failed_percent5 = 0
            test_status1 = -1
            test_status2 = -1
            test_status3 = -1
            test_status4 = -1
            test_status5 = -1
            test_status6 = -1
            test_status7 = -1
            test_status8 = -1
            test_status9 = -1
            test_status10 = -1
            test_comment1 = -1
            test_comment2 = -1
            test_comment3 = -1
            test_comment4 = -1
            test_comment5 = -1
            test_comment6 = -1
            test_comment7 = -1
            test_comment8 = -1
            test_comment9 = -1
            test_comment10 = -1
            spectype=-1
                
            
            #  Equations to get today - days
            #~~~~~~~~~~~~~ Time ~~~~~~~~~~~~~~~~~
            days=30 # start_date is today - days 
            time_code = TimeCode(days)
            friday = time_code.friday()
            #print('friday=',friday)
            today = datetime.datetime.today()
            today = make_aware(today)
            #print('today =', today)
            #start_date  = time_code.today_minus() # end_date is today - days 
            #start_date = make_aware(start_date)
            #end_date = today
            #print('start_date =',start_date)
            #print('end_date =',end_date)
            year = time_code.this_year()
            month_num = time_code.this_month()
            month_string = time_code.month_string()
            day = time_code.this_day()
            hour = time_code.this_hour()
            minute = time_code.this_minute()
            sec = time_code.this_sec()
            #print('Today=',day,'/',month_num,'/',year,'/ ',hour,':',minute,':',sec)
            #print('Month=',month_string)
            #~~~~~~~~~~~~~ Time ~~~~~~~~~~~~~~~~~
            
            #~~~~~~~~~~Get Post Values~~~~~~~~~~~~~~~
            job_num = request.POST.get('_job', -1)
            if job_num=='None' or job_num=='' or job_num==None or job_num=='All Job Numbers':
                job_num=-1
            #print('job_num=',job_num)
            part_num = request.POST.get('_part', -1)
            if part_num=='None' or part_num=='' or part_num==None or part_num=='All Part Numbers':
                part_num=-1
            #print('part_num=',part_num)
            workstation = request.POST.get('_workstation', -1)
            print('workstation=',workstation)
            if workstation=='None' or workstation=='' or workstation==None or workstation=='All Workstations':
                workstation=-1
            operator = request.POST.get('_operator', -1)
            if operator=='None' or operator=='' or operator==None or operator=='All Operators':
                operator=-1
            start_date = request.POST.get('_start_date', -1)
            end_date = request.POST.get('_end_date', -1)
            report = request.POST.get('_report', -1)
            #print('report123=',report)
            analyze = request.POST.get('_analyze', -1)
            trace = request.POST.get('_trace', -1)
            artwork = request.POST.get('_art', -1)
            if artwork=='None' or artwork=='' or artwork==None or artwork=='All Artworks':
                artwork=-1
            
            if analyze != -1 :
                analyze = 1
            #print('analyze=',analyze)
            spec1 = request.POST.get('_spec1', -1)
            spec2 = request.POST.get('_spec2', -1)
            spec3 = request.POST.get('_spec3', -1)
            spec4 = request.POST.get('_spec4', -1)
            spec5 = request.POST.get('_spec5', -1)
            print('spec1=',spec1,'spec2=',spec2,'spec3=',spec3,'spec4=',spec4,'spec5=',spec5)
            if spec1!=-1:
                spec1 = float(spec1)
                spec2 = float(spec2)
                spec3 = float(spec3)
                spec4 = float(spec4)
                spec5 = float(spec5)
            
            #~~~~~~~~~~Get Post Values~~~~~~~~~~~~~~~
            #https://openpyxl.readthedocs.io/en/stable/
            #https://www.softwaretestinghelp.com/python-openpyxl-tutorial/
            print('^^^^^^^^^^^^^Look here^^^^^^^^^^^^^^^^^^^^^')
            print('job_num=',job_num)
            print('part_num=',part_num)
            print('workstation=',workstation)
            print('operator=',operator)
            print('artwork=',artwork)
            print('report=',report)
            print('^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^')
            if report!=-1:
                print('excel report start')
                reporting = ExcelReports(job_num,operator,workstation)
                reporting.test_data()
                print('excel report end')   
                return redirect('excel:reports')
            
            job_list = Testdata.objects.using('TEST').order_by('-jobnumber').values_list('jobnumber', flat=True).distinct()
            part_list = Testdata.objects.using('TEST').order_by('-partnumber').values_list('partnumber', flat=True).distinct() 
            operator_list = Effeciency.objects.using('TEST').order_by('operator').values_list('operator', flat=True).distinct()
            workstation_list = Workstation.objects.using('TEST').order_by('computername').values_list('computername', flat=True).distinct()
            
            got_enough=False
            if job_num !=-1:#Job
                part = Testdata.objects.using('TEST').filter(jobnumber=job_num).last()
                job_num = part.jobnumber 
                operator = part.operator
                if part:
                    part_num = part.partnumber
                workstation= part.workstation
                report_data = Testdata.objects.using('TEST').filter(jobnumber=job_num).all()
                spec_data = Specifications.objects.using('TEST').filter(jobnumber=job_num).first()
                if spec_data:
                    got_enough=True
                
                print('spec_data',spec_data)
                #print('we are here',report_data)
            elif part_num!=-1 and artwork!=-1 and workstation!=-1  and operator!=-1 : #part_number, Artwork, Workstation, Operator
                report_data = Testdata.objects.using('TEST').filter(partnumber=part_num).filter(artwork_rev=artwork).filter(workstation=workstation).filter(operator=operator).all()
                spec_data = Specifications.objects.using('TEST').filter(partnumber=part_num).first()
                if spec_data:
                    got_enough=True
                print('in part1')
            elif part_num!=-1 and artwork!=-1 and workstation!=-1  and operator==-1 : #part_number, Artwork, Workstation
                report_data = Testdata.objects.using('TEST').filter(partnumber=part_num).filter(artwork_rev=artwork).filter(workstation=workstation).all()
                spec_data = Specifications.objects.using('TEST').filter(partnumber=part_num).first()
                if spec_data:
                    got_enough=True
                print('in part2')
            elif part_num!=-1 and artwork!=-1 and workstation==-1  and operator==-1 : #part_number, Artwork
                report_data = Testdata.objects.using('TEST').filter(partnumber=part_num).filter(artwork_rev=artwork).all()
                spec_data = Specifications.objects.using('TEST').filter(partnumber=part_num).first()
                if spec_data:
                    got_enough=True
                print('in part3')
            elif part_num!=-1 and artwork==-1 and workstation==-1  and operator==-1 : #part_number
                report_data = Testdata.objects.using('TEST').filter(partnumber=part_num).all()
                spec_data = Specifications.objects.using('TEST').filter(partnumber=part_num).first()
                if spec_data:
                    got_enough=True
                print('spec_data=',spec_data)
                print('in part4')
            elif part_num==-1 and artwork!=-1 and workstation!=-1  and operator!=-1 : #Artwork, Workstation, Operator
                report_data = Testdata.objects.using('TEST').filter(artwork_rev=artwork).filter(workstation=workstation).filter(operator=operator).all()
                if report_data:
                    part_num=report_data[0].partnumber
                    if part_num:
                        spec_data = Specifications.objects.using('TEST').filter(partnumber=part_num).first()
                        if spec_data:
                            got_enough=True
                print('in part5')
            elif part_num==-1 and artwork!=-1 and workstation!=-1  and operator==-1 : #Artwork, Workstation
                report_data = Testdata.objects.using('TEST').filter(artwork_rev=artwork).filter(workstation=workstation).all() 
                if report_data:
                    part_num=report_data[0].partnumber
                    if part_num:
                        spec_data = Specifications.objects.using('TEST').filter(partnumber=part_num).first()
                        if spec_data:
                            got_enough=True                
                print('in part6')                
            elif part_num==-1 and artwork!=-1 and workstation==-1  and operator==-1 : #Artwork
                report_data = Testdata.objects.using('TEST').filter(artwork_rev=artwork).all()   
                if report_data:
                    part_num=report_data[0].partnumber
                    if part_num:
                        spec_data = Specifications.objects.using('TEST').filter(partnumber=part_num).first()
                        if spec_data:
                            got_enough=True    
                print('in part7')
            elif part_num==-1 and artwork==-1 and workstation!=-1  and operator!=-1 : #Workstation, Operator
                report_data = Testdata.objects.using('TEST').filter(workstation=workstation).filter(operator=operator).all()   
                if report_data:
                    part_num=report_data[0].partnumber
                    if part_num:
                        spec_data = Specifications.objects.using('TEST').filter(partnumber=part_num).first()
                        if spec_data:
                            got_enough=True    
                print('in part8')
            elif part_num==-1 and artwork==-1 and workstation!=-1  and operator==-1 : #Workstation
                report_data = Testdata.objects.using('TEST').filter(workstation=workstation).all() 
                if report_data:
                    part_num=report_data[0].partnumber
                    if part_num:
                        spec_data = Specifications.objects.using('TEST').filter(partnumber=part_num).first()
                        if spec_data:
                            got_enough=True    
                print('in part9')
            elif part_num==-1 and artwork==-1 and workstation==-1  and operator!=-1 : #Operator
                report_data = Testdata.objects.using('TEST').filter(operator=operator).all() 
                if report_data:
                    part_num=report_data[0].partnumber
                    if part_num:
                        spec_data = Specifications.objects.using('TEST').filter(partnumber=part_num).first()
                        if spec_data:
                            got_enough=True    
                print('in part10')
            print('report_data',report_data)
            if got_enough and spec_data and report_data:
                #filter blanks
                temp_list = []
                for artwork_rev in artwork_list:
                    if not artwork_rev == '':
                        temp_list.append(artwork_rev)
                artwork_list = temp_list
                #print('artwork_list=',artwork_list)
                #print('job_num=',job_num)
                print('spec_data=',spec_data)
                print('spec_data.vswr=',spec_data.vswr)
                if spec_data.vswr:
                    conversions = Conversions(spec_data.vswr,'')
                    spec_rl = round(conversions.vswr_to_rl(),3)
                else:
                    spec_rl = 0
                print('spec_rl=',spec_rl)
                spectype=spec_data.spectype
                try:
                    if '90 DEGREE COUPLER' in spec_data.spectype or 'BALUN' in spec_data.spectype:
                        if spec_data.insertionloss:
                            spec1 = round(spec_data.insertionloss,3)
                        if spec_rl:
                            spec2 = spec_rl
                        if spec_data.isolation:
                            spec3 = round(spec_data.isolation,3)
                        if spec_data.amplitudebalance:
                            spec4 = round(spec_data.amplitudebalance,3)
                        if spec_data.phasebalance:
                            spec5 = round(spec_data.phasebalance,3)
                    elif 'DIRECTIONAL COUPLER' in spec_data.spectype: 
                        if spec_data.insertionloss:
                            spec1 = round(spec_data.insertionloss,3)
                        if spec_rl:
                            spec2 = spec_rl
                        if spec_data.coupling:
                            spec3 = round(spec_data.coupling,3)
                        if spec_data.directivity:
                            spec4 = round(spec_data.directivity,3)
                        if spec_data.coupledflatness:
                            spec5 = round(spec_data.coupledflatness,3)
                except ValueError as e:
                    print('error = ',e) 
                   
                spectype=spec_data.spectype
                print('spectype=',spec_data.spectype)
                total=0
                temp_list = []
                
                print('report_data=',report_data)
                for data in report_data:
                   good_data=True
                   print('good_data1=',good_data)
                   #~~~~~~~~~~~~~~~Check for good data~~~~~~~~~~~~~~~~~
                   print('IL&RL ',data.insertionloss,data.insertionloss)
                   if not data.insertionloss and not data.returnloss:
                        good_data=False
                        print('IL no good')
                   if '90 DEGREE COUPLER' in spec_data.spectype or 'BALUN' in spec_data.spectype:
                        #print('ISo&AM&PB ',data.isolation,data.phasebalance)
                        if not data.isolation and not data.phasebalance: 
                            good_data=False
                            print('no good ISO or PB')
                        if spec_data.ab_exp_tf :
                            if not data.amplitudebalance1:
                                good_data=False
                                print('no good AB')
                        else:
                            if not data.amplitudebalance:
                                good_data=False
                                print('no good AB2')
                   else:
                        #print('coup&dir&cf ',data.coupling,data.directivity,data.coupledflatness)
                        if not data.coupling and not data.directivity and not data.coupledflatness: 
                            good_data=False
                            print('no good CPL DIR CF')
                   print('good_data2=',good_data)
                   #~~~~~~~~~~~~~~~Check for good data~~~~~~~~~~~~~~~~~
                   if good_data:
                        go = True
                        if data.insertionloss:  #does data.insertionloss: have any data? 
                            if data.insertionloss > spec1 * 3:
                                bad1_list.append(data.insertionloss)
                                #print('bad1=',data.insertionloss,'spec=',spec1 * 3)
                        if data.returnloss:  #does data.returnloss: have any data? 
                            if data.returnloss < spec2 * 3:
                                bad2_list.append(data.returnloss)
                                #print('bad2=',data.returnloss,'spec=',spec2 * 3)
                        
                        if ('90 DEGREE COUPLER' in spec_data.spectype or 'BALUN' in spec_data.spectype) and data.isolation:  #does data.isolation: have any data?      
                            if abs(data.isolation) > spec3 * 3:  
                                bad3_list.append(data.isolation)
                                #print('bad3=',abs(data.isolation),'spec=',spec3 * 3)
                        
                        if ('90 DEGREE COUPLER' in spec_data.spectype or 'BALUN' in spec_data.spectype) and data.amplitudebalance:  #does data.amplitudebalance have any data?      
                            if abs(data.amplitudebalance) > spec4 * 3:  
                                bad4_list.append(data.amplitudebalance)
                                #print('bad4=',abs(data.amplitudebalance),'spec=',spec4 * 3)
                       
                        if ('90 DEGREE COUPLER' in spec_data.spectype or 'BALUN' in spec_data.spectype) and data.phasebalance:  #does data.coupling have any data?  
                            if abs(data.phasebalance) > spec5 * 3:  
                                bad5_list.append(data.phasebalance)
                                #print('bad5=',abs(data.phasebalance),'spec=',spec5 * 3)
                            
                        if 'DIRECTIONAL COUPLER' in spec_data.spectype and data.coupling: #does data.coupling have any data?
                            if abs(data.coupling) > spec3 * 3:  
                                bad3_list.append(data.coupling)
                                
                        if 'DIRECTIONAL COUPLER' in spec_data.spectype and data.directivity:#does data.directivity have any data?
                            if abs(data.directivity) > spec4 * 3:  
                                bad4_list.append(data.directivity)
                        if 'DIRECTIONAL COUPLER' in spec_data.spectype  and data.coupledflatness: #does data.coupledflatness have any data?
                            if 'DIRECTIONAL COUPLER' in spec_data.spectype  and abs(data.coupledflatness) > spec5 * 3:  
                                bad5_list.append(data.coupledflatness)
                             
                        
                        temp_list.append(data)
                        art_rev_list.append(data.artwork_rev)
                        test1_list.append(data.insertionloss)
                        test2_list.append(data.returnloss)
                        if '90 DEGREE COUPLER' in spec_data.spectype or 'BALUN' in spec_data.spectype:
                            test3_list.append(data.isolation)
                            test4_list.append(data.amplitudebalance)
                            test5_list.append(data.phasebalance)
                        else:
                            test3_list.append(data.coupling)
                            test4_list.append(data.directivity)
                            test5_list.append(data.coupledflatness)
                   else:
                        blank+=1
                   total+=1
                        
                if len(bad1_list)>=1:
                    bad1+=1
                    if len(bad1_list)==1:
                        bad_data1 = round(bad1_list[0],2)
                    else:
                        bad_data1 = round(statistics.mean(bad1_list),2)
                if len(bad2_list)>=1:
                    bad2+=1
                    if len(bad2_list)==1:
                        bad_data2 = round(bad1_list[0],2)
                    else:
                        bad_data2 = round(statistics.mean(bad2_list),2)
                if len(bad3_list)>=1:
                    bad3+=1
                    if len(bad3_list)==1:
                        bad_data3 = round(bad3_list[0],2)
                    else:
                        bad_data3 = round(statistics.mean(bad3_list),2)
                if len(bad4_list)>=1:
                    bad4+=1
                    if len(bad4_list)==1:
                        bad_data4 = round(bad4_list[0],2)
                    else:
                        bad_data4 = round(statistics.mean(bad4_list),2)
                if len(bad5_list)>=1:
                    bad5+=1
                    if len(bad5_list)==1:
                        bad_data5 = round(bad5_list[0],2)
                    else:
                        bad_data5 = round(statistics.mean(bad5_list),2)
                
                report_data = temp_list 
                
                if len(test1_list) > 1:# must have at least two tests
                    test_list = [test1_list,test2_list,test3_list,test4_list,test5_list]
                    spec_list = [spec1,spec2,spec3,spec4,spec5]
                    #print('test_list =',test_list)
                    histo_data = Histogram_data(test_list,spec_list,'test1') 
                    il_histo_data = histo_data.Hist_data()
                    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~statistics~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                    try:
                        stat_data = Statistics(test1_list,test2_list,test3_list,test4_list,test5_list) 
                        stat_list = stat_data.get_stats()
                        #print('stat_list=',stat_list)
                        stat1_min = stat_list[0][0]
                        #print('stat1_min=',stat1_min)
                        stat1_max = stat_list[0][1]
                        #print('stat1_max=',stat1_max)
                        stat1_avg = stat_list[0][2]
                        #print('stat1_avg=',stat1_avg)
                        stat1_std = stat_list[0][3]
                        #print('stat1_std=',stat1_std)
                        stat2_min = stat_list[1][0]
                        stat2_max = stat_list[1][1]
                        stat2_avg = stat_list[1][2]
                        stat2_std = stat_list[1][3]
                        stat3_min = stat_list[2][0]
                        stat3_max = stat_list[2][1]
                        stat3_avg = stat_list[2][2]
                        stat3_std = stat_list[2][3]
                        stat4_min = stat_list[3][0]
                        stat4_max = stat_list[3][1]
                        stat4_avg = stat_list[3][2]
                        stat4_std = stat_list[3][3]
                        stat5_min = stat_list[4][0]
                        stat5_max = stat_list[4][1]
                        stat5_avg = stat_list[4][2]
                        stat5_std = stat_list[4][3]
                        for x in range(len(test1_list)):
                            if test1_list[x] <= spec1:
                                passed1 +=1
                            else:
                                failed1 +=1
                                
                            if test2_list[x] <= spec2:
                                passed2 +=1
                            else:
                                failed2 +=1
                            if test3_list[x] <= spec3:
                                passed3 +=1
                            else:
                                failed3 +=1
                            if test4_list[x] <= spec4:
                                passed4 +=1
                            else:
                                failed4 +=1
                            if test5_list[x] <= spec5:
                                passed5 +=1
                            else:
                                failed5 +=1
                                
                        if passed1==0:
                            failed_percent1 = '100%'
                        elif failed1==0:
                            failed_percent1 = '0%'
                        else:    
                            failed_percent1 = str(round((failed1/passed1)* 100,3)) + '%'
                        
                        if passed2==0:
                            failed_percent2 = '100%'
                        elif failed2==0:
                            failed_percent2 = '0%'
                        else:    
                            failed_percent2 = str(round((failed2/passed2)* 100,3)) + '%'
                        
                        if passed3==0:
                            failed_percent3 = '100%'
                        elif failed3==0:
                            failed_percent3 = '0%'
                        else:    
                            failed_percent3 = str(round((failed3/passed3)* 100,3)) + '%'
                        
                        if passed4==0:
                            failed_percent4 = '100%'
                        elif failed4==0:
                            failed_percent4 = '0%'
                        else:    
                            failed_percent4 = str(round((failed4/passed4)* 100,3)) + '%'
                        
                        if passed5==0:
                            failed_percent5 = '100%'
                        elif failed5==0:
                            failed_percent5 = '0%'
                        else:    
                            failed_percent5 = str(round((failed5/passed5)* 100,2)) + '%'
                        
                        mean = statistics.mean(test3_list)
                        sd = statistics.stdev(test3_list)
                        #print('mean=',mean)
                        #print('sd=',sd)   
                    except BaseException as err:
                        print(f"Unexpected {err=}, {type(err)=}")
                    
                    histo_data = Histogram_data(test_list,spec_list,'test1')
                    il_histo = histo_data.Hist_data()
                    
                    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~IL x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                    x_range_list = []
                    x_range = X_Range(il_histo,spec1, stat1_min,stat1_max,stat1_avg)                
                    x_range_list=x_range.list()
                    #print('x_range_list=',len(x_range_list))
                    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~IL  x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                    
                    #~~~~~~~~~~~~~IL standard deviation gassian distribution line~~~~~~~~~~~~~~
                    sd_list = []
                    sd=SDEV_Dist(spec1, test1_list,stat1_std,stat1_min,stat1_max,stat1_avg) 
                    sd_list =sd.gauss()
                    #print('sd_list=',sd_list)
                    min_max = sd.gauss_min_max()
                    #~~~~~~~~~~~~~IL standard deviation gassian distribution line~~~~~~~~~~~~~~
                    
                    #~~~~~~~~~~ IL Standard deviation Chart~~~~~~~~~~~~~~~~~
                    #print('chart_data=',chart_data)
                    custom_style = Style(colors=('#000000','#FF0000'),title_font_size=39, label_font_size=15,
                                        background='transparent', plot_background='transparent')
                    xy_chart = pygal.XY(style=custom_style,show_dots=False,show_y_labels=False)
                    xy_chart.title = 'IL Histogram'
                    xy_chart.add('IL Histogram', sd_list)
                    #print('sd_list=',sd_list)
                    #print('spec1=',spec1)
                    # add  spec line
                    xy_chart.add('spec max', [(spec1, min_max[0]),(spec1, min_max[1])])
                    
                    #~~~~~~~~~~ IL Standard deviation Chart~~~~~~~~~~~~~~~~~
                    
                    #~~~~~~~~~~~~~~~~ IL Histogram Chart~~~~~~~~~~~~~~~~~~~~
                    try:
                        il_histo_data2 = xy_chart.render_data_uri()
                        #print('rl_histo=',rl_histo)
                        custom_style = Style(colors=('#991593','#201599'),title_font_size=39, label_font_size=17)
                        hist = pygal.Histogram(fill=True,style=custom_style, human_readable=True,show_x_labels=False)
                        hist.x_labels = x_range_list
                        hist.add('IL Histogram', il_histo)
                        hist.title = 'IL Histogram' 
                        il_histo_data = hist.render_data_uri()
                        #print('il_histo_data=',il_histo_data)
                    except ValueError as e:
                             il_histo_data =  -1
                    #~~~~~~~~~~~~~~~~ IL Histogram Chart~~~~~~~~~~~~~~~~~~~~
                    
                    #~~~~~~~~~~~~~~~~~ IL Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                    histo_data = Histogram_data(test_list,spec_list,'test1')
                    il_histo = histo_data.Hist_data()
                    charts = XY_Chart(test_list,spec_list,'test1') 
                    chart_data = charts.Chart_data()
                    #print('chart_data=',chart_data)
                    custom_style = Style(colors=('#991593','#201599'),title_font_size=39, label_font_size=15)
                    xy_chart = pygal.XY(style=custom_style, human_readable=True)
                    xy_chart.title = 'IL XY Plot'
                    xy_chart.add('Insertion Loss', chart_data)
                    xy_chart.add('spec max', [(0, spec1), (len(chart_data), spec1)])
                    chart1 = xy_chart.render_data_uri()
                    #~~~~~~~~~~~~~~~~~ IL Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                    
                    histo_data = Histogram_data(test_list,spec_list,'test2')
                    rl_histo = histo_data.Hist_data()
                    
                    #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
                    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~RL x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                    x_range_list = []
                    x_range = X_Range(rl_histo,spec2, stat2_min,stat2_max,stat2_avg)                
                    x_range_list=x_range.list()
                    print('x_range_list=',len(x_range_list))
                    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~RL  x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                    
                    #~~~~~~~~~~~~~RL standard deviation gassian distribution line~~~~~~~~~~~~~~
                    sd_list = []
                    sd=SDEV_Dist(spec2, test2_list,stat2_std,stat2_min,stat2_max,stat2_avg) 
                    sd_list =sd.gauss()
                    #print('sd_list=',sd_list)
                    min_max = sd.gauss_min_max()
                    #~~~~~~~~~~~~~RL standard deviation gassian distribution line~~~~~~~~~~~~~~
                    
                    #~~~~~~~~~~ RL Standard deviation Chart~~~~~~~~~~~~~~~~~
                    #print('chart_data=',chart_data)
                    custom_style = Style(colors=('#000000','#FF0000'),title_font_size=39, label_font_size=15,
                                        background='transparent', plot_background='transparent')
                    xy_chart = pygal.XY(style=custom_style,show_dots=False,show_y_labels=False)
                    xy_chart.title = 'RL Histogram'
                    xy_chart.add('RL Histogram', sd_list)
                    #print('sd_list=',sd_list)
                    #print('spec2=',spec2)
                    # add  spec line
                    xy_chart.add('spec max', [(spec2, min_max[0]),(spec2, min_max[1])])
                    #~~~~~~~~~~ RL Standard deviation Chart~~~~~~~~~~~~~~~~~
                    
                    #~~~~~~~~~~~~~~~~ RL Histogram Chart~~~~~~~~~~~~~~~~~~~~
                    try:
                        rl_histo_data2 = xy_chart.render_data_uri()
                        #print('rl_histo=',rl_histo)
                        custom_style = Style(colors=('#47ff7b','#201599'),title_font_size=39, label_font_size=17)
                      
                      
                        hist = pygal.Histogram(fill=True,style=custom_style, human_readable=True,show_x_labels=False,print_values=True)
                        hist.x_labels = x_range_list
                        hist.add('RL Histogram', rl_histo)
                        hist.title = 'RL Histogram' 
                        rl_histo_data = hist.render_data_uri()
                        #print('rl_histo_data=',il_histo_data)
                    except ValueError as e:
                        rl_histo_data =  -1
                    #~~~~~~~~~~~~~~~~ RL Histogram Chart~~~~~~~~~~~~~~~~~~~~
                    
                    #~~~~~~~~~~~~~~~~~ RL Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                    histo_data = Histogram_data(test_list,spec_list,'test2')
                    rl_histo = histo_data.Hist_data()
                    charts = XY_Chart(test_list,spec_list,'test2') 
                    chart_data = charts.Chart_data()
                    #print('chart_data=',chart_data)
                    custom_style = Style(colors=('#47ff7b','#201599'),title_font_size=39, label_font_size=15)
                    xy_chart = pygal.XY(style=custom_style, human_readable=True)
                    xy_chart.title = 'RL XY Plot'
                    xy_chart.add('Return Losss', chart_data)
                    xy_chart.add('spec max', [(0, spec2), (len(chart_data), spec2)])
                    chart2 = xy_chart.render_data_uri()
                    #~~~~~~~~~~~~~~~~~ RL Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                     
                    #print('rl_histo_data=',rl_histo_data)
                    if '90 DEGREE COUPLER' in spec_data.spectype or 'BALUN' in spec_data.spectype:
                        histo_data = Histogram_data(test_list,spec_list,'test3')
                        iso_histo = histo_data.Hist_data()
                        
                        #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
                        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ISO x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        x_range_list = []
                        x_range = X_Range(iso_histo,0-spec3, stat3_min,stat3_max,stat3_avg)                
                        x_range_list=x_range.list()
                        #print('x_range_list=',len(x_range_list))
                        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ISO  x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~ISO standard deviation gassian distribution line~~~~~~~~~~~~~~
                        sd_list = []
                        sd=SDEV_Dist(spec3, test3_list,stat3_std,stat3_min,stat3_max,stat3_avg) 
                        sd_list =sd.gauss()
                        #print('sd_list=',sd_list)
                        min_max = sd.gauss_min_max()
                        #~~~~~~~~~~~~~ISO standard deviation gassian distribution line~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~ ISO Standard deviation Chart~~~~~~~~~~~~~~~~~
                        #print('chart_data=',chart_data)
                        custom_style = Style(colors=('#000000','#FF0000'),title_font_size=39, label_font_size=15,
                                            background='transparent', plot_background='transparent')
                        xy_chart = pygal.XY(style=custom_style,show_dots=False,show_y_labels=False)
                        xy_chart.title = 'ISO Histogram'
                        xy_chart.add('ISO Histogram', sd_list)
                        #print('sd_list=',sd_list)
                        #print('spec3=',spec3)
                        # add  spec line
                        xy_chart.add('spec max', [(0-spec3, min_max[0]),(0-spec3, min_max[1])])
                        
                        #~~~~~~~~~~ ISO Standard deviation Chart~~~~~~~~~~~~~~~~~
                        try:
                            iso_histo_data2 = xy_chart.render_data_uri()
                            #~~~~~~~~~~~~~~~~ ISO Histogram Chart~~~~~~~~~~~~~~~~~~~~
                            #print('rl_histo=',rl_histo)
                            custom_style = Style(colors=('#ffd138','#201599'),title_font_size=39, label_font_size=17)
                            hist = pygal.Histogram(fill=True,style=custom_style, human_readable=True,show_x_labels=False,print_values=True)
                            hist.x_labels = x_range_list
                            hist.add('ISO Histogram', iso_histo)
                            hist.title = 'ISO Histogram' 
                            iso_histo_data = hist.render_data_uri()
                            #print('iso_histo_data=',iso_histo_data)
                            #~~~~~~~~~~~~~~~~ ISO Histogram Chart~~~~~~~~~~~~~~~~~~~~
                        except ValueError as e:
                             iso_histo_data =  -1
                        #~~~~~~~~~~~~~~~~~ ISO Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                        histo_data = Histogram_data(test_list,spec_list,'test3')
                        iso_histo = histo_data.Hist_data()
                        charts = XY_Chart(test_list,spec_list,'test3') 
                        chart_data = charts.Chart_data()
                        #print('chart_data=',chart_data)
                        custom_style = Style(colors=('#ffd138','#201599'),title_font_size=39, label_font_size=15)
                        xy_chart = pygal.XY(style=custom_style, human_readable=True)
                        xy_chart.title = 'ISO XY Plot'
                        xy_chart.add('Isolation', chart_data)
                        print('0-spec3=',0-spec3)
                        xy_chart.add('spec max', [(0, -16), (len(chart_data), -16)])
                        chart3 = xy_chart.render_data_uri()
                        #~~~~~~~~~~~~~~~~~ ISO Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                        
                        histo_data = Histogram_data(test_list,spec_list,'test4')
                        ab_histo = histo_data.Hist_data()
                        
                        #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
                        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~AB x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        x_range_list = []
                        x_range = X_Range(ab_histo,spec4, stat4_min,stat4_max,stat4_avg)                
                        x_range_list=x_range.list()
                        #print('x_range_list=',len(x_range_list))
                        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~AB  x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~AB standard deviation gassian distribution line~~~~~~~~~~~~~~
                        sd_list = []
                        sd=SDEV_Dist(spec4, test4_list,stat4_std,stat4_min,stat4_max,stat4_avg) 
                        sd_list =sd.gauss()
                        #print('sd_list=',sd_list)
                        min_max = sd.gauss_min_max()
                        #~~~~~~~~~~~~~AB standard deviation gassian distribution line~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~ AB Standard deviation Chart~~~~~~~~~~~~~~~~~
                        #print('chart_data=',chart_data)
                        custom_style = Style(colors=('#000000','#FF0000'),title_font_size=39, label_font_size=15,
                                            background='transparent', plot_background='transparent')
                        xy_chart = pygal.XY(style=custom_style,show_dots=False,show_y_labels=False)
                        xy_chart.title = 'AB Histogram'
                        xy_chart.add('AB Histogram', sd_list)
                        #print('sd_list=',sd_list)
                        #print('spec4=',spec4)
                        # add  spec line
                        xy_chart.add('spec min', [(spec4, 0-min_max[0]),(spec4, 0-min_max[1])])
                        xy_chart.add('spec max', [(spec4, min_max[0]),(spec4, min_max[1])])
                        #~~~~~~~~~~ AB Standard deviation Chart~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~~~~ AB Histogram Chart~~~~~~~~~~~~~~~~~~~~
                        try:
                            ab_histo_data2 = xy_chart.render_data_uri()
                            #print('rl_histo=',rl_histo)
                            custom_style = Style(colors=('#130fff','#201599'),title_font_size=39, label_font_size=17)
                            hist = pygal.Histogram(fill=True,style=custom_style, human_readable=True,show_x_labels=False,print_values=True)
                            hist.x_labels = x_range_list
                            hist.add('AB Histogram', ab_histo)
                            hist.title = 'AB Histogram' 
                            ab_histo_data = hist.render_data_uri()
                            #print('ab_histo_data=',ab_histo_data)
                        except ValueError as e:
                            ab_histo_data =  -1
                        #~~~~~~~~~~~~~~~~ AB Histogram Chart~~~~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~~~~~ AB Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                        histo_data = Histogram_data(test_list,spec_list,'test4')
                        ab_histo = histo_data.Hist_data()
                        charts = XY_Chart(test_list,spec_list,'test4') 
                        chart_data = charts.Chart_data()
                        #print('chart_data=',chart_data)
                        custom_style = Style(colors=('#130fff','#201599'),title_font_size=39, label_font_size=15)
                        xy_chart = pygal.XY(style=custom_style, human_readable=True)
                        xy_chart.title = 'AB XY Plot'
                        xy_chart.add('Amplitude Balance', chart_data)
                        xy_chart.add('spec min', [(0, 0-spec4), (len(chart_data), 0-spec4)])
                        xy_chart.add('spec max', [(0, spec4), (len(chart_data), spec4)])
                        chart4 = xy_chart.render_data_uri()
                        #~~~~~~~~~~~~~~~~~ AB Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                       
                        histo_data = Histogram_data(test_list,spec_list,'test5')
                        pb_histo = histo_data.Hist_data()
                        
                        #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
                        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~PB x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        x_range_list = []
                        x_range = X_Range(pb_histo,spec5, stat5_min,stat5_max,stat5_avg)                
                        x_range_list=x_range.list()
                        #print('x_range_list=',len(x_range_list))
                        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~PB  x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~PB standard deviation gassian distribution line~~~~~~~~~~~~~~
                        sd_list = []
                        sd=SDEV_Dist(spec5, test5_list,stat5_std,stat5_min,stat5_max,stat5_avg) 
                        sd_list =sd.gauss()
                        #print('sd_list=',sd_list)
                        min_max = sd.gauss_min_max()
                        #~~~~~~~~~~~~~PB standard deviation gassian distribution line~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~ PB Standard deviation Chart~~~~~~~~~~~~~~~~~
                        #print('chart_data=',chart_data)
                        custom_style = Style(colors=('#000000','#FF0000'),title_font_size=39, label_font_size=15,
                                            background='transparent', plot_background='transparent')
                        xy_chart = pygal.XY(style=custom_style,show_dots=False,show_y_labels=False)
                        xy_chart.title = 'PB Histogram'
                        xy_chart.add('PB Histogram', sd_list)
                        #print('sd_list=',sd_list)
                        #print('spec5=',spec5)
                        # add  spec line
                        xy_chart.add('spec min', [(spec5, 0-min_max[0]),(spec5, 0-min_max[1])])
                        xy_chart.add('spec max', [(spec5, min_max[0]),(spec5, min_max[1])])
                        #~~~~~~~~~~ PB Standard deviation Chart~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~~~~ PB Histogram Chart~~~~~~~~~~~~~~~~~~~~
                        try:
                            pb_histo_data2 = xy_chart.render_data_uri()
                            #print('rl_histo=',rl_histo)
                            custom_style = Style(colors=('#130fff','#201599'),title_font_size=39, label_font_size=17,
                                            value_font_family='googlefont:Raleway',value_font_size=30,value_colors=('white',))
                            hist = pygal.Histogram(fill=True,style=custom_style, human_readable=True,show_x_labels=False,print_values=True)
                            hist.x_labels = x_range_list
                            hist.add('PB Histogram', pb_histo)
                            hist.title = 'PB Histogram' 
                            pb_histo_data = hist.render_data_uri()
                            #print('pb_histo_data=',pb_histo_data)
                        except ValueError as e:
                            pb_histo_data =  -1
                        #~~~~~~~~~~~~~~~~ PB Histogram Chart~~~~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~~~~~ PB Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                        histo_data = Histogram_data(test_list,spec_list,'test5')
                        pb_histo = histo_data.Hist_data()
                        charts = XY_Chart(test_list,spec_list,'test5') 
                        chart_data = charts.Chart_data()
                        #print('chart_data=',chart_data)
                        custom_style = Style(colors=('#130fff','#201599'),title_font_size=39, label_font_size=15)
                        xy_chart = pygal.XY(style=custom_style, human_readable=True)
                        xy_chart.title = 'PB XY Plot'
                        xy_chart.add('Amplitude Balance', chart_data)
                        xy_chart.add('spec min', [(0, 0-spec5), (len(chart_data), 0-spec5)])
                        xy_chart.add('spec max', [(0, spec5), (len(chart_data), spec5)])
                        chart5 = xy_chart.render_data_uri()
                        #~~~~~~~~~~~~~~~~~ PB Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                    else:
                        histo_data = Histogram_data(test_list,spec_list,'test3')
                        coup_histo = histo_data.Hist_data()
                        
                        #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
                        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~COUP x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        x_range_list = []
                        x_range = X_Range(coup_histo,spec3, stat3_min,stat3_max,stat3_avg)                
                        x_range_list=x_range.list()
                        #print('x_range_list=',len(x_range_list))
                        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~COUP  x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~COUP standard deviation gassian distribution line~~~~~~~~~~~~~~
                        sd_list = []
                        sd=SDEV_Dist(spec3, test3_list,stat3_std,stat3_min,stat3_max,stat3_avg) 
                        sd_list =sd.gauss()
                        #print('sd_list=',sd_list)
                        min_max = sd.gauss_min_max()
                        #~~~~~~~~~~~~~COUP standard deviation gassian distribution line~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~ COUP Standard deviation Chart~~~~~~~~~~~~~~~~~
                        #print('chart_data=',chart_data)
                        custom_style = Style(colors=('#000000','#FF0000'),title_font_size=39, label_font_size=15,
                                            background='transparent', plot_background='transparent')
                        xy_chart = pygal.XY(style=custom_style,show_dots=False,show_y_labels=False)
                        xy_chart.title = 'COUP Histogram'
                        xy_chart.add('COUP Histogram', sd_list)
                        #print('sd_list=',sd_list)
                        #print('spec3=',spec3)
                        # add  spec line
                        xy_chart.add('spec max', [(spec3, min_max[0]),(spec3, min_max[1])])
                        #~~~~~~~~~~ COUP Standard deviation Chart~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~~~~ COUP Histogram Chart~~~~~~~~~~~~~~~~~~~~
                        try:
                            coup_histo_data2 = xy_chart.render_data_uri()
                            #print('rl_histo=',rl_histo)
                            custom_style = Style(colors=('#ffd138','#201599'),title_font_size=39, label_font_size=17,
                                            value_font_family='googlefont:Raleway',value_font_size=30,value_colors=('white',))
                            hist = pygal.Histogram(fill=True,style=custom_style, human_readable=True,show_x_labels=False,print_values=True)
                            hist.x_labels = x_range_list
                            hist.add('COUP Histogram', coup_histo)
                            hist.title = 'COUP Histogram' 
                            coup_histo_data = hist.render_data_uri()
                            #print('coup_histo_data=',coup_histo_data)
                        except ValueError as e:
                            coup_histo_data =  -1
                        #~~~~~~~~~~~~~~~~ COUP Histogram Chart~~~~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~~~~~ COUP Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                        histo_data = Histogram_data(test_list,spec_list,'test3')
                        coup_histo = histo_data.Hist_data()
                        charts = XY_Chart(test_list,spec_list,'test3') 
                        chart_data = charts.Chart_data()
                        #print('chart_data=',chart_data)
                        custom_style = Style(colors=('#ffd138','#201599'),title_font_size=39, label_font_size=15)
                        xy_chart = pygal.XY(style=custom_style, human_readable=True)
                        xy_chart.title = 'COUP XY Plot'
                        xy_chart.add('Coupling', chart_data)
                        xy_chart.add('spec max', [(0, spec3), (len(chart_data), spec3)])
                        chart3 = xy_chart.render_data_uri()
                        #~~~~~~~~~~~~~~~~~ COUP Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                        
                                           
                        histo_data = Histogram_data(test_list,spec_list,'test4')
                        dir_histo = histo_data.Hist_data()
                        
                        #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
                        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~DIR x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        x_range_list = []
                        x_range = X_Range(dir_histo,spec4, stat4_min,stat4_max,stat4_avg)                
                        x_range_list=x_range.list()
                        #print('x_range_list=',len(x_range_list))
                        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~DIR  x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~DIR standard deviation gassian distribution line~~~~~~~~~~~~~~
                        sd_list = []
                        sd=SDEV_Dist(spec4, test4_list,stat4_std,stat4_min,stat4_max,stat4_avg) 
                        sd_list =sd.gauss()
                        #print('sd_list=',sd_list)
                        min_max = sd.gauss_min_max()
                        #~~~~~~~~~~~~~DIR standard deviation gassian distribution line~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~ DIR Standard deviation Chart~~~~~~~~~~~~~~~~~
                        #print('chart_data=',chart_data)
                        custom_style = Style(colors=('#000000','#FF0000'),title_font_size=39, label_font_size=15,
                                            background='transparent', plot_background='transparent')
                        xy_chart = pygal.XY(style=custom_style,show_dots=False,show_y_labels=False)
                        xy_chart.title = 'DIR Histogram'
                        xy_chart.add('DIR Histogram', sd_list)
                        #print('sd_list=',sd_list)
                        #print('spec4=',spec4)
                        # add  spec line
                        xy_chart.add('spec min', [(0-spec4, min_max[0]),(0-spec4, min_max[1])])
                        xy_chart.add('spec max', [(spec4, min_max[0]),(spec4, min_max[1])])
                        #~~~~~~~~~~ DIR Standard deviation Chart~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~~~~ DIR Histogram Chart~~~~~~~~~~~~~~~~~~~~
                        try:
                            dir_histo_data2 = xy_chart.render_data_uri()
                            #print('rl_histo=',rl_histo)
                            custom_style = Style(colors=('#130fff','#201599'),title_font_size=39, label_font_size=17,
                                            value_font_family='googlefont:Raleway',value_font_size=30,value_colors=('white',))
                            hist = pygal.Histogram(fill=True,style=custom_style, human_readable=True,show_x_labels=False,print_values=True)
                            hist.x_labels = x_range_list
                            hist.add('DIR Histogram', dir_histo)
                            hist.title = 'DIR Histogram' 
                            dir_histo_data = hist.render_data_uri()
                            #print('dir_histo_data=',dir_histo_data)
                        except ValueError as e:
                            dir_histo_data  =  -1
                        #~~~~~~~~~~~~~~~~ DIR Histogram Chart~~~~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~~~~~ DIR Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                        histo_data = Histogram_data(test_list,spec_list,'test4')
                        dir_histo = histo_data.Hist_data()
                        charts = XY_Chart(test_list,spec_list,'test4') 
                        chart_data = charts.Chart_data()
                        #print('chart_data=',chart_data)
                        custom_style = Style(colors=('#130fff','#201599'),title_font_size=39, label_font_size=15)
                        xy_chart = pygal.XY(style=custom_style, human_readable=True)
                        xy_chart.title = 'DIR XY Plot'
                        xy_chart.add('Directivity', chart_data)
                        xy_chart.add('spec min', [(0, 0-spec4), (len(chart_data), 0-spec4)])
                        xy_chart.add('spec max', [(0, spec4), (len(chart_data), spec4)])
                        chart4 = xy_chart.render_data_uri()
                        #~~~~~~~~~~~~~~~~~ DIR Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                        
                        #
                        histo_data = Histogram_data(test_list,spec_list,'test5')
                        cb_histo = histo_data.Hist_data()
                        
                        #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
                        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~CB x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        x_range_list = []
                        x_range = X_Range(cb_histo,spec5, stat5_min,stat5_max,stat5_avg)                
                        x_range_list=x_range.list()
                        #print('x_range_list=',len(x_range_list))
                        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~CB  x-range~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~CB standard deviation gassian distribution line~~~~~~~~~~~~~~
                        sd_list = []
                        sd=SDEV_Dist(spec5, test5_list,stat5_std,stat5_min,stat5_max,stat5_avg) 
                        sd_list =sd.gauss()
                        #print('sd_list=',sd_list)
                        min_max = sd.gauss_min_max()
                        #~~~~~~~~~~~~~CB standard deviation gassian distribution line~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~ CB Standard deviation Chart~~~~~~~~~~~~~~~~~
                        #print('chart_data=',chart_data)
                        custom_style = Style(colors=('#000000','#FF0000'),title_font_size=39, label_font_size=15,
                                            background='transparent', plot_background='transparent')
                        xy_chart = pygal.XY(style=custom_style,show_dots=False,show_y_labels=False)
                        xy_chart.title = 'CB Histogram'
                        xy_chart.add('CB Histogram', sd_list)
                        #print('sd_list=',sd_list)
                        #print('spec5=',spec5)
                        # add  spec line
                        xy_chart.add('spec min', [(spec5, 0-min_max[0]),(spec5, 0-min_max[1])])
                        xy_chart.add('spec max', [(spec5, min_max[0]),(spec5, min_max[1])])
                        #~~~~~~~~~~ CB Standard deviation Chart~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~~~~ CB Histogram Chart~~~~~~~~~~~~~~~~~~~~
                        try:
                            cb_histo_data2 = xy_chart.render_data_uri()
                            custom_style = Style(colors=('#ff2617','#201599'),title_font_size=39, label_font_size=17,
                                            value_font_family='googlefont:Raleway',value_font_size=30,value_colors=('white',))
                            hist = pygal.Histogram(fill=True,style=custom_style, human_readable=True,show_x_labels=False,print_values=True)
                            hist.x_labels = x_range_list
                            hist.add('CB Histogram', cb_histo)
                            hist.title = 'CB Histogram' 
                            cb_histo_data = hist.render_data_uri()
                            #print('cb_histo_data=',cb_histo_data)
                        except ValueError as e:
                            cb_histo_data  =  -1
                        #~~~~~~~~~~~~~~~~ CB Histogram Chart~~~~~~~~~~~~~~~~~~~~
                        
                        #~~~~~~~~~~~~~~~~~ CB Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                        histo_data = Histogram_data(test_list,spec_list,'test5')
                        cb_histo = histo_data.Hist_data()
                        charts = XY_Chart(test_list,spec_list,'test5') 
                        chart_data = charts.Chart_data()
                        #print('chart_data=',chart_data)
                        custom_style = Style(colors=('#ff2617','#201599'),title_font_size=39, label_font_size=15)
                        xy_chart = pygal.XY(style=custom_style, human_readable=True)
                        xy_chart.title = 'CB XY Plot'
                        xy_chart.add('Coupling Balance', chart_data)
                        xy_chart.add('spec min', [(0, 0-spec5), (len(chart_data), 0-spec5)])
                        xy_chart.add('spec max', [(0, spec5), (len(chart_data), spec5)])
                        chart5 = xy_chart.render_data_uri()
                        #~~~~~~~~~~~~~~~~~ CB Data XY Chart~~~~~~~~~~~~~~~~~~~~~
                        
                       
           
            workstation_status = ReportQueue.objects.using('TEST').filter(reportstatus='test running').values_list('workstation','jobnumber','partnumber','operator','value','maxvalue','ping').all()
            x=1
            for station, jobs, parts, opera, value, maxvalue, ping in workstation_status:
                gauge = pygal.SolidGauge(
                show_legend=False, half_pie=True, inner_radius=0.70,
                style=pygal.style.styles['default'](value_font_size=80,plot_background="gray"))
                efficiency = Effeciency.objects.using('TEST').filter(workstation=station).filter(jobnumber=jobs).filter(operator=opera).last()
                print('efficiency=',efficiency)
                #print('************************ping=',ping)
                if efficiency:
                    comment = 'Workstation: ' + str(station) + '\nOperator: ' + str(opera) + '\nJob: ' + str(jobs) + '\nPart: ' + str(parts) + '\nTotal Parts: ' + str(efficiency.totaluuts) + '\nParts Complete: ' + str(efficiency.completeuuts) + '\nOperator Effeciency: ' + str(efficiency.effeciencystatus)
                else:
                    comment = 'Workstation: ' + str(station) + '\nOperator: ' + str(opera) + '\nJob: ' + str(jobs)
                percent_formatter = lambda x: '{:.10g}%'.format(x)
                dollar_formatter = lambda x: '{:.10g}$'.format(x)
                gauge.value_formatter = percent_formatter
                if value:                
                    new_val = ((value/maxvalue) * 100)
                    gauge.add('', [{'value': int(new_val), 'max_value': 100}])
                    #print('value=',value,' maxvalue=',maxvalue)
                if x == 1:
                    test_status1=gauge.render_data_uri()
                    test_comment1 = comment
                elif x == 2:
                    test_status2=gauge.render_data_uri()
                    test_comment2 = comment
                elif x == 3:
                    test_status3=gauge.render_data_uri()
                    test_comment3=comment
                elif x == 4:
                    test_status4=gauge.render_data_uri()
                    test_comment4 = comment
                elif x == 5:
                    test_status5=gauge.render_data_uri()
                    test_comment5 = comment
                elif x == 6:
                    test_status6=gauge.render_data_uri()
                    test_comment6 = comment
                elif x == 7:
                    test_status7=gauge.render_data_uri()
                    test_comment7 = comment
                elif x == 8:
                    test_status8=gauge.render_data_uri()
                    test_comment8 = comment
                elif x == 9:
                    test_status9=gauge.render_data_uri()
                    test_comment9 = comment
                elif x == 10:
                    test_status10=gauge.render_data_uri()
                    test_comment10 = comment
                x+=1
                #print('test_status1',test_status1)
                #print('comment=',comment)
            
        except ValueError as e:
            print ("Lists load Failure ", e)
            print('error = ',e)     
        return render (self.request,"excel/index.html",{'job_num':job_num,'part_num':part_num,'workstation':workstation,'operator':operator,'start_date':start_date,'end_date':end_date,'artwork_list':artwork_list,'artwork':artwork,
                                                        'job_list':job_list,'part_list':part_list,'workstation_list':workstation_list,'operator_list':operator_list,'spec1':spec1,'spec2':spec1,'spec3':spec3,'spectype':spectype,
                                                        'spec4':spec4,'spec5':spec5,'report_data':report_data,'test1_list':test1_list,'test2_list':test2_list,'test3_list':test3_list,'test4_list':test4_list,'test5_list':test5_list,
                                                        'stat1_min':stat1_min,'stat1_max':stat1_max,'stat1_avg':stat1_avg,'stat1_std':stat1_std,'stat2_min':stat2_min,'stat2_max':stat2_max,'stat2_avg':stat2_avg,'stat2_std':stat2_std,
                                                        'stat3_min':stat3_min,'stat3_max':stat3_max,'stat3_avg':stat3_avg,'stat3_std':stat3_std,'stat4_min':stat4_min,'stat4_max':stat4_max,'stat4_avg':stat4_avg,'stat4_std':stat4_std,
                                                        'stat5_min':stat3_min,'stat5_max':stat5_max,'stat5_avg':stat5_avg,'stat5_std':stat5_std,'analyze':analyze,'il_histo_data':il_histo_data,'rl_histo_data':rl_histo_data,
                                                        'iso_histo_data':iso_histo_data,'ab_histo_data':ab_histo_data,'pb_histo_data':pb_histo_data,'coup_histo_data':coup_histo_data,'iso_histo_data':iso_histo_data,
                                                        'passed1':passed1,'failed1':failed1,'failed_percent1':failed_percent1,'passed2':passed2,'failed2':failed2,'failed_percent2':failed_percent2,'passed3':passed3,'failed3':failed3,   
                                                        'cb_histo_data':cb_histo_data,'failed_percent3':failed_percent3,'passed4':passed4,'failed4':failed4,'failed_percent4':failed_percent4,'passed5':passed5,'failed5':failed5,
                                                        'failed_percent5':failed_percent5,'bad_data1':bad_data1,'bad_data2':bad_data2,'bad_data3':bad_data3,'bad_data4':bad_data4,'bad_data5':bad_data5,'il_histo_data2':il_histo_data2,
                                                        'chart1':chart1,'chart2':chart2,'chart3':chart3,'chart4':chart4,'chart5':chart5,'blank':blank,'total':total,'art_rev_list':art_rev_list,'blank':blank,'test_status1':test_status1,
                                                        'test_status2':test_status2,'test_status3':test_status3,'test_status4':test_status4,'test_status5':test_status5,'test_status6':test_status6,'test_status7':test_status7,
                                                        'test_status8':test_status8,'test_status9':test_status9,'test_status10':test_status10,'test_comment1':test_comment1,'test_comment2':test_comment2,'test_comment3':test_comment3,'test_comment4':test_comment4,
                                                        'test_comment5':test_comment5,'test_comment6':test_comment6,'test_comment7':test_comment7,'test_comment8':test_comment8,'test_comment9':test_comment9,'test_comment10':test_comment10,
                                                        'rl_histo_data2':rl_histo_data2,'iso_histo_data2':iso_histo_data2,'ab_histo_data2':ab_histo_data2,'pb_histo_data2':pb_histo_data2,'coup_histo_data2':coup_histo_data2,'dir_histo_data2':dir_histo_data2,
                                                        'cb_histo_data2':cb_histo_data2,'bad_data1':bad_data1,'bad_data2':bad_data2,'bad_data3':bad_data3,'bad_data4':bad_data4,'bad_data5':bad_data5,'bad1':bad1,'bad2':bad2,'bad3':bad3,'bad4':bad4,'bad5':bad5})
                                                        


def export_users_xls(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="users.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Users Data') # this will make a sheet named Users Data
    
    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['Username', 'First Name', 'Last Name', 'Email Address', ]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style) # at 0 row 0 column 

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    rows = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response
    
#This code will explain how to Style your Excel File. The bellow code will explain Wrap text in the cell, background color, border, and text color.    
def export_styling_xls(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="users.xlsx"'

    wb = Workbook()
    sheet = wb['Sheet']
    sheet.title = 'Summary'
    sheet = wb.create_sheet('Styling Data') # this will make a sheet named Users Data - First Sheet
       
    makesheet=CreateSheets('Styling Data',sheet)
    makesheet.tabular_data()
    print('we are here after sytling')
    
    
    '''
    styles = dict(
        bold = 'font: bold 1',
        italic = 'font: italic 1',
        # Wrap text in the cell
        wrap_bold = 'font: bold 1; align: wrap 1;',
        # White text on a blue background
        reversed = 'pattern: pattern solid, fore_color blue; font: color white;',
        # Light orange checkered background
        light_orange_bg = 'pattern: pattern fine_dots, fore_color white, back_color orange;',
        # Heavy borders
        bordered = 'border: top thick, right thick, bottom thick, left thick;',
        # 16 pt red text
        bordered = 'border: top thick, right thick, bottom thick, left thick;',
        # 16 pt red text
        big_red = 'font: height 320, color red;',
    )

    for idx, k in enumerate(sorted(styles)):
        style = xlwt.easyxf(styles[k])
        ws.write(idx, 0, k)
        ws.write(idx, 1, styles[k], style)
    '''
    
    wb.save(response)

    return response

#The below code will explain how to write data in Exisiting excel file and the content inside it.
def export_write_xls(request):
    # EG: path = excel_app/sample.xls
    path = os.path.dirname(__file__)
    path = path + '/excel_templates/'
    print('path=',path)
    file = os.path.join(path, 'TestData.xlsx')
    print('file=',file)

    wb = load_workbook(file)
    print('wb=',wb)
    sheet = wb["Raw Data1"]
    print('sheet=',sheet)
    sheet['F2'] = '398789-02' 
    sheet['F3'] = 'IPP-89348' 
    sheet['F4'] = '90 degree coupler' 
    print("sheet['F2']=",sheet['F2'])
    print(wb.sheetnames)
    wb.save("C:/ATE Data/demo4.xlsx")
    wb = load_workbook("C:/ATE Data/demo4.xlsx")
    return response
    
    

def test_report():
    

    # EG: path = report/excel_templates/TestData.xls
    path = os.path.dirname(__file__)
    path = path + '/excel_templates/'
    print('path=',path)
    file = os.path.join(path, 'TestData.xls')

    rb = open_workbook(file, formatting_info=True)
    sh = rb.sheet_by_name('Data')

   

    row_num = 2 # index start from 0
    rows = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num])
    
    # wb.save(file) # will replace original file
    # wb.save(file + '.out' + os.path.splitext(file)[-1]) # will save file where the excel file is
    wb.save(response)
    return response
   