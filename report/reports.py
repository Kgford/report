import xlwt
from xlutils.copy import copy # http://pypi.python.org/pypi/xlutils
from xlrd import open_workbook # http://pypi.python.org/pypi/xlrd
from openpyxl import Workbook
from openpyxl import load_workbook,drawing
from openpyxl.styles import PatternFill, Alignment#Connect cell styles
from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill#Connect styles for text
from openpyxl.styles import colors#Connect colors for text and cells
from openpyxl.chart import LineChart,ScatterChart,Reference,Series

from django.http import HttpResponse
from report.overhead import TimeCode, Security, StringThings,Conversions
from test_db.models import Specifications,Workstation,Workstation1,Testdata,Testdata3,Trace,Tracepoints,Tracepoints2,Effeciency,ReportQueue
import os
import statistics 
import time


class CreateSheets:
    def __init__ (self, sheet_name,worksheet,artwork_len):
        self.sheet_name = sheet_name
        self.worksheet = worksheet 
        self.artwork_len = artwork_len          
         
        print('self.sheet_name=',self.sheet_name)
        print('self.worksheet=',self.worksheet)
    
    def set_outside_thin_border(self, cell_range):
        rows = self.worksheet[cell_range]
        side = Side(border_style='thin', color="FF000000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border
    
                
    def set_outside_thick_border(self, cell_range):
        rows = self.worksheet[cell_range]
        side = Side(border_style='thick', color="FF000000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border
    
    def set_border_range(self, cell_range):
        thin = Side(border_style="thin", color="000000")#Border style, color
        border = Border(left=thin, right=thin, top=thin, bottom=thin)#Position of border

        for row in self.worksheet[cell_range]:
            for cell in row:
                cell.border = border
    
    
    def summary(self):
        
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~mearge rows~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        self.worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6) #first row
        
        print('Summary setting widths')
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~set column witdhs~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # set the width of the column
        self.worksheet.column_dimensions['A'].width = 12.22
        self.worksheet.column_dimensions['B'].width = 12.89
        self.worksheet.column_dimensions['C'].width = 15.67
        self.worksheet.column_dimensions['D'].width = 17.43
        self.worksheet.column_dimensions['E'].width = 18.00
        self.worksheet.column_dimensions['F'].width = 18.57
        self.worksheet.column_dimensions['G'].width = 12.00
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~set column witdhs~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        
        
        
        print('Adding fonts and words')
        row=1
        font = Font(name='Arial',size=18,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='7FFF00')
        aa=self.worksheet['A'+str(row)]
        aa.font=font
        aa.alignment = Alignment(horizontal='center')
        self.worksheet['A'+str(row)]='Summary Data'
        
        row=3
        font = Font(name='Arial',size=10,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        g=self.worksheet['G'+str(row)]
        g.font=font
        g.alignment = Alignment(horizontal='center')
        self.worksheet['G'+str(row)]='Total Tested'
        h=self.worksheet['H'+str(row)]
        h.font=font
        h.alignment = Alignment(horizontal='center')
        self.worksheet['H'+str(row)]=' Passed'
        i=self.worksheet['I'+str(row)]
        i.font=font
        i.alignment = Alignment(horizontal='center')
        self.worksheet['I'+str(row)]='Failed'
        
        for x in range(3):
            print('x in range=',x)
            if x==0:
                row=3
            else:
                row = row + 3 + self.artwork_len 
                
            for col_range in range(2, 7):
                cell_title = self.worksheet.cell(row+1, col_range)
                cell_title.fill = PatternFill(start_color="e2eb34", end_color="e2eb34", fill_type="solid") #Yellow
            
            print('summary row=',row,' artwork_len=',self.artwork_len)
            font = Font(name='Arial',size=10,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
            a=self.worksheet['A'+str(row)]
            a.font=font
            a.alignment = Alignment(horizontal='center')
            self.worksheet['A'+str(row)]='Artwork Rev'
            print('Summary setting borders')
            self.set_border_range('B' + str(row) + ':F' + str(row+1)) 
            print('Summary setting borders')
            if x == 0:
                meas='Avg'
            elif x==1:
                meas='Min'
            elif x==2:
                meas='Max'    
            
            b=self.worksheet['B'+str(row)]
            b.font=font
            b.alignment = Alignment(horizontal='center')
            self.worksheet['B'+str(row)]= meas + '. IL (dB)'
            c=self.worksheet['C'+str(row)]
            c.font=font
            c.alignment = Alignment(horizontal='center')
            self.worksheet['C'+str(row)]= meas + '. RL (dB)'
            d=self.worksheet['D'+str(row)]
            d.font=font
            d.alignment = Alignment(horizontal='center')
            self.worksheet['D'+str(row)]= meas + '. Isolation (dB)'
            e=self.worksheet['E'+str(row)]
            e.font=font
            e.alignment = Alignment(horizontal='center')
            self.worksheet['E'+str(row)]= meas + '. Amp Bal (dB)'
            f=self.worksheet['F'+str(row)]
            f.font=font
            f.alignment = Alignment(horizontal='center')
            self.worksheet['F'+str(row)]= meas + '. Phase Bal (dB)'
            
    
    
    def chart_data(self):
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~set row heights~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        self.worksheet.row_dimensions[1].height = 13.8
        self.worksheet.row_dimensions[2].height = 11.40
        self.worksheet.row_dimensions[3].height = 11.40
        self.worksheet.row_dimensions[4].height = 11.40
        self.worksheet.row_dimensions[5].height = 11.40
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~set row heights~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~set column witdhs~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # set the width of the column
        self.worksheet.column_dimensions['A'].width = 10.50
        self.worksheet.column_dimensions['B'].width = 10.50 
        self.worksheet.column_dimensions['C'].width = 13.8
        self.worksheet.column_dimensions['D'].width = 8.00
        self.worksheet.column_dimensions['E'].width = 6.78 
        self.worksheet.column_dimensions['F'].width = 22.21
        self.worksheet.column_dimensions['G'].width = 19.40
       #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~set column witdhs~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        print('merging cells')
        
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~mearge rows~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        self.worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6) #first row
        
        self.worksheet.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)  #second row
        self.worksheet.merge_cells(start_row=3, start_column=4, end_row=3, end_column=5)  #third row
        self.worksheet.merge_cells(start_row=4, start_column=4, end_row=4, end_column=5)  #fourth row 
        self.worksheet.merge_cells(start_row=5, start_column=4, end_row=5, end_column=5)  #fith row
        self.worksheet.merge_cells(start_row=47, start_column=2, end_row=47, end_column=5) #53 row
        self.worksheet.merge_cells(start_row=48, start_column=1, end_row=48, end_column=3) #48 row
        self.worksheet.merge_cells(start_row=48, start_column=4, end_row=48, end_column=5) #48 row
        self.worksheet.merge_cells(start_row=48, start_column=6, end_row=48, end_column=7) #48 row
        self.worksheet.merge_cells(start_row=48, start_column=8, end_row=48, end_column=10) #48 row
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~mearge rows~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                
        print('at image') 
        
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~add logo~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        image_name = "\\\ippdc\\Test Automation\\Excel_Templates\\logo.png"
        img = drawing.image.Image(image_name)
        img.anchor = 'A2' # Or whatever cell location you want to use.
        self.worksheet.add_image(img)
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~add logo~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
       
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~add words~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        a1 = self.worksheet['A1']
        font = Font(name='Arial',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        a1.font = font
        a1.alignment = Alignment(horizontal='center')
        self.worksheet['A1']='Chart Data'
        
        d2 = self.worksheet['D2']
        d3 = self.worksheet['D3']
        d4 = self.worksheet['D4']
        d5 = self.worksheet['D5']
        font = Font(name='Arial',size=10,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        d2.font = font
        d2.alignment = Alignment(horizontal='right')
        d3.font = font
        d3.alignment = Alignment(horizontal='right')
        d4.font = font
        d4.alignment = Alignment(horizontal='right')
        d5.font = font
        d5.alignment = Alignment(horizontal='right')
        self.worksheet['D2']='Job Number:'
        self.worksheet['D3']='Part Number:'
        self.worksheet['D4']='Part Type:'
        self.worksheet['D5']='Artwork Rev:'
        
        b47 = self.worksheet['B47']
        a48 = self.worksheet['A48']
        d48 = self.worksheet['D48']
        f48 = self.worksheet['F48']
        h48 = self.worksheet['H48']
        font = Font(name='Arial',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        b47.font = font
        b47.alignment = Alignment(horizontal='center')
        a48.font = font
        a48.alignment = Alignment(horizontal='center')
        f48.font = font
        f48.alignment = Alignment(horizontal='center')
        h48.font = font
        h48.alignment = Alignment(horizontal='center')
        self.worksheet['B47']='Test Data Follows:'
        self.worksheet['A48']='Chart 1'
        self.worksheet['D48']='Chart 2:'
        self.worksheet['F48']='Chart 3'
        self.worksheet['H48']='Chart 4'
        
        a49 = self.worksheet['A49']
        b49 = self.worksheet['B49']
        c49 = self.worksheet['C49']
        d49 = self.worksheet['D49']
        e49 = self.worksheet['E49']
        f49 = self.worksheet['F49']
        g49 = self.worksheet['G49']
        h49 = self.worksheet['H49']
        i49 = self.worksheet['I49']
        j49 = self.worksheet['J49']
        font = Font(name='Arial',size=10,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        a49.font = font
        a49.alignment = Alignment(horizontal='center')
        b49.font = font
        b49.alignment = Alignment(horizontal='center')
        c49.font = font
        c49.alignment = Alignment(horizontal='center')
        d49.font = font
        d49.alignment = Alignment(horizontal='center')
        e49.font = font
        e49.alignment = Alignment(horizontal='center')
        f49.font = font
        f49.alignment = Alignment(horizontal='center')
        g49.font = font
        g49.alignment = Alignment(horizontal='center')
        h49.font = font
        h49.alignment = Alignment(horizontal='center')
        i49.font = font
        i49.alignment = Alignment(horizontal='center')
        j49.font = font
        j49.alignment = Alignment(horizontal='center')
        self.worksheet['A49']='Freq MHz'
        self.worksheet['B49']='Trace 1:'
        self.worksheet['C49']='Trace 2'
        self.worksheet['D49']='Freq MHz'
        self.worksheet['E49']='Trace 1:'
        self.worksheet['F49']='Freq MHz'
        self.worksheet['G49']='Trace 1:'
        self.worksheet['H49']='Freq MHz'
        self.worksheet['I49']='Trace 1:'
        self.worksheet['J49']='Trace 2'
        
        
        
    def tabular_data(self):
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~set row heights~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        print('in styling')
        
        # set the height of the row1
        self.worksheet.row_dimensions[1].height = 30.6
        self.worksheet.row_dimensions[2].height = 30.6
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~set row heights~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~set column witdhs~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # set the width of the column
        self.worksheet.column_dimensions['A'].width = 15.44 
        self.worksheet.column_dimensions['B'].width = 7.11 
        self.worksheet.column_dimensions['C'].width = 6.11 
        self.worksheet.column_dimensions['D'].width = 6.11 
        self.worksheet.column_dimensions['E'].width = 6.11 
        self.worksheet.column_dimensions['F'].width = 6.11 
        self.worksheet.column_dimensions['G'].width = 6.11 
        self.worksheet.column_dimensions['H'].width = 9.00 
        self.worksheet.column_dimensions['I'].width = 8.00 
        self.worksheet.column_dimensions['J'].width = 7.00
        self.worksheet.column_dimensions['K'].width = 8.11

        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~set column witdhs~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        print('merging cells')
        
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~mearge rows~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        self.worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11) #first row
        
        self.worksheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=7)  #second row
        self.worksheet.merge_cells(start_row=2, start_column=8, end_row=2, end_column=11) #second row
        self.worksheet.merge_cells(start_row=3, start_column=2, end_row=3, end_column=4)  #third row
        self.worksheet.merge_cells(start_row=3, start_column=5, end_row=3, end_column=7)  #third row
        self.worksheet.merge_cells(start_row=3, start_column=8, end_row=3, end_column=11) #third row
        self.worksheet.merge_cells(start_row=4, start_column=2, end_row=4, end_column=4)  #fourth row 
        self.worksheet.merge_cells(start_row=4, start_column=5, end_row=4, end_column=7)  #fourth row
        self.worksheet.merge_cells(start_row=4, start_column=8, end_row=4, end_column=11) #fourth row
        self.worksheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=4)  #fith row
        self.worksheet.merge_cells(start_row=5, start_column=5, end_row=5, end_column=7)  #fith row
        self.worksheet.merge_cells(start_row=5, start_column=9, end_row=5, end_column=10) #fith row
        
        
        self.worksheet.merge_cells(start_row=7, start_column=2, end_row=7, end_column=11) #seventh row
        
        self.worksheet.merge_cells(start_row=8, start_column=2, end_row=8, end_column=3) #Nineteenth row
        self.worksheet.merge_cells(start_row=8, start_column=4, end_row=8, end_column=5) #Nineteenth row
        self.worksheet.merge_cells(start_row=8, start_column=6, end_row=8, end_column=7) #Nineteenth row
        self.worksheet.merge_cells(start_row=8, start_column=8, end_row=8, end_column=9) #Nineteenth row
        self.worksheet.merge_cells(start_row=8, start_column=10, end_row=8, end_column=11)#Nineteenth row
        
        self.worksheet.merge_cells(start_row=9, start_column=2, end_row=9, end_column=3) #Nineteenth row
        self.worksheet.merge_cells(start_row=9, start_column=4, end_row=9, end_column=5) #Nineteenth row
        self.worksheet.merge_cells(start_row=9, start_column=6, end_row=9, end_column=7) #Nineteenth row
        self.worksheet.merge_cells(start_row=9, start_column=8, end_row=9, end_column=9) #Nineteenth row
        self.worksheet.merge_cells(start_row=9, start_column=10, end_row=9, end_column=11)#Nineteenth row
        
        self.worksheet.merge_cells(start_row=10, start_column=2, end_row=10, end_column=3) #Nineteenth row
        self.worksheet.merge_cells(start_row=10, start_column=4, end_row=10, end_column=5) #Nineteenth row
        self.worksheet.merge_cells(start_row=10, start_column=6, end_row=10, end_column=7) #Nineteenth row
        self.worksheet.merge_cells(start_row=10, start_column=8, end_row=10, end_column=9) #Nineteenth row
        self.worksheet.merge_cells(start_row=10, start_column=10, end_row=10, end_column=11)#Nineteenth row
        
        self.worksheet.merge_cells(start_row=11, start_column=2, end_row=11, end_column=3) #Nineteenth row
        self.worksheet.merge_cells(start_row=11, start_column=4, end_row=11, end_column=5) #Nineteenth row
        self.worksheet.merge_cells(start_row=11, start_column=6, end_row=11, end_column=7) #Nineteenth row
        self.worksheet.merge_cells(start_row=11, start_column=8, end_row=11, end_column=9) #Nineteenth row
        self.worksheet.merge_cells(start_row=11, start_column=10, end_row=11, end_column=11)#Nineteenth row
        
        self.worksheet.merge_cells(start_row=12, start_column=2, end_row=12, end_column=3) #Nineteenth row
        self.worksheet.merge_cells(start_row=12, start_column=4, end_row=12, end_column=5) #Nineteenth row
        self.worksheet.merge_cells(start_row=12, start_column=6, end_row=12, end_column=7) #Nineteenth row
        self.worksheet.merge_cells(start_row=12, start_column=8, end_row=12, end_column=9) #Nineteenth row
        self.worksheet.merge_cells(start_row=12, start_column=10, end_row=12, end_column=11)#Nineteenth row
        
        self.worksheet.merge_cells(start_row=13, start_column=2, end_row=13, end_column=3) #Nineteenth row
        self.worksheet.merge_cells(start_row=13, start_column=4, end_row=13, end_column=5) #Nineteenth row
        self.worksheet.merge_cells(start_row=13, start_column=6, end_row=13, end_column=7) #Nineteenth row
        self.worksheet.merge_cells(start_row=13, start_column=8, end_row=13, end_column=9) #Nineteenth row
        self.worksheet.merge_cells(start_row=13, start_column=10, end_row=13, end_column=11)#Nineteenth row
        
        self.worksheet.merge_cells(start_row=14, start_column=2, end_row=14, end_column=3) #Nineteenth row
        self.worksheet.merge_cells(start_row=14, start_column=4, end_row=14, end_column=5) #Nineteenth row
        self.worksheet.merge_cells(start_row=14, start_column=6, end_row=14, end_column=7) #Nineteenth row
        self.worksheet.merge_cells(start_row=14, start_column=8, end_row=14, end_column=9) #Nineteenth row
        self.worksheet.merge_cells(start_row=14, start_column=10, end_row=14, end_column=11)#Nineteenth row
        
        self.worksheet.merge_cells(start_row=15, start_column=2, end_row=15, end_column=3) #Nineteenth row
        self.worksheet.merge_cells(start_row=15, start_column=4, end_row=15, end_column=5) #Nineteenth row
        self.worksheet.merge_cells(start_row=15, start_column=6, end_row=15, end_column=7) #Nineteenth row
        self.worksheet.merge_cells(start_row=15, start_column=8, end_row=15, end_column=9) #Nineteenth row
        self.worksheet.merge_cells(start_row=15, start_column=10, end_row=15, end_column=11)#Nineteenth row
        
        self.worksheet.merge_cells(start_row=16, start_column=2, end_row=16, end_column=3) #Nineteenth row
        self.worksheet.merge_cells(start_row=16, start_column=4, end_row=16, end_column=5) #Nineteenth row
        self.worksheet.merge_cells(start_row=16, start_column=6, end_row=16, end_column=7) #Nineteenth row
        self.worksheet.merge_cells(start_row=16, start_column=8, end_row=16, end_column=9) #Nineteenth row
        self.worksheet.merge_cells(start_row=16, start_column=10, end_row=16, end_column=11)#Nineteenth row
        
        
        self.worksheet.merge_cells(start_row=18, start_column=2, end_row=18, end_column=11) #Eighteenth row
        
        
        self.worksheet.merge_cells(start_row=19, start_column=1, end_row=20, end_column=1) #Nineteenth row
        self.worksheet.merge_cells(start_row=19, start_column=2, end_row=19, end_column=3) #Nineteenth row
        self.worksheet.merge_cells(start_row=19, start_column=4, end_row=19, end_column=5) #Nineteenth row
        self.worksheet.merge_cells(start_row=19, start_column=6, end_row=19, end_column=7) #Nineteenth row
        self.worksheet.merge_cells(start_row=19, start_column=8, end_row=19, end_column=9) #Nineteenth row
        self.worksheet.merge_cells(start_row=19, start_column=10, end_row=19, end_column=11)#Nineteenth row
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~mearge rows~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                
        print('setting borders')
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~set borders~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        self.set_border_range('A1:K5')
        self.set_border_range('B8:K8')
        self.set_border_range('A9:K16')
        self.set_border_range('A19:K20')
        self.set_outside_thick_border('A2:K5')
        self.set_outside_thick_border('B8:K8')
        self.set_outside_thick_border('A9:K16')
        self.set_outside_thick_border('A19:K20')
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~set borders~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        print('at image') 
        
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~add logo~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        image_name = "\\\ippdc\\Test Automation\\Excel_Templates\\logo.png"
        img = drawing.image.Image(image_name)
        img.anchor = 'A2' # Or whatever cell location you want to use.
        self.worksheet.add_image(img)
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~add logo~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        
        print('adding color')
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~add color~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        for col_range in range(1, 12):
            cell_title = self.worksheet.cell(9, col_range)
            cell_title.fill = PatternFill(start_color="e2eb34", end_color="e2eb34", fill_type="solid") #Yellow
            cell_title = self.worksheet.cell(14, col_range)
            cell_title.fill = PatternFill(start_color="5ceb34", end_color="5ceb34", fill_type="solid") #Green
            cell_title = self.worksheet.cell(15, col_range)
            cell_title.fill = PatternFill(start_color="eb3434", end_color="eb3434", fill_type="solid")#Green
            cell_title = self.worksheet.cell(16, col_range)
            cell_title.fill = PatternFill(start_color="eb3434", end_color="eb3434", fill_type="solid") #Red
            
        for col_range in range(2, 12):
            cell_title = self.worksheet.cell(20, col_range)
            cell_title.fill = PatternFill(start_color="e2eb34", end_color="e2eb34", fill_type="solid") #Yellow
        print('finished color')
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~add color~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~Align cells~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        for x in range(8):
            row=x+9
            b=self.worksheet['B'+str(row)]
            b.alignment = Alignment(horizontal='center')
            d=self.worksheet['D'+str(row)]
            d.alignment = Alignment(horizontal='center')
            f=self.worksheet['F'+str(row)]
            f.alignment = Alignment(horizontal='center')
            h=self.worksheet['H'+str(row)]
            h.alignment = Alignment(horizontal='center')
            j=self.worksheet['J'+str(row)]
            j.alignment = Alignment(horizontal='center')
         #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~Align cells~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~  


       #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~add words~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        a1 = self.worksheet['A1']
        font = Font(name='Arial',size=22,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        a1.font = font
        a1.alignment = Alignment(horizontal='center')
        self.worksheet['A1']='Test Data'
        
        e2 = self.worksheet['E2']
        e3 = self.worksheet['E3']
        e4 = self.worksheet['E4']
        e5 = self.worksheet['E5']
        a4 = self.worksheet['A4']
        a5 = self.worksheet['A5']
        a9 = self.worksheet['A9']
        i5 = self.worksheet['I5']
        font = Font(name='Arial',size=10,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        e2.font = font
        e2.alignment = Alignment(horizontal='right')
        e3.font = font
        e3.alignment = Alignment(horizontal='right')
        e4.font = font
        e4.alignment = Alignment(horizontal='right')
        e5.font = font
        e5.alignment = Alignment(horizontal='right')
        a4.font = font
        a4.alignment = Alignment(horizontal='right')
        a5.font = font
        a5.alignment = Alignment(horizontal='right')
        a9.font = font
        a9.alignment = Alignment(horizontal='right')
        i5.font = font
        i5.alignment = Alignment(horizontal='right')
        self.worksheet['E2']='Job Number:'
        self.worksheet['E3']='Part Number:'
        self.worksheet['E4']='Part Type:'
        self.worksheet['E5']='Artwork Rev:'
        self.worksheet['A4']='Operator:'
        self.worksheet['A5']='Workstation:'
        self.worksheet['I5']='Location:'
        
        font = Font(name='Arial',size=10,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        a9 = self.worksheet['A9']
        a10 = self.worksheet['A10']
        a11 = self.worksheet['A11']
        a12 = self.worksheet['A12']
        a13 = self.worksheet['A13']
        a14 = self.worksheet['A14']
        a15 = self.worksheet['A15']
        a16 = self.worksheet['A16']
        a9.font = font
        a9.alignment = Alignment(horizontal='right')
        a10.font = font
        a10.alignment = Alignment(horizontal='right')
        a11.font = font
        a11.alignment = Alignment(horizontal='right')
        a12.font = font
        a12.alignment = Alignment(horizontal='right')
        a13.font = font
        a13.alignment = Alignment(horizontal='right')
        a14.font = font
        a14.alignment = Alignment(horizontal='right')
        a15.alignment = Alignment(horizontal='right')
        a16.alignment = Alignment(horizontal='right')
        self.worksheet['A9']='Specification'
        self.worksheet['A10']='Average'
        self.worksheet['A11']='Minumum'
        self.worksheet['A12']='Maximum'
        self.worksheet['A13']='standard Deviation'
        self.worksheet['A14']='Qty Passed'
        font = Font(name='Arial',size=10,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FFFFFFFF')
        a15.font = font
        a16.font = font
        self.worksheet['A15']='Qty Failed'
        self.worksheet['A16']='% Failed'
        
        font = Font(name='Arial',size=9,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        a19 = self.worksheet['A19']
        b8 = self.worksheet['B8']
        d8 = self.worksheet['D8']
        f8 = self.worksheet['F8']
        h8 = self.worksheet['H8']
        j8 = self.worksheet['J8']
        b7 = self.worksheet['B7']
        a19.font = font
        a19.alignment = Alignment(horizontal='center')
        b8.font = font
        b8.alignment = Alignment(horizontal='center')
        d8.font = font
        d8.alignment = Alignment(horizontal='center')
        f8.font = font
        f8.alignment = Alignment(horizontal='center')
        h8.font = font
        h8.alignment = Alignment(horizontal='center')
        j8.font = font
        j8.alignment = Alignment(horizontal='center')
        b7.alignment = Alignment(horizontal='center')
        b7.font = font
        
        self.worksheet['B7']='Statistics'
        self.worksheet['B8']='IL'
        self.worksheet['D8']='RL'
        self.worksheet['F8']='ISO'
        self.worksheet['H8']='AB'
        self.worksheet['J8']='PB'
        self.worksheet['A19']='Serial Number'
        
        
        b18 = self.worksheet['B18']
        b19 = self.worksheet['B19']
        d19 = self.worksheet['D19']
        f19 = self.worksheet['F19']
        h19 = self.worksheet['H19']
        j19 = self.worksheet['J19']
        b18.font = font
        b18.alignment = Alignment(horizontal='center')
        b19.font = font
        b19.alignment = Alignment(horizontal='center')
        d19.font = font
        d19.alignment = Alignment(horizontal='center')
        f19.font = font
        f19.alignment = Alignment(horizontal='center')
        h19.font = font
        h19.alignment = Alignment(horizontal='center')
        j19.font = font
        j19.alignment = Alignment(horizontal='center')
        self.worksheet['B18']='Data'
        self.worksheet['B19']='Insertion Loss'
        self.worksheet['D19']='Return loss'
        self.worksheet['F19']='Isolation'
        self.worksheet['H19']='Amplitude Balance'
        self.worksheet['J19']='Phase Balance'
    

class ExcelReports:
    def __init__ (self, job_num,operator,workstation):
        self.job_num = job_num
        self.operator = operator
        self.workstation = workstation
        #print('job_num=',self.job_num)
               
    
    def test_data(self):
        filter_bad_data=True
        job_list = Testdata.objects.using('TEST').filter(jobnumber=self.job_num).order_by('jobnumber').values_list('jobnumber', flat=True).distinct()
        part_list = Testdata.objects.using('TEST').filter(jobnumber=self.job_num).order_by('partnumber').values_list('partnumber', flat=True).distinct()
        artwork_list = Testdata.objects.using('TEST').filter(jobnumber=self.job_num).order_by('partnumber').values_list('artwork_rev', flat=True).distinct()
        report_data = Testdata.objects.using('TEST').filter(jobnumber=self.job_num).all()
        print('running report')
        ReportQueue.objects.using('TEST').filter(jobnumber=self.job_num).filter(workstation=self.workstation).update(reportstatus='running report')
        spec_data=None
        if report_data:
            part_num = report_data[0].partnumber
            spec_data = Specifications.objects.using('TEST').filter(jobnumber=self.job_num).last()
            spectype = spec_data.spectype
            paths = ReportFiles(self.job_num,part_num,spectype)
            data_path = paths.data_path()
            template_path = paths.template()
            #print('template_path=',template_path)
        
        wb = Workbook()
        print('workbook=',wb)
        print('wb.sheetnames=',wb.sheetnames)
        
        sheet = wb['Sheet']
        sheet.title = 'Summary'
        
        print('artwork_list1',artwork_list)
        if not artwork_list:
            artwork_list = ['RawData 1',]
        
        print('artwork_list2',artwork_list)
        #filter blanks
        temp_list = []
        for artwork_rev in artwork_list:
            if artwork_rev and artwork_rev != '':
                temp_list.append(artwork_rev)
        artwork_list = temp_list
        print('artwork_list3',artwork_list)
        if not artwork_list:
            artwork_list = ['UNKN REV',]
        print('artwork_list4',artwork_list)
               
                
        x=1
        z=1
        print('loading data')
        sum_row = 5
        ReportQueue.objects.using('TEST').filter(jobnumber=self.job_num).filter(workstation=self.workstation).update(reportstatus='loading data')
        if spec_data:
            for artwork_rev in artwork_list:
                if 'UNKN REV' in artwork_rev:
                    report_data = Testdata.objects.using('TEST').filter(jobnumber=self.job_num).all()
                    data_count = Testdata.objects.using('TEST').filter(jobnumber=self.job_num).count()
                else:
                    report_data = Testdata.objects.using('TEST').filter(jobnumber=self.job_num).filter(artwork_rev=artwork_rev).all()
                    data_count = Testdata.objects.using('TEST').filter(jobnumber=self.job_num).filter(artwork_rev=artwork_rev).count()
                
                print('data_count=',data_count)
                
                conversions = Conversions(spec_data.vswr,'')
                spec_rl = round(conversions.vswr_to_rl(),2)
                
                print('spec_rl=',spec_rl)
                print('spec_data=',spec_data)
                print('report_data=',report_data)
                #time.sleep(20)
                if '90 DEGREE COUPLER' in spectype or 'BALUN' in spectype:
                    spec_list = [spec_data.insertionloss,spec_rl,spec_data.isolation,spec_data.amplitudebalance,spec_data.phasebalance,spec_data.ab_ex] 
                elif 'DIRECTIONAL COUPLER' in spectype: 
                    spec_list = [spec_data.insertionloss,spec_rl,spec_data.coupling,spec_data.directivity,spec_data.coupledflatness]
                    
                #print('spec_list=',spec_list)
                if report_data:
                    part_num = report_data[0].partnumber
                    print('part_num=',part_num)
                    spectype = spec_data.spectype
                    print('artwork_rev=',artwork_rev)
                    if '/' in artwork_rev:
                        artwork_rev=artwork_rev.replace("/", "_") 
                        print('artwork_rev2=',artwork_rev)
                    if artwork_rev == None or artwork_rev =='':
                        #artwork_rev = 'UNKN REV'
                        #unknown_rev='UNKN REV_0'
                        num=int(unknown_rev[9])
                        #print('num=',num)
                        unknown_rev = unknown_rev[:-1]
                        artwork_rev=unknown_rev + str(num+1)
                        #print('artwork_rev=',artwork_rev)
                        
                    
                    #create new sheet and format 
                    print('making data sheet',artwork_rev)
                    sheet = wb.create_sheet(artwork_rev) 
                    makesheet=CreateSheets(artwork_rev,sheet,len(artwork_list))
                    makesheet.tabular_data()
                    #format summary sheet
                    
                    sheet = wb['Summary']
                    print('sheet before=',sheet)
                    
                    makesheet=CreateSheets(artwork_rev,sheet,len(artwork_list))
                    makesheet.summary()
                    print ('sheetnames=',wb.sheetnames)
                    print('artwork_rev=',artwork_rev)
                    sheet = wb[artwork_rev]
                    
                    print('sheet=',sheet)
                    sheet['B4'] = self.operator 
                    sheet['B5'] = self.workstation 
                    sheet['H2'] = self.job_num
                    sheet['H3'] = part_num 
                    sheet['H4'] = spectype 
                    sheet['H5'] = artwork_rev 
                    
                    #~~~~~~~~~~~~configure  the tests~~~~~~~~~~~~~
                    if 'DIRECTIONAL COUPLER' in spectype:
                        sheet['F19'] = "Coupling"
                        sheet['H19'] = "Directivity"
                        sheet['J19'] = "Coupling Flatness"
                    elif 'BALUN' in spectype:
                        sheet['F19'] = "No Test"
                        sheet['H19'] = "Amplitude Balance"
                        sheet['J19'] = "Phase Balance"
                    elif 'TRANSFORMER' in spectype:
                        sheet['F19'] = "No Test"
                        sheet['H19'] = "No Test"
                        sheet['J19'] = "No Test"
                    else:
                        sheet['F19'] = "Isolation"
                        sheet['H19'] = "Amplitude Balance"
                        sheet['J19'] = "Phase Balance"
                    #~~~~~~~~~~~~choose the tests~~~~~~~~~~~~~
                    
                    #~~~~~~~~~~~~~~~~~~~~~format the spec cells for data or multiband data~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                    #Mearge split cells for normal data
                    if spec_data.ab_exp_tf or spec_data.dir_exp_tf:  # Dual Band AB/DIR only Don't mearge AB cels
                        sheet.merge_cells(start_row=20, start_column=2, end_row=20, end_column=3) #IL
                        sheet.merge_cells(start_row=20, start_column=4, end_row=20, end_column=5) #RL
                        sheet.merge_cells(start_row=20, start_column=6, end_row=20, end_column=7)  #ISO/Coup
                        sheet.merge_cells(start_row=20, start_column=10, end_row=20, end_column=11) #PB/COUP Flat
                    elif spec_data.il_exp_tf:  # Dual Band IL only Don't mearge if spec_data.il_exp_tf:  # Dual Band AB only Don't mearge AB cels
                        sheet.merge_cells(start_row=20, start_column=4, end_row=20, end_column=5) #RL
                        sheet.merge_cells(start_row=20, start_column=6, end_row=20, end_column=7)  #ISO/Coup
                        sheet.merge_cells(start_row=20, start_column=8, end_row=20, end_column=9) #AB/DIR
                        sheet.merge_cells(start_row=20, start_column=10, end_row=20, end_column=11) #PB/COUP Flat cels
                    elif spec_data.coup_exp_tf or spec_data.iso_exp_tf:  # Dual Band Coupling/Isolation only Don't mearge if spec_data.coup_exp_tf:  
                        sheet.merge_cells(start_row=20, start_column=2, end_row=20, end_column=3) #IL
                        sheet.merge_cells(start_row=20, start_column=4, end_row=20, end_column=5) #RL
                        sheet.merge_cells(start_row=20, start_column=8, end_row=20, end_column=9) #AB/DIR
                        sheet.merge_cells(start_row=20, start_column=10, end_row=20, end_column=11) #PB/COUP Flat    
                    elif spec_data.pb_exp_tf or spec_data.cf_exp_tf:  # Dual Band PB/Coup Flatness only Don't mearge if spec_data.pb_exp_tf:  
                        sheet.merge_cells(start_row=20, start_column=2, end_row=20, end_column=3) #IL
                        sheet.merge_cells(start_row=20, start_column=4, end_row=20, end_column=5) #RL
                        sheet.merge_cells(start_row=20, start_column=6, end_row=20, end_column=7)  #ISO/Coup
                        sheet.merge_cells(start_row=20, start_column=8, end_row=20, end_column=9) #AB/DIR
                    else:
                        sheet.merge_cells(start_row=20, start_column=2, end_row=20, end_column=3) #IL
                        sheet.merge_cells(start_row=20, start_column=4, end_row=20, end_column=5) #RL
                        sheet.merge_cells(start_row=20, start_column=6, end_row=20, end_column=7)  #ISO/Coup
                        sheet.merge_cells(start_row=20, start_column=8, end_row=20, end_column=9) #AB/DIR
                        sheet.merge_cells(start_row=20, start_column=10, end_row=20, end_column=11) #PB/COUP Flat    
                    #~~~~~~~~~~~~~~~~~~~~~format the spec cells for data or multiband data~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                    
                    thisRow = len(artwork_list) + 3 
                    # ~~~~~~~~~~~~~~~~~~~~~~~~~~Load the specs ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                    if spec_data.il_exp_tf:
                        sheet['B20'] = str(spec_data.insertionloss) 
                        sheet['C20'] = str(spec_data.il_ex)
                        sheet['B9'] = str(spec_data.insertionloss) + '/' + str(spec_data.il_ex) + ' Max'
                    else:
                        sheet['B20'] = str(spec_data.insertionloss) + ' Max' 
                        sheet['B9'] = str(spec_data.insertionloss) + ' Max'
                    sheet['D20'] = str(spec_rl) + ' Max'
                    sheet['D9'] = str(spec_rl) + ' Max'
                    if '90 DEGREE COUPLER' in spectype or 'BALUN' in spectype:
                        if spec_data.iso_exp_tf:
                            sheet['F20'] = "- " + str(spec_data.isolation) + ' Max' 
                            sheet['G20'] = str(spec_data.iso_ex) + ' Max'
                            sheet['F9'] = "- " + str(spec_data.isolation) + '/' + str(spec_data.iso_ex) + ' Max'
                        else:
                            sheet['F20'] = str(spec_data.isolation) + ' Max'
                            sheet['F9'] = "- " + str(spec_data.isolation) + ' Max'
                        
                        if spec_data.ab_exp_tf:
                            sheet['H20'] = "+/- " + str(spec_data.amplitudebalance) 
                            sheet['I20'] = str(spec_data.ab_ex) + ' dB'
                            sheet['H9'] = "+/- " + str(spec_data.amplitudebalance) + '/' + str(spec_data.ab_ex) + ' deg'
                        else:
                            sheet['H20'] = "+/- " + str(spec_data.amplitudebalance) + ' dB'
                            sheet['H9'] = "+/- " + str(spec_data.amplitudebalance) + ' dB'
                        
                        if spec_data.pb_exp_tf:
                            sheet['J20'] = "+/- " + str(spec_data.phasebalance) 
                            sheet['K20'] = str(spec_data.pb_ex) + ' dB'
                            sheet['J9'] = "+/- " + str(spec_data.phasebalance) + '/' + str(spec_data.pb_ex) + ' deg'
                        else:
                            sheet['J20'] = "+/- " + str(spec_data.phasebalance) + ' deg'
                            sheet['J9'] = "+/- " + str(spec_data.phasebalance) + ' deg'
                        
                    elif 'DIRECTIONAL COUPLER' in spectype:
                        if spec_data.coup_exp_tf:
                            sheet['F20'] = str(spec_data.coupling) + "+/- " + str(spec_data.coupplusminus) 
                            sheet['G20'] = str(spec_data.coup_ex) + "+/- " + str(spec_data.coupplusminus) 
                            sheet['F9'] =  str(spec_data.coupling) + "+/- " + str(spec_data.coupplusminus)  + '/' +  str(spec_data.coup_ex) + " +/- " + str(spec_data.coupplusminus) 
                        else:
                            sheet['F20'] = str(spec_data.coupling) + "+/- " + str(spec_data.coupplusminus) 
                            sheet['F9'] = str(spec_data.coupling) + "+/- " + str(spec_data.coupplusminus) 
                                                        
                        
                        if spec_data.dir_exp_tf:
                            sheet['H20'] = str(spec_data.directivity) + ' dB Min'
                            sheet['I20'] = str(spec_data.dir_ex) + '/' + str(spec_data.directivity) + ' dB Min'
                            sheet['H9'] =  str(spec_data.directivity) + '/' + str(spec_data.dir_ex) + ' dB Min'
                        else:
                            sheet['H20'] = str(spec_data.directivity) + ' dB Min'
                            sheet['H9'] =  str(spec_data.directivity) + ' dB Min'
                        
                        if spec_data.cf_exp_tf:
                            sheet['J20'] = "+/- " + str(spec_data.coupledflatness) 
                            sheet['K20'] = str(spec_data.cf_ex) + ' dB'
                            sheet['J9'] = "+/- " + str(spec_data.coupledflatness) + '/' + str(spec_data.pb_ex) + ' dB'
                        else:
                            sheet['J20'] = "+/- " + str(spec_data.coupledflatness) + ' dB'
                            sheet['J9'] = "+/- " + str(spec_data.coupledflatness) + ' dB'
                    
                    elif 'TRANSFORMER' in spectype:
                        sheet['F20'] = 'N/A'
                        sheet['H20'] = 'N/A'
                        sheet['J20'] = 'N/A'
                        sheet['B9'] = 'N/A'
                        sheet['D9'] = 'N/A'
                        sheet['F9'] = 'N/A'
                        sheet['H9'] = 'N/A'
                        sheet['J9'] = 'N/A'
                   
                   #Tabular data
                    rownum = 21
                    insertion_loss1 = []
                    return_loss1 = []
                    isolation1 = []
                    coupling1 = []
                    amplitude_balance1 = []
                    phase_balance1 = []
                    directivity1 = []
                    coupledflatness1 = []
                    insertion_loss2 = []
                    return_loss2 = []
                    isolation2 = []
                    coupling2 = []
                    amplitude_balance2 = []
                    phase_balance2 = []
                    directivity2 = []
                    coupledflatness2 = []
                    
                    stat_list = []
                    il_pass = 0
                    rl_pass = 0
                    iso_pass = 0
                    ab_pass = 0
                    pb_pass = 0
                    coup_pass = 0
                    dir_pass = 0
                    cf_pass = 0                
                    il_fail = 0
                    rl_fail = 0
                    iso_fail = 0
                    ab_fail = 0
                    pb_fail = 0
                    coup_fail = 0
                    dir_fail = 0
                    cf_fail = 0
                    total_pass = 0
                    uut = 1
                    spec1=0
                    spec2=0
                    spec3=0
                    spec4=0
                    spec5=0
                    
                    print('report_data=',report_data)
                    if spec_data.vswr:
                        conversions = Conversions(spec_data.vswr,'')
                        spec_rl = round(conversions.vswr_to_rl(),3)
                    else:
                        spec_rl = 0
                    print('spec_rl=',spec_rl)
                    print('spec_data.insertionloss=',spec_data.insertionloss)
                    spectype=spec_data.spectype
                    try:
                        if '90 DEGREE COUPLER' in spec_data.spectype or 'BALUN' in spec_data.spectype:
                            if spec_data.insertionloss:
                                spec1 = round(spec_data.insertionloss,3)
                            if spec_rl:
                                spec2 = spec_rl
                            if spec_data.isolation:
                                spec3 = round(spec_data.isolation,3)
                            if spec_data.amplitudebalance:
                                spec4 = round(spec_data.amplitudebalance,3)
                            if spec_data.phasebalance:
                                spec5 = round(spec_data.phasebalance,3)
                        elif 'DIRECTIONAL COUPLER' in spec_data.spectype: 
                            if spec_data.insertionloss:
                                spec1 = round(spec_data.insertionloss,3)
                            if spec_rl:
                                spec2 = spec_rl
                            if spec_data.coupling:
                                spec3 = round(spec_data.coupling,3)
                            if spec_data.directivity:
                                spec4 = round(spec_data.directivity,3)
                            if spec_data.coupledflatness:
                                spec5 = round(spec_data.coupledflatness,3)
                    except ValueError as e:
                        print('error = ',e) 
                    
                    for data in report_data:
                        good_data=True
                        #~~~~~~~~~~~~~~~Check for good data~~~~~~~~~~~~~~~~~
                        #print('IL&RL ',data.insertionloss,data.insertionloss)
                        if not data.insertionloss:
                            good_data=False
                        elif data.insertionloss > spec1 * 3:
                            good_data=False
                        
                        if not data.returnloss:
                            good_data=False
                        elif data.returnloss < spec2 * 3:
                            good_data=False
                        
                        if '90 DEGREE COUPLER' in spectype or 'BALUN' in spectype:
                            #print('ISo&AM&PB ',data.isolation,data.phasebalance)
                            if not data.isolation: 
                                good_data=False
                            elif abs(data.isolation) > spec3 * 3:
                                good_data=False
                            
                            if not data.phasebalance: 
                                good_data=False
                            elif abs(data.phasebalance) > spec3 * 3:
                                good_data=False
                            
                            if spec_data.ab_exp_tf :
                                if not data.amplitudebalance1:
                                    good_data=False
                                elif abs(data.amplitudebalance) > spec4 * 3:
                                    good_data=False
                            else:
                                if not data.amplitudebalance:
                                    good_data=False
                                    #print('no good')
                        else:
                            #print('coup&dir&cf ',data.coupling,data.directivity,data.coupledflatness)
                            if not data.coupling: 
                                good_data=False
                            elif abs(data.coupling) > spec3 * 3: 
                                good_data=False
                            
                            if not data.directivity: 
                                good_data=False
                            elif abs(data.directivity) > spec4 * 3: 
                                good_data=False
                            
                            if not data.coupledflatness: 
                                good_data=False
                            elif abs(data.coupledflatness) > spec5 * 3:
                                good_data=False
                        #~~~~~~~~~~~~~~~Check for good data~~~~~~~~~~~~~~~~~
                        #~~~~~~~~~~~~~~~~  Never mind ): ~~~~~~~~~~~~~~~~~~~
                        if not filter_bad_data:
                            good_data=True
                        #~~~~~~~~~~~~~~~~  Never mind ): ~~~~~~~~~~~~~~~~~~~
                        
                        #time.sleep(20)
                        if good_data:
                            sheet.cell(row=rownum, column=1).value= 'UUT ' + str(uut)
                            print('data.serialnumber=',data.serialnumber)
                            
                            ##~~~~~~~~~~~~~~~~~~~~~~~~IL Dual Band ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~  
                            if spec_data.il_exp_tf:
                                if data.insertionloss and data.insertionloss2:
                                    sheet.cell(row=rownum, column=2).value= round(data.insertionloss,2)
                                    sheet.cell(row=rownum, column=3).value= round(data.insertionloss2,2)
                                    testdata1a = sheet.cell(row=rownum, column=2)#Created a variable that contains cell
                                    testdata1b = sheet.cell(row=rownum, column=3)#Created a variable that contains cell
                                    insertionloss.append(data.insertionloss)
                                    insertionloss2.append(data.insertionloss2)
                                    if data.insertionloss <= spec_data.insertionloss and data.insertionloss2 <= spec_data.il2:
                                        il_pass+=1
                                    else:
                                        il_fail+=1
                                        if data.insertionloss > spec_data.insertionloss:
                                            testdata1a.font = Font(color='FF3342', bold=True, italic=True) #W
                                        if data.ainsertionloss2 > spec_data.il2:
                                            testdata1b.font = Font(color='FF3342', bold=True, italic=True) #W
                            ##~~~~~~~~~~~~~~~~~~~~~~~~IL Dual Band ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~  
                            else:
                                if data.insertionloss:
                                    sheet.cell(row=rownum, column=3).value= round(data.insertionloss,2)
                                    testdata1 = sheet.cell(row=rownum, column=3)#Created a variable that contains cell
                                    insertion_loss1.append(data.insertionloss)
                                    if data.insertionloss <= spec_data.insertionloss:
                                        il_pass+=1
                                    else:
                                        il_fail+=1
                                        testdata1.font = Font(color='FF3342', bold=True, italic=True) #W
                            
                            if data.returnloss:
                                sheet.cell(row=rownum, column=5).value= round(data.returnloss,2)
                                testdata2 = sheet.cell(row=rownum, column=5)#Created a variable that contains cell
                                return_loss1.append(data.returnloss)
                                if data.returnloss <= spec_rl:
                                    rl_pass+=1
                                else:
                                    rl_fail+=1
                                    testdata2.font = Font(color='FF3342', bold=True, italic=True) #W
                            
                            if '90 DEGREE COUPLER' in spectype or 'BALUN' in spectype:
                                if data.isolation:
                                    sheet.cell(row=rownum, column=7).value= round(data.isolation,2)
                                    testdata3 = sheet.cell(row=rownum, column=7)#Created a variable that contains cell
                                    isolation1.append(data.isolation)
                                    if data.isolation <= 0-spec_data.isolation:
                                        iso_pass+=1
                                    else:
                                        iso_fail+=1
                                        testdata3.font = Font(color='FF3342', bold=True, italic=True) #W
                                ##~~~~~~~~~~~~~~~~~~~~~~~~AB Dual Band ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~  
                                if spec_data.ab_exp_tf:
                                    if data.amplitudebalance1 and data.amplitudebalance2:
                                        sheet.cell(row=rownum, column=8).value= round(data.amplitudebalance1,2)
                                        sheet.cell(row=rownum, column=9).value= round(data.amplitudebalance2,2)
                                        testdata4a = sheet.cell(row=rownum, column=8)#Created a variable that contains cell
                                        testdata4b = sheet.cell(row=rownum, column=9)#Created a variable that contains cell
                                        amplitude_balance1.append(data.amplitudebalance)
                                        amplitude_balance2.append(data.amplitudebalance2)
                                        if data.amplitudebalance1 <= spec_list[3] and data.amplitudebalance2 <= spec_list[5]:
                                            ab_pass+=1
                                        else:
                                            ab_fail+=1
                                            if data.amplitudebalance1 > spec_list[3]:
                                                testdata4a.font = Font(color='FF3342', bold=True, italic=True) #W
                                            if data.amplitudebalance2 > spec_list[5]:
                                                testdata4b.font = Font(color='FF3342', bold=True, italic=True) #W
                                ##~~~~~~~~~~~~~~~~~~~~~~~~AB Dual Band ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~  
                                else:
                                    if data.amplitudebalance:
                                        sheet.cell(row=rownum, column=9).value= round(data.amplitudebalance,2)
                                        testdata4 = sheet.cell(row=rownum, column=9)#Created a variable that contains cell
                                        amplitude_balance1.append(data.amplitudebalance)
                                        if data.amplitudebalance <= spec_list[3]:
                                            ab_pass+=1
                                        else:
                                            ab_fail+=1
                                            testdata4.font = Font(color='FF3342', bold=True, italic=True) #W

                                if spec_data.ab_exp_tf:
                                    if data.phasebalance and data.phasebalance2:
                                        sheet.cell(row=rownum, column=10).value= round(data.phasebalance,2)
                                        sheet.cell(row=rownum, column=11).value= round(data.phasebalance2,2)
                                        testdata5a = sheet.cell(row=rownum, column=10)#Created a variable that contains cell
                                        testdata5b = sheet.cell(row=rownum, column=11)#Created a variable that contains cell
                                        amplitude_balance1.append(data.amplitudebalance)
                                        amplitude_balance2.append(data.phasebalance2)
                                        if data.phasebalance <= spec_list[3] and data.phasebalance2 <= spec_data.pb_ex:
                                            ab_pass+=1
                                        else:
                                            ab_fail+=1
                                            if data.phasebalance > spec_list[4]:
                                                testdata5a.font = Font(color='FF3342', bold=True, italic=True) #W
                                            if data.phasebalance2> spec_list[5]:
                                                testdata5b.font = Font(color='FF3342', bold=True, italic=True) #W
                                else:
                                    if data.phasebalance:
                                        sheet.cell(row=rownum, column=11).value= round(data.phasebalance,2)
                                        testdata5 = sheet.cell(row=rownum, column=11)#Created a variable that contains cell
                                        phase_balance1.append(data.phasebalance)
                                        if data.phasebalance <= spec_list[4]:
                                            pb_pass+=1
                                        else:
                                            pb_fail+=1
                                            testdata5.font = Font(color='FF3342', bold=True, italic=True) #W
                            else:
                                ##~~~~~~~~~~~~~~~~~~~~~~~~COUP Dual Band ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~  
                                if spec_data.coup_exp_tf:
                                    if data.coupling and data.coupling2:
                                        sheet.cell(row=rownum, column=6).value= round(data.coupling,2)
                                        sheet.cell(row=rownum, column=7).value= round(data.coupling2,2)
                                        testdata3a = sheet.cell(row=rownum, column=6)#Created a variable that contains cell
                                        testdata3b = sheet.cell(row=rownum, column=7)#Created a variable that contains cell
                                        coupling1.append(data.coupling)
                                        coupling2.append(data.coupling2)
                                        if data.coupling <= spec_data.coupling + spec_data.coupplusminus and data.coupling >= spec_data.coupling - spec_data.coupplusminus and data.coupling2 <= spec_data2.coupling + spec_data.coupplusminus and data.coupling2 >= spec_data2.coupling - spec_data.coup_ex:
                                            coup_pass+=1
                                        else:
                                            coup_fail+=1
                                            if data.coupling <= spec_data.coupling + spec_data.coupplusminus and data.coupling >= spec_data.coupling - spec_data.coupplusminus:
                                                testdata3a.font = Font(color='FF3342', bold=True, italic=True) #W
                                            if data.coupling2 <= spec_data2.coupling + spec_data.coupplusminus and data.coupling2 >= spec_data2.coupling - spec_data.coupplusminus:
                                                testdata3b.font = Font(color='FF3342', bold=True, italic=True) #W
                                ##~~~~~~~~~~~~~~~~~~~~~~~~COUP Dual Band ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~  
                                else:
                                    if data.coupling:
                                        sheet.cell(row=rownum, column=7).value= round(data.coupling,2)
                                        testdata3 = sheet.cell(row=rownum, column=7)#Created a variable that contains cell
                                        coupling1.append(data.coupling)
                                        if data.coupling <= spec_data.coupling + spec_data.coupplusminus and data.coupling >= spec_data.coupling - spec_data.coupplusminus:
                                            coup_pass+=1
                                        else:
                                            coup_fail+=1
                                            testdata3 = sheet.cell(row=rownum, column=7)#Created a variable that contains cell
                                            testdata3.font = Font(color='FF3342', bold=True, italic=True) #W
                                
                                ##~~~~~~~~~~~~~~~~~~~~~~~~directivity Dual Band ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~  
                                if spec_data.dir_exp_tf:
                                    if data.directivity and data.directivity2:
                                        sheet.cell(row=rownum, column=8).value= round(data.directivity,2)
                                        sheet.cell(row=rownum, column=9).value= round(data.directivity2,2)
                                        testdata4a = sheet.cell(row=rownum, column=8)#Created a variable that contains cell
                                        testdata4b = sheet.cell(row=rownum, column=9)#Created a variable that contains cell
                                        directivity1.append(data.directivity)
                                        directivity2.append(data.directivity2)
                                        if data.directivity >= spec_list[3] and data.directivity2 >= spec_data.dir_ex:
                                            coup_pass+=1
                                        else:
                                            coup_fail+=1
                                            if data.directivity >= spec_list[3]:
                                                testdata4a.font = Font(color='FF3342', bold=True, italic=True) #W
                                            if data.directivity2 >= spec_data.dir_ex:
                                                testdata4b.font = Font(color='FF3342', bold=True, italic=True) #W
                                ##~~~~~~~~~~~~~~~~~~~~~~~~COUP Dual Band ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~  
                                else:
                                    if data.directivity:
                                        sheet.cell(row=rownum, column=9).value= round(data.directivity,2)
                                        testdata4 = sheet.cell(row=rownum, column=9)#Created a variable that contains cell
                                        directivity1.append(data.directivity)
                                        if data.directivity >= spec_list[3]:
                                            dir_pass+=1
                                        else:
                                            dir_fail+=1
                                            testdata4.font = Font(color='FF3342', bold=True, italic=True) #W
                                
                                ##~~~~~~~~~~~~~~~~~~~~~~~~CF Dual Band ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~  
                                if spec_data.cf_exp_tf:
                                    if data.coupledflatness and data.coupledflatness2:
                                        sheet.cell(row=rownum, column=10).value= round(data.coupledflatness,2)
                                        sheet.cell(row=rownum, column=11).value= round(data.coupledflatness2,2)
                                        testdata5a = sheet.cell(row=rownum, column=10)#Created a variable that contains cell
                                        testdata5b = sheet.cell(row=rownum, column=11)#Created a variable that contains cell
                                        coupledflatness1.append(data.coupledflatness)
                                        coupledflatness2.append(data.coupledflatness2)
                                        if data.coupledflatness <= spec_data.coupledflatness and data.coupledflatness2 <= spec_data.coupledflatness2:
                                            cf_pass+=1
                                        else:
                                            cf_fail+=1
                                            if data.coupledflatness > spec_data.coupledflatness:
                                                testdata5a.font = Font(color='FF3342', bold=True, italic=True) #W
                                            if coupledflatness2 > spec_data.coupledflatness2:
                                                testdata5b.font = Font(color='FF3342', bold=True, italic=True) #W
                                ##~~~~~~~~~~~~~~~~~~~~~~~~AB Dual Band ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~  
                                else:    
                                    if data.coupledflatness:
                                        sheet.cell(row=rownum, column=11).value= round(data.coupledflatness,2)
                                        testdata5 = sheet.cell(row=rownum, column=11)#Created a variable that contains cell
                                        coupledflatness1.append(data.coupledflatness)
                                        if data.coupledflatness <= spec_list[4]:
                                            cf_pass+=1
                                        else:
                                            cf_fail+=1
                                            testdata5.font = Font(color='FF3342', bold=True, italic=True) #W
                            rownum+=1
                            #print('rownum=',rownum)
                        uut+=1
                   
                    uut=uut-1    
                    #~~~~~~~~~~~~~~~~Statics and Summary ~~~~~~~~~~~~~~~~~~~~
                    il_list=[]
                    rl_list=[]
                    ab_list=[]
                    iso_list=[]
                    pb_list=[]
                    coup_list=[]
                    dir_list=[]
                    cf_list=[]
                    if len(insertion_loss1) > 1:# must have at least two tests
                        list_names = ['Min','Max','Avg','Stdev']
                        #print('insertion_loss1=',insertion_loss1)
                        il_stdev = round(statistics.stdev(insertion_loss1),2) #Standard deviation
                        il_var = round(statistics.variance(insertion_loss1),2) #Variance
                        il_avg = round(statistics.mean(insertion_loss1),2) #Mean Average
                        il_min = round(min(insertion_loss1),2) #Min
                        il_max = round(max(insertion_loss1),2) #Max
                        sheet['B10'] = il_avg
                        sheet['B11'] = il_min
                        sheet['B12'] = il_max
                        sheet['B13'] = il_stdev
                        sheet['B14'] = il_pass
                        sheet['B15'] = il_fail
                        print('uut=',uut)
                        print('rownum=',rownum)
                        sheet['B16'] = str(round((il_fail/uut)*100,2)) + '%'
                        il_list = [il_min,il_max,il_avg,il_stdev]
                        #print('il_list=',il_list)

                        #print('return_loss1=',return_loss1)
                        
                        if len(return_loss1)>1:
                            rl_stdev = round(statistics.stdev(return_loss1),2) #Standard deviation
                            rl_var = round(statistics.variance(return_loss1),2) #Variance
                            rl_avg = round(statistics.mean(return_loss1),2) #Mean Average
                            rl_min = round(min(return_loss1),2) #Min
                            rl_max = round(max(return_loss1),2) #Max
                            rl_list = [rl_min,rl_max,rl_avg,rl_stdev]
                            #print('rl_list=',rl_list)
                        else:
                            rl_var = 'N/A'
                            rl_stdev = 'N/A'
                            rl_avg  = 'N/A'
                            rl_min  = 'N/A'
                            rl_max  = 'N/A'
                            rl_list = [rl_min,rl_max,rl_avg,rl_stdev]

                        sheet['D10'] = rl_avg
                        sheet['D11'] = rl_min
                        sheet['D12'] = rl_max
                        sheet['D13'] = rl_stdev
                        sheet['D14'] = rl_pass
                        sheet['D15'] = rl_fail
                        print('uut=',uut)
                        print('rownum=',rownum)
                        sheet['D16'] = str(round((rl_fail/uut)*100,2)) + '%'
                            
                        if '90 DEGREE COUPLER' in spectype or 'BALUN' in spectype:
                            #print('isolation1=',isolation1)
                            #print('amplitude_balance1=',amplitude_balance1)
                            #print('phase_balance1=',phase_balance1)
                            
                            if len(isolation1)>1:
                                iso_stdev = round(statistics.stdev(isolation1),2) #Standard deviation
                                iso_var = round(statistics.variance(isolation1),2) #Variance
                                iso_avg = round(statistics.mean(isolation1),2) #Mean Average
                                iso_min = round(min(isolation1),2) #Min
                                iso_max = round(max(isolation1),2) #Max
                                iso_list = [iso_min,iso_max,iso_avg,iso_stdev]
                                #print('iso_list=',iso_list)
                            else:
                                iso_var = 'N/A'
                                iso_stdev = 'N/A'
                                iso_avg  = 'N/A'
                                iso_min  = 'N/A'
                                iso_max  = 'N/A'
                                iso_list = [iso_min,iso_max,iso_avg,iso_stdev]
                                
                            sheet['F8'] = 'ISO'
                            sheet['F10'] = iso_avg
                            sheet['F11'] = iso_min
                            sheet['F12'] = iso_max
                            sheet['F13'] = iso_stdev
                            sheet['F14'] = iso_pass
                            sheet['F15'] = iso_fail
                            print('uut=',uut)
                            print('rownum=',rownum)
                            sheet['F16'] = str(round((iso_fail/uut)*100,2)) + '%'
                                
                            if len(amplitude_balance1)>1:
                                ab_var = round(statistics.variance(amplitude_balance1),2) #Variance
                                ab_stdev = round(statistics.stdev(amplitude_balance1),2) #Standard deviation
                                ab_avg = round(statistics.mean(amplitude_balance1),2) #Mean Average
                                ab_min = round(min(amplitude_balance1),2) #Min
                                ab_max = round(max(amplitude_balance1),2) #Max
                                ab_list = [ab_min,ab_max,ab_avg,ab_stdev]
                            else:
                                ab_var = 'N/A'
                                ab_stdev = 'N/A'
                                ab_avg  = 'N/A'
                                ab_min  = 'N/A'
                                ab_max  = 'N/A'
                                ab_list = [ab_list_min,ab_list_max,ab_list_avg,ab_list_stdev]
                                
                            sheet['H8'] = 'AB'
                            sheet['H10'] = ab_avg
                            sheet['H11'] = ab_min
                            sheet['H12'] = ab_max
                            sheet['H13'] = ab_stdev
                            sheet['H14'] = ab_pass
                            sheet['H15'] = ab_fail
                            print('uut=',uut)
                            print('rownum=',rownum)
                            sheet['H16'] = str(round((ab_fail/uut)*100,2)) + '%'
                            #print('ab_list=',ab_list)

                            if len(phase_balance1)>1:
                                pb_stdev = round(statistics.stdev(phase_balance1),2) #Standard deviation
                                pb_var = round(statistics.variance(phase_balance1),2) #Variance
                                pb_avg = round(statistics.mean(phase_balance1),2) #Mean Average
                                pb_min = round(min(phase_balance1),2) #Min
                                pb_max = round(max(phase_balance1),2) #Max
                                pb_list = [pb_min,pb_max,pb_avg,pb_stdev]
                                #print('pb_list=',pb_list)
                            else:
                                pb_var = 'N/A'
                                pb_stdev = 'N/A'
                                pb_avg  = 'N/A'
                                pb_min  = 'N/A'
                                pb_max  = 'N/A'
                                pb_list = [pb_min,pb_max,pb_avg,pb_stdev]
                                
                            sheet['J8'] = 'PB'
                            sheet['J10'] = pb_avg
                            sheet['J11'] = pb_min
                            sheet['J12'] = pb_max
                            sheet['J13'] = pb_stdev
                            sheet['J14'] = pb_pass
                            sheet['J15'] = pb_fail
                            print('uut=',uut)
                            print('rownum=',rownum)
                            sheet['J16'] = str(round((pb_fail/uut)*100,2)) + '%'
                            stat_list = [il_list,rl_list,iso_list,ab_list,pb_list]
                        else:
                            #print('coupling1=',coupling1)
                            #print('directivity1=',directivity1)
                            #print('coupledflatness1=',coupledflatness1)
                            if len(coupling1)>1:
                                coup_stdev = round(statistics.stdev(coupling1),2) #Standard deviation
                                coup_var = round(statistics.variance(coupling1),2) #Variance
                                coup_avg = round(statistics.mean(coupling1),2) #Mean Average
                                coup_min = round(min(coupling1),2) #Min
                                coup_max = round(max(coupling1),2) #Max
                                coup_list = [coup_min,coup_max,coup_avg,coup_stdev]
                                
                                #print('iso_list=',iso_list)
                            else:
                                coup_var = 'N/A'
                                coup = 'N/A'
                                coup_avg  = 'N/A'
                                coup_min  = 'N/A'
                                coup_max  = 'N/A'
                                coup_list = [coup_min,coup_max,coup_avg,coup_stdev]
                            
                            sheet['F8'] = 'COUP'
                            sheet['F10'] = coup_avg
                            sheet['F11'] = coup_min
                            sheet['F12'] = coup_max
                            sheet['F13'] = coup_stdev
                            sheet['F14'] = coup_pass
                            sheet['F15'] = coup_fail
                            print('uut=',uut)
                            print('rownum=',rownum)
                            sheet['F16'] = str(round((coup_fail/uut)*100,2)) + '%'
                            
                            if len(directivity1)>1:
                                dir_stdev = round(statistics.stdev(directivity1),2) #Standard deviation
                                dir_var = round(statistics.variance(directivity1),2) #Variance
                                dir_avg = round(statistics.mean(directivity1),2) #Mean Average
                                dir_min = round(min(directivity1),2) #Min
                                dir_max = round(max(directivity1),2) #Max
                                dir_list = [dir_min,dir_max,dir_avg,dir_stdev]
                                
                                #print('ab_list=',ab_list)
                            else:
                                dir_var = 'N/A'
                                dir_stdev = 'N/A'
                                dir_avg  = 'N/A'
                                dir_min  = 'N/A'
                                dir_max  = 'N/A'
                                dir_list = [dir_min,dir_max,dir_avg,dir_stdev]

                            
                            sheet['H8'] = 'DIR'
                            sheet['H10'] = dir_avg
                            sheet['H11'] = dir_min
                            sheet['H12'] = dir_max
                            sheet['H13'] = dir_stdev
                            sheet['H14'] = dir_pass
                            sheet['H15'] = dir_fail
                            sheet['H17'] = uut
                            print('uut=',uut)
                            print('rownum=',rownum)
                            sheet['H16'] = str(round((dir_fail/uut)*100,2)) + '%'
                            if len(coupledflatness1)>1:
                                cf_var = round(statistics.variance(coupledflatness1),2) #Variance
                                cf_stdev = round(statistics.stdev(coupledflatness1),2) #Standard deviation
                                cf_avg = round(statistics.mean(coupledflatness1),2) #Mean Average
                                cf_min = round(min(coupledflatness1),2) #Min
                                cf_max = round(max(coupledflatness1),2) #Max
                                cf_list = [cf_min,cf_max,cf_avg,cf_stdev]
                                #print('pb_list=',pb_list)
                            else:
                                cf_var = 'N/A'
                                cf_stdev = 'N/A'
                                cf_avg  = 'N/A'
                                cf_min  = 'N/A'
                                cf_max  = 'N/A'
                                cf_list = [cf_min,cf_max,cf_avg,cf_stdev]
                            
                            sheet['J8'] = 'CF'
                            sheet['J10'] = cf_avg
                            sheet['J11'] = cf_min
                            sheet['J12'] = cf_max
                            sheet['J13'] = cf_stdev
                            sheet['J14'] = cf_pass
                            sheet['J15'] = cf_fail
                            sheet['J16'] = str(round((cf_fail/rownum+1)*100,2)) + '%'
                                
                            stat_list = [il_list,rl_list,coup_list,dir_list,cf_list]

                        
                        #print('stat_list=',stat_list)
                        sheet.title = artwork_rev
                        
                        #~~~~~~~~~~~~~~~~~~~~~~Summary sheet~~~~~~~~~~~~~~~~~~~~~~~~
                        sheet = wb["Summary"]
                        #print('sheet=',sheet)
                        print('sum_row=',sum_row)
                        if sum_row==5:
                            thisRow = len(artwork_list) + 3
                            print('thisRow=',thisRow)
                        
                        
                        if '90 DEGREE COUPLER' in spectype or 'BALUN' in spectype:
                            #AVG
                            sheet['A' + str(sum_row)] = artwork_rev
                            if sum_row==5:
                                sheet['B' + str(sum_row-1)] = str(spec_list[0]) + ' Max'
                                sheet['C' + str(sum_row-1)] = str(spec_list[1]) + ' Max'
                                sheet['D' + str(sum_row-1)] = str(spec_list[2]) + ' Max'
                                sheet['E' + str(sum_row-1)] = "'+/- " + str(spec_list[3]) + ' dB'
                                sheet['F' + str(sum_row-1)] = "'+/- " + str(spec_list[4]) + ' deg'
                            sheet['B' + str(sum_row)] = il_avg
                            sheet['C' + str(sum_row)] = rl_avg
                            sheet['D' + str(sum_row)] = iso_avg
                            sheet['E' + str(sum_row)] = ab_avg
                            sheet['F' + str(sum_row)] = pb_avg
                            
                            sheet['G' + str(sum_row)] = (il_pass + rl_pass + iso_pass + ab_pass + pb_pass)/5 + (il_fail + rl_fail + iso_fail + ab_fail + pb_fail)/5
                            sheet['H' + str(sum_row)] = (il_pass + rl_pass + iso_pass + ab_pass + pb_pass)/5
                            sheet['I' + str(sum_row)] = il_fail + rl_fail + iso_fail + ab_fail + pb_fail
                            
                         
                            #MIN
                            sheet['A' + str(sum_row + thisRow)] = artwork_rev
                            if sum_row==5:
                                sheet['B' + str(sum_row + thisRow-1)] = spec_list[0]  = str(spec_list[0]) + ' Max'
                                sheet['C' + str(sum_row + thisRow-1)] = str(spec_list[1]) + ' Max'
                                sheet['D' + str(sum_row + thisRow-1)] = str(spec_list[2]) + ' Max'
                                sheet['E' + str(sum_row + thisRow-1)] = "+/- " + str(spec_list[3]) + ' dB'
                                sheet['F' + str(sum_row + thisRow-1)] = "+/- " + str(spec_list[4]) + ' deg'
                            sheet['B' + str(sum_row + thisRow)] = il_min
                            sheet['C' + str(sum_row + thisRow)] = rl_min
                            sheet['D' + str(sum_row + thisRow)] = iso_min
                            sheet['E' + str(sum_row + thisRow)] = ab_min
                            sheet['F' + str(sum_row + thisRow)] = pb_min
                            sheet['G' + str(sum_row + thisRow)] = int((il_pass + rl_pass + iso_pass + ab_pass + pb_pass)/5) + int((il_fail + rl_fail + iso_fail + ab_fail + pb_fail)/5)
                            sheet['H' + str(sum_row + thisRow)] = int((il_pass + rl_pass + iso_pass + ab_pass + pb_pass)/5)  
                            sheet['I' + str(sum_row + thisRow)] = il_fail + rl_fail + iso_fail + ab_fail + pb_fail
                            
                            #Max
                            sheet['A' + str(sum_row + 2*thisRow)] = artwork_rev
                            if sum_row==5:
                                sheet['B' + str(sum_row + 2*thisRow-1)] = str(spec_list[0]) + ' Max'
                                sheet['C' + str(sum_row + 2*thisRow-1)] = str(spec_list[1]) + ' Max'
                                sheet['D' + str(sum_row + 2*thisRow-1)] = str(spec_list[2]) + ' Max'
                                sheet['E' + str(sum_row + 2*thisRow-1)] = "+/- " + str(spec_list[3]) + ' dB'
                                sheet['F' + str(sum_row + 2*thisRow-1)] = "+/- " + str(spec_list[4]) + ' deg'
                            sheet['B' + str(sum_row + 2*thisRow)] = il_max
                            sheet['C' + str(sum_row + 2*thisRow)] = rl_max
                            sheet['D' + str(sum_row + 2*thisRow)] = iso_max
                            sheet['E' + str(sum_row + 2*thisRow)] = ab_max
                            sheet['F' + str(sum_row + 2*thisRow)] = pb_max
                            sheet['G' + str(sum_row + 2*thisRow)] = int((il_pass + rl_pass + iso_pass + ab_pass + pb_pass)/5) + int((il_fail + rl_fail + iso_fail + ab_fail + pb_fail)/5)
                            sheet['H' + str(sum_row + 2*thisRow)] = int((il_pass + rl_pass + iso_pass + ab_pass + pb_pass)/5) 
                            sheet['I' + str(sum_row + 2*thisRow)] = il_fail + rl_fail + iso_fail + ab_fail + pb_fail
                        else:
                                                    #AVG
                            sheet['A' + str(sum_row)] = artwork_rev
                            if sum_row==5:
                                sheet['B' + str(sum_row-1)] = str(spec_list[0]) + ' Max'
                                sheet['C' + str(sum_row-1)] = str(spec_list[1]) + ' Max'
                                sheet['D' + str(sum_row-2)] = 'Avg.Coupling (dB)'
                                sheet['D' + str(sum_row-1)] = str(spec_list[2]) + '+/-' + str(spec_data.coupplusminus)
                                sheet['E' + str(sum_row-2)] = 'Avg.Directivity (dB)'
                                sheet['E' + str(sum_row-1)] = str(spec_list[3]) + ' dB Min'
                                sheet['F' + str(sum_row-2)] = 'Avg.Coup Flat(dB)'
                                sheet['F' + str(sum_row-1)] = "'+/- " + str(spec_list[4]) + ' dB'
                            sheet['B' + str(sum_row)] = il_avg
                            sheet['C' + str(sum_row)] = rl_avg
                            sheet['D' + str(sum_row)] = coup_avg
                            sheet['E' + str(sum_row)] = dir_avg
                            sheet['F' + str(sum_row)] = cf_avg
                            sheet['G' + str(sum_row)] = int((il_pass + rl_pass + coup_pass + dir_pass + cf_pass)/5) + int((il_fail + rl_fail + coup_fail + dir_fail + cf_fail)/5)
                            sheet['H' + str(sum_row)] = int((il_pass + rl_pass + coup_pass + dir_pass + cf_pass)/5) 
                            sheet['I' + str(sum_row)] = il_fail + rl_fail + coup_fail + dir_fail + cf_fail
                            
                            #MIN
                            sheet['A' + str(sum_row + thisRow)] = artwork_rev
                            if sum_row==5:
                                sheet['B' + str(sum_row + thisRow-1)] = spec_list[0]  = str(spec_list[0]) + ' Max'
                                sheet['C' + str(sum_row + thisRow-1)] = str(spec_list[1]) + ' Max'
                                sheet['D' + str(sum_row + thisRow-2)] = 'Min.Coupling (dB)'
                                sheet['D' + str(sum_row + thisRow-1)] = str(spec_list[2]) + '+/-' + str(spec_data.coupplusminus)
                                sheet['E' + str(sum_row + thisRow-2)] = 'Min.Directivity (dB)'
                                sheet['E' + str(sum_row + thisRow-1)] = str(spec_list[3]) + ' dB Min'
                                sheet['F' + str(sum_row + thisRow-2)] = 'Min.Coup Flat(dB)'
                                sheet['F' + str(sum_row + thisRow-1)] = "'+/- " + str(spec_list[4]) + ' dB'
                            sheet['B' + str(sum_row + thisRow)] = il_min
                            sheet['C' + str(sum_row + thisRow)] = rl_min
                            sheet['D' + str(sum_row + thisRow)] = coup_min
                            sheet['E' + str(sum_row + thisRow)] = dir_min
                            sheet['F' + str(sum_row + thisRow)] = cf_min
                            sheet['G' + str(sum_row + thisRow)] = int((il_pass + rl_pass + coup_pass + dir_pass + cf_pass)/5) + int((il_fail + rl_fail + coup_fail + dir_fail + cf_fail)/5)
                            sheet['H' + str(sum_row + thisRow)] = int((il_pass + rl_pass + coup_pass + dir_pass + cf_pass)/5) 
                            sheet['I' + str(sum_row + thisRow)] = il_fail + rl_fail + coup_fail + dir_fail + cf_fail
                            
                            #Max
                            sheet['A' + str(sum_row + 2*thisRow)] = artwork_rev
                            if sum_row==5:
                                sheet['B' + str(sum_row + 2*thisRow-1)] = str(spec_list[0]) + ' Max'
                                sheet['C' + str(sum_row + 2*thisRow-1)] = str(spec_list[1]) + ' Max'
                                sheet['D' + str(sum_row + 2*thisRow-2)] = 'Max.Coupling (dB)'
                                sheet['D' + str(sum_row + 2*thisRow-1)] = str(spec_list[2]) + '+/-' + str(spec_data.coupplusminus)
                                sheet['E' + str(sum_row + 2*thisRow-2)] = 'Max.Directivity (dB)'
                                sheet['E' + str(sum_row + 2*thisRow-1)] = str(spec_list[3]) + ' dB Min'
                                sheet['F' + str(sum_row + 2*thisRow-2)] = 'Max.Coup Flat(dB)'
                                sheet['F' + str(sum_row + 2*thisRow-1)] = "'+/- " + str(spec_list[4]) + ' dB'
                            sheet['B' + str(sum_row + 2*thisRow)] = il_max
                            sheet['C' + str(sum_row + 2*thisRow)] = rl_max
                            sheet['D' + str(sum_row + 2*thisRow)] = coup_max
                            sheet['E' + str(sum_row + 2*thisRow)] = dir_max
                            sheet['F' + str(sum_row + 2*thisRow)] = cf_max
                            sheet['G' + str(sum_row + 2*thisRow)] = int((il_pass + rl_pass + coup_pass + dir_pass + cf_pass)/5) + int((il_fail + rl_fail + coup_fail + dir_fail + cf_fail)/5)
                            sheet['H' + str(sum_row + 2*thisRow)] = int((il_pass + rl_pass + coup_pass + dir_pass + cf_pass)/5)
                            sheet['I' + str(sum_row + 2*thisRow)] = il_fail + rl_fail + coup_fail + dir_fail + cf_fail

                        #~~~~~~~~~~~~~~~~~~~~~~Summary sheet~~~~~~~~~~~~~~~~~~~~~~~~
                        #rename the sheet to the artworkrev
                        x+=1
                sum_row=sum_row+1
                print('###########################!!!!!!!!!sum_row=',sum_row,'$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$%%%%%%%%%%%%')
                z+=1  
                #~~~~~~~~~~~~~~~~~~~~~~~~~charts~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                ReportQueue.objects.using('TEST').filter(jobnumber=self.job_num).filter(workstation=self.workstation).update(reportstatus='loading charts')
                trace_num = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title='Insertion Loss J3').count()
                loadcharts=LoadCharts(self.job_num,part_num,spectype,wb,artwork_rev,len(artwork_list))
                loadcharts.charts()
                print('Charts Loaded')
        
            #~~~~~~~~~~~~~~~~~~~~~~~~~Save~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
            print('saving',self.job_num,part_num,spectype,self.operator,self.workstation,wb)
        try:
            savenow = SaveReports(self.job_num,part_num,spectype,self.operator,self.workstation,wb)
            print('savenow=',savenow)
            savenow.save()
            ReportQueue.objects.using('TEST').filter(jobnumber=self.job_num).filter(workstation=self.workstation).update(reportstatus='report complete')
            print("Report for ",self.job_num, " is complete")
        except BaseException as err:
            print('save error=',err) 
            try:            
               print("Report for ",self.job_num, " has failed at save",err)
               ReportQueue.objects.using('TEST').filter(jobnumber=self.job_num).filter(workstation=self.workstation).update(failurestatus='report save failure mes: ' + err)
            except BaseException as err:
               print('save error and report error=',err) 


class ReportFiles:
    def __init__ (self, job_num,part_num,spec_type):
        self.job_num = job_num
        self.part_num = part_num
        self.spec_type = spec_type
    
    def data_path(self):
        top_folder = "\\\ippdc\\Test Automation\\Test Data\\"
        report_path = "90_Degree\\"
        if '90 DEGREE COUPLER SMD' in self.spec_type:
            report_path = '90_Degree_SMD\\'
        elif '90 DEGREE COUPLER' in self.spec_type:
            report_path = "90_Degree\\"
        elif 'BALUN' in self.spec_type:
            report_path = "Balun\\"
        elif 'DIRECTIONAL COUPLER SMD' in self.spec_type:
            report_path = "Directional_Couplers_SMD\\"
        elif 'DIRECTIONAL COUPLER' in self.spec_type:
            report_path = "Directional_Couplers\\"
        elif 'COMBINER/DIVIDER SMD' in self.spec_type:
            report_path = "Combiner-Divider_SMD\\"
        elif 'COMBINER/DIVIDER' in self.spec_type:
            report_path = "Combiner-Divider\\"
        elif 'TRANSFORMER' in self.spec_type:
            report_path = "Transformer\\"
        
        #Create the path if it doesn't exist
        new_path = top_folder + report_path 
        #print('new_path=',new_path)
        if not os.path.exists(new_path):
            os.makedirs(new_path)
        return new_path
    
    def template(self):
        top_folder = "\\\ippdc\\Test Automation\\Excel_Templates\\"
        template = "90DEGREE_STANDARD.xlsx"
        if '90 DEGREE COUPLER SMD' in self.spec_type:
            template = '90DEGREE_STANDARD.xlsx'
        elif '90 DEGREE COUPLER' in self.spec_type:
            template = "90DEGREE_STANDARD.xlsx"
        elif 'BALUN' in self.spec_type:
            template = "90DEGREE_STANDARD.xlsx"
        elif 'DIRECTIONAL COUPLER SMD' in self.spec_type:
            template = "90DEGREE_STANDARD.xlsx"
        elif 'DIRECTIONAL COUPLER' in self.spec_type:
            template = "90DEGREE_STANDARD.xlsx"
        elif 'COMBINER/DIVIDER SMD' in self.spec_type:
            template = "90DEGREE_STANDARD.xlsx"
        elif 'COMBINER/DIVIDER' in self.spec_type:
            template = "90DEGREE_STANDARD.xlsx"
     
        new_path = top_folder + template 
        return new_path
        
class SaveReports:
    def __init__ (self, job_num,part_num,spec_type,operator,workstation,workbook):
        self.job_num = job_num
        self.part_num = part_num
        self.spec_type = spec_type
        self.operator = operator
        self.workstation = workstation
        self.workbook = workbook
        
    def save(self):
        paths = ReportFiles(self.job_num,self.part_num,self.spec_type)
        data_path = paths.data_path()
        x=1
        new_name = self.part_num + " " + self.job_num
        file_exists=os.path.isfile(data_path + new_name + ".xlsx")
        print('file_exists=',file_exists)
        
        if not file_exists :
             self.workbook.save(data_path + new_name + ".xlsx")
        else:
            while file_exists:
                new_name = new_name + '_' + str(x)
                file_exists=os.path.isfile(data_path + new_name + ".xlsx")
                if not file_exists :
                    self.workbook.save(data_path + new_name + ".xlsx")
        print('saving ',data_path + new_name + ".xlsx")
        ReportQueue.objects.using('TEST').filter(reportstatus='in process').filter(jobnumber = self.job_num).filter(partnumber=self.part_num).filter(operator=self.operator).filter(workstation=self.workstation).update(reportstatus='complete')




'''
rows = [
    ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
    [date(2015,9, 1), 40, 30, 25],
    [date(2015,9, 2), 40, 25, 30],
    [date(2015,9, 3), 50, 30, 45],
    [date(2015,9, 4), 30, 25, 40],
    [date(2015,9, 5), 25, 35, 30],
    [date(2015,9, 6), 20, 40, 35],
]
'''
class MakeCharts:
    def __init__ (self,sheet,spec_type,chartdata):
        self.sheet = sheet
        self.spec_type = spec_type
        self.chartdata = chartdata
        #print('loading Makecharts')
        #print('chartdata=',chartdata)
    
    def chart1(self):
        #print('self.spec_type=',self.spec_type)
        chart1 = ScatterChart()
        chart1.style = 13
        chart1.y_axis.title = 'dB'
        chart1.x_axis.title = 'Frequency MHz'
        chart1.x_axis.tickLblPos = "low"
        if not self.chartdata:
            return 0
        try:
            chart1.x_axis.scaling.min = min([sublist[1] for sublist in self.chartdata])
            chart1.x_axis.scaling.max = max([sublist[1] for sublist in self.chartdata])
        except IndexError as e:
            return 0
                
        chart1.legend = None
        chart1.width = 13
        chart1.height = 10
        
        print('xdata=',self.chartdata[0][1])
        print('TIdata=',self.chartdata[0][2])
        
        if '90 DEGREE COUPLER' in self.spec_type or 'BALUN' in self.spec_type: # type with 2 IL traces
            chart1.title='Insertion Loss'
            #~~~~~~~~~~~~~~~~~~~~~~~~~Calculate y-Axis~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
            try:
                t1_min=min([sublist[2] for sublist in self.chartdata])
                t2_min=min([sublist[3] for sublist in self.chartdata])
            except IndexError as e:
                return 0
            
            if t2_min<t1_min:
                y_min= t2_min
            else:
                y_min=t1_min
            
            t1_max=max([sublist[2] for sublist in self.chartdata])
            t2_max=max([sublist[3] for sublist in self.chartdata])
            if t2_max>t1_max:
                y_max=t2_max
            else:
                y_max=t1_max
            
            y_delta=y_max-y_min
            chart1.y_axis.scaling.min = y_min-(2*y_delta)
            chart1.y_axis.scaling.max = y_max+(2*y_delta)
            xvalues = Reference(self.sheet, min_col=1, min_row=50, max_row=249)
            yvalues = Reference(self.sheet, min_col=2, min_row=50, max_row=249)
            print('yvalues=',yvalues,' column=',2)
            series = Series(values=yvalues, xvalues=xvalues, title_from_data=True)
            chart1.series.append(series)
            yvalues = Reference(self.sheet, min_col=3, min_row=50, max_row=249)
            print('yvalues=',yvalues,' column=',3)
            series = Series(values=yvalues, xvalues=xvalues, title_from_data=True)
            chart1.series.append(series)
        else:
            chart1.title = "Insertion Loss"
            if not self.chartdata:
                return 0
            try:
                y_min=min([sublist[2] for sublist in self.chartdata])
                y_max=max([sublist[2] for sublist in self.chartdata])
            except IndexError as e:
                return 0
            
            y_delta=y_max-y_min
            chart1.y_axis.scaling.min = y_min-(2*y_delta)
            chart1.y_axis.scaling.max = y_max+(2*y_delta)
            xvalues = Reference(self.sheet, min_col=1, min_row=50, max_row=249)
            yvalues = Reference(self.sheet, min_col=2, min_row=50, max_row=249)
            series = Series(values=yvalues, xvalues=xvalues, title_from_data=True)
            print('yvalues=',yvalues)
            chart1.series.append(series)
        
        self.sheet.add_chart(chart1, "A6")
    
    def chart2(self):
        chart2 = ScatterChart()
        chart2.style = 12
        chart2.y_axis.title = 'dB'
        chart2.x_axis.title = 'Frequency MHz'
        chart2.x_axis.tickLblPos = "low"
        if not self.chartdata:
            return 0
        try:
            chart2.x_axis.scaling.min = min([sublist[1] for sublist in self.chartdata])
            chart2.x_axis.scaling.max = max([sublist[1] for sublist in self.chartdata])
        except IndexError as e:
            return 0
       
        chart2.legend = None
        chart2.width = 13
        chart2.height = 10
        chart2.title = "Return Loss"
        if not self.chartdata:
            return 0
        try:
            y_min=min([sublist[2] for sublist in self.chartdata])
            y_max=max([sublist[2] for sublist in self.chartdata])
            print('chart2 y_min=', y_min,'y_max=', y_max)
        except IndexError as e:
            return 0
       
        y_delta=y_max-y_min
        chart2.y_axis.scaling.min = y_min-(2*y_delta)
        chart2.y_axis.scaling.max = y_max+(2*y_delta)
        xvalues2 = Reference(self.sheet, min_col=4, min_row=50, max_row=249)
        yvalues2 = Reference(self.sheet, min_col=5, min_row=50, max_row=249)
        series = Series(values=yvalues2, xvalues=xvalues2, title_from_data=True)
        print('yvalues2=',yvalues2)
        chart2.series.append(series)
        self.sheet.add_chart(chart2, "G6")  
    
    def chart3(self):
        #print('self.spec_type=',self.spec_type)
        chart3 = ScatterChart()
        chart3.style = 11
        chart3.y_axis.title = 'dB'
        chart3.x_axis.title = 'Frequency MHz'
        chart3.x_axis.tickLblPos = "low"
        if not self.chartdata:
            return 0
        try:
            chart3.x_axis.scaling.min = min([sublist[1] for sublist in self.chartdata])
            chart3.x_axis.scaling.max = max([sublist[1] for sublist in self.chartdata])
        except IndexError as e:
            return 0
        
        chart3.legend = None
        chart3.width = 13
        chart3.height = 10
        chart3.title = "Isolation"
        if not self.chartdata:
            return 0
        try:
            y_min=min([sublist[2] for sublist in self.chartdata])
            y_max=max([sublist[2] for sublist in self.chartdata])
        except IndexError as e:
            return 0
        y_delta=y_max-y_min
        chart3.y_axis.scaling.min = y_min-(2*y_delta)
        chart3.y_axis.scaling.max = y_max+(2*y_delta)
        xvalues3 = Reference(self.sheet, min_col=6, min_row=50, max_row=249)
        yvalues3 = Reference(self.sheet, min_col=7, min_row=50, max_row=249)
        series = Series(values=yvalues3, xvalues=xvalues3, title_from_data=True)
        chart3.series.append(series)
        print('yvalues3=',yvalues3)
        self.sheet.add_chart(chart3, "A27")     

    def chart4(self):
        #print('self.spec_type=',self.spec_type)
        chart4 = ScatterChart()
        chart4.style = 10
        chart4.y_axis.title = 'dB'
        chart4.x_axis.title = 'Frequency MHz'
        chart4.x_axis.tickLblPos = "low"
        if not self.chartdata:
            return 0
        if not self.chartdata:
            return 0
        try:
            chart4.x_axis.scaling.min = min([sublist[1] for sublist in self.chartdata])
            chart4.x_axis.scaling.max = max([sublist[1] for sublist in self.chartdata])
        except IndexError as e:
            return 0
        
        chart4.legend = None
        chart4.width = 13
        chart4.height = 10
        if '90 DEGREE COUPLER' in self.spec_type or 'BALUN' in self.spec_type:
            chart4.title='Phase Balance'
            title1='Phase Balance J3'
            title2='Phase Balance J4'
        else:
            chart4.title='Coupled Flatness'
            title1='Coupled Flatness J3' 
            title2='Coupled Flatness J4'  
        #~~~~~~~~~~~~~~~~~~~~~~~~~Calculate y-Axis~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        if not self.chartdata:
            return 0
        try:
            y_min=min([sublist[2] for sublist in self.chartdata])
            y_max=max([sublist[2] for sublist in self.chartdata])
        except IndexError as e:
            return 0
        
        y_delta=y_max-y_min
        chart4.y_axis.scaling.min = y_min-(2*y_delta)
        chart4.y_axis.scaling.max = y_max+(2*y_delta)
        #~~~~~~~~~~~~~~~~~~~~~~~~~Calculate y-Axis~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        xvalues4 = Reference(self.sheet, min_col=8, min_row=50, max_row=249)
        yvalues4 = Reference(self.sheet, min_col=9, min_row=50, max_row=249)
        series = Series(values=yvalues4, xvalues=xvalues4, title_from_data=True)
        chart4.series.append(series)
        yvalues4 = Reference(self.sheet, min_col=10, min_row=50, max_row=249)
        series = Series(values=yvalues4, xvalues=xvalues4, title_from_data=True)
        print('yvalues4=',yvalues4)
        chart4.series.append(series)
        self.sheet.add_chart(chart4, "G27")
        
 
class LoadCharts:    
    def __init__ (self, job_num,part_num,spec_type,workbook,artwork_rev,artwork_len):
        self.job_num = job_num
        self.part_num = part_num
        self.spec_type = spec_type
        self.workbook = workbook
        self.artwork_rev=artwork_rev
        self.artwork_len = artwork_len
        print('loading charts')
    
    def charts(self):
        chart1_data = []
        chart_title = []
        chart3_chart_data = []
        chart4_chart_data = []
       
        f = []
        d = []
        d1 = []
        d2 = []
        rows = []
        chart_data = []
        x=0
        print('self.spec_type=',self.spec_type)
        for idx in range(5): 
            serialnumber = 'UUT' + str(idx+1)
            print('serialnumber=',serialnumber)
            #create new sheet and format
            new_sheetname = str(self.artwork_rev) + '_UUT' + str(idx+1)    
            print('new_sheetname=',new_sheetname)
            sheet = self.workbook.create_sheet(new_sheetname) 
            makesheet=CreateSheets(new_sheetname,sheet,self.artwork_len)
            makesheet.chart_data()
           
           #~~~~~~~~~~~Load Header~~~~~~
            sheet = self.workbook[new_sheetname]
            sheet['F2'] = self.job_num
            sheet['F3'] = self.part_num
            sheet['F4'] = self.spec_type
            #~~~~~~~~~~~Load Header~~~~~~
            all_charts = 0
            chart_data = []
            if '90 DEGREE COUPLER' in self.spec_type or 'BALUN' in self.spec_type:
                #~~~~~~~~~~~~~~~~~~~~Chart1~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                chart_data = []
                title1='Insertion Loss J3'
                title2='Insertion Loss J4'
                trace_id1 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title1).filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()
                trace_id2 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title2).filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()
                #~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J3~~~~~~~~~~~~~~~~~~~~~~~~
                print('trace_id 1=',trace_id1)
                trace_points = []
                f_list = []
                d1_list=[]
                x=0
                if not trace_id1:
                    serialnumber = 'UUT ' + str(idx+1)
                    #print('serialnumber=',serialnumber)
                    trace_id1 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title1).filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()
                    trace_id2 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title2).filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()
                if not trace_id1:
                    serialnumber = 'UUT  ' + str(idx+1)
                    #print('serialnumber=',serialnumber)
                    trace_id1 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title1).filter(serialnumber=serialnumber).values_list('id').last()
                    trace_id2 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title2).filter(serialnumber=serialnumber).values_list('id').last()    
                if not trace_id1:
                    serialnumber = 'UUT' + str(idx+1)
                    #print('serialnumber=',serialnumber)
                    trace_id1 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title1).filter(serialnumber=serialnumber).values_list('id').last()
                    trace_id2 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title2).filter(serialnumber=serialnumber).values_list('id').last()
                if not trace_id1:
                    serialnumber = 'UUT ' + str(idx+1)
                    #print('serialnumber=',serialnumber)
                    trace_id1 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title1).filter(serialnumber=serialnumber).values_list('id').last()
                    trace_id2 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title2).filter(serialnumber=serialnumber).values_list('id').last()
                if not trace_id1:
                    serialnumber = 'UUT  ' + str(idx+1)
                    #print('serialnumber=',serialnumber)
                    trace_id1 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title1).filter(serialnumber=serialnumber).values_list('id').last()
                    trace_id2 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title2).filter(serialnumber=serialnumber).values_list('id').last()   
                
                
                if trace_id1:
                    if trace_id1[0] > 171666:
                        trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id1[0]).all()
                    else:
                        trace_points = Tracepoints.objects.using('TEST').filter(traceid=trace_id1[0]).all()
                    #print('trace_points=',trace_points)
                    rownum=50
                    #print(len(trace_points))
                    x=0
                    for point in trace_points:
                        if point.xdata==0 or x>200:
                            break    # break here
                        #print('rownum=',rownum)
                        sheet.cell(row=rownum, column=1).value= round(point.xdata,2)
                        sheet.cell(row=rownum, column=2).value= round(point.ydata,2)
                        #print('rownum=',rownum,' point.xdata=',point.xdata,' point.ydata=',point.ydata)
                        f_list.append(round(point.xdata,2))
                        d1_list.append(round(point.ydata,2))
                        rownum+=1
                        x+=1 
                    #~~~~~~~~~~~~~~~~~~~~~~Insertion Loss J4~~~~~~~~~~~~~~~~~~~~~~~~
                    trace_points = []
                    title='Insertion Loss'
                    #print('trace_id2=',trace_id2)
                    #print('f_list=',len(f_list))
                    #print('d1_list=',len(d1_list))
                if trace_id2 and f_list:
                    if trace_id2[0] > 171666:
                        trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id2[0]).all()
                    else:
                        trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id2[0]).all()
                    rownum=50
                    print(len(trace_points))
                    x=0
                    for point in trace_points:
                        if point.xdata==0 or x>200:
                            break    # break here
                        sheet.cell(row=rownum, column=3).value= round(point.ydata,2)
                        d2=round(point.ydata,2)
                        #print('x=',x)
                        try:
                            chart_data.append([title,f_list[x],d1_list[x],d2])
                            #print('f=',f_list[x],'d1=',d1_list[x],'d2=',d2)
                        except IndexError as e:
                            print('indexError=',e)
                        
                        rownum+=1 
                        x+=1 
            else:
                #~~~~~~~~~~~~~~~~~~~~~~Insertion Loss ~~~~~~~~~~~~~~~~~~~~~~~~
                title='Insertion Loss'
                trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title).filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()
                
                trace_points = []
                if not trace_id:
                    serialnumber = 'UUT ' + str(idx+1)
                    #print('serialnumber=',serialnumber)
                    trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title).filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()
                if not trace_id:
                    serialnumber = 'UUT  ' + str(idx+1)
                    #print('serialnumber=',serialnumber)
                    trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title).filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last() 
                if not trace_id:
                    serialnumber = 'UUT' + str(idx+1)
                    #print('serialnumber=',serialnumber)
                    trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title).filter(serialnumber=serialnumber).values_list('id').last()
                if not trace_id:
                    serialnumber = 'UUT ' + str(idx+1)
                    #print('serialnumber=',serialnumber)
                    trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title).filter(serialnumber=serialnumber).values_list('id').last()
                if not trace_id:
                    serialnumber = 'UUT  ' + str(idx+1)
                    #print('serialnumber=',serialnumber)
                    trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title).filter(serialnumber=serialnumber).values_list('id').last()
                
                if trace_id:
                    print('trace_id=',trace_id[0])
                    if trace_id[0] > 171666:
                        trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
                    else:
                        trace_points = Tracepoints.objects.using('TEST').filter(traceid=trace_id[0]).all()
                    #print('trace_points=',trace_points)
                    rownum=50
                    x=0
                    for point in trace_points:
                        if point.xdata==0 or x>200:
                            break    # break here
                        sheet.cell(row=rownum, column=1).value= round(point.xdata,2)
                        sheet.cell(row=rownum, column=2).value= round(point.ydata,2)
                        #print('rownum=',rownum,' point.xdata=',point.xdata,' point.ydata=',point.ydata)
                        f=round(point.xdata,2)
                        d=round(point.ydata,2)
                        rownum+=1
                        x+=1 
                        chart_data.append([title,f,d])
            #print('chart1_data=',chart_data)
            if chart_data:            
                load_chart= MakeCharts(sheet,self.spec_type,chart_data)
                load_chart.chart1()
                all_charts =+ 1
                print('good chart1 data')
            else:
                print('no chart1 data')
            
                
            #~~~~~~~~~~~~~~~~~~~~Chart1~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
            
            #~~~~~~~~~~~~~~~~~~~~Chart2~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        
            title='Return Loss'
            #~~~~~~~~~~~~~~~~~~~~~~Return Loss~~~~~~~~~~~~~~~~~~~~~~~~
            chart_data = []
            trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title='Return Loss').filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()       
            trace_points = []
            if not trace_id:
                serialnumber = 'UUT ' + str(idx+1)
                #print('serialnumber=',serialnumber)
                trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title='Return Loss').filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()  
            if not trace_id:
                serialnumber = 'UUT  ' + str(idx+1)
                #print('serialnumber=',serialnumber)
                trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title='Return Loss').filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()  
            if not trace_id:
                serialnumber = 'UUT' + str(idx+1)
                #print('serialnumber=',serialnumber)
                trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title='Return Loss').filter(serialnumber=serialnumber).values_list('id').last()
            if not trace_id:
                serialnumber = 'UUT ' + str(idx+1)
                #print('serialnumber=',serialnumber)
                trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title='Return Loss').filter(serialnumber=serialnumber).values_list('id').last()
            if not trace_id:
                serialnumber = 'UUT  ' + str(idx+1)
                #print('serialnumber=',serialnumber)
                trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title='Return Loss').filter(serialnumber=serialnumber).values_list('id').last()
            
            if trace_id:
                if trace_id[0] > 171666:
                    trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
                else:
                    trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
                rownum=50
                x=0
                for point in trace_points:
                    if point.xdata==0 or x>200:
                        break    # break here
                    sheet.cell(row=rownum, column=4).value= round(point.xdata,2)
                    sheet.cell(row=rownum, column=5).value= round(point.ydata,2)
                    f=round(point.xdata,2)
                    d=round(point.ydata,2)
                    rownum+=1
                    x+=1 
                    chart_data.append([title,f,d])
            
            #print('chart2_data=',chart_data)
            if chart_data:
                load_chart= MakeCharts(sheet,self.spec_type,chart_data)
                load_chart.chart2()
                all_charts =+ 1
                print('good chart2 data')
            else:
                print('no chart2 data')
            
            #~~~~~~~~~~~~~~~~~~~~Chart2~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
            
            #~~~~~~~~~~~~~~~~~~~~Chart3~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
            #~~~~~~~~~~~~~~~~~~~~~~isolation~~~~~~~~~~~~~~~~~~~~~~~~
            chart_data = []
            trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title='Isolation').filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()    
            trace_points = []
            title='Isolation'
            if not trace_id:
                serialnumber = 'UUT ' + str(idx+1)
                #print('serialnumber=',serialnumber)
                trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title='Isolation').filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()  
            if not trace_id:
                serialnumber = 'UUT  ' + str(idx+1)
                #print('serialnumber=',serialnumber)
                trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title='Isolation').filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()  
            if not trace_id:
                serialnumber = 'UUT' + str(idx+1)
                #print('serialnumber=',serialnumber)
                trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title='Isolation').filter(serialnumber=serialnumber).values_list('id').last()
            if not trace_id:
                serialnumber = 'UUT ' + str(idx+1)
                #print('serialnumber=',serialnumber)
                trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title='Isolation').filter(serialnumber=serialnumber).values_list('id').last()
            if not trace_id:
                serialnumber = 'UUT  ' + str(idx+1)
                #print('serialnumber=',serialnumber)
                trace_id = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title='Isolation').filter(serialnumber=serialnumber).values_list('id').last()
            
            if trace_id:
                if trace_id[0] > 171666:
                    trace_points = Tracepoints2.objects.using('TEST').filter(traceid=trace_id[0]).all()
                else:
                    trace_points = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id[0]).all()
                rownum=50
                x=0
                for point in trace_points:
                    if point.xdata==0 or x>200:
                        break    # break here
                    sheet.cell(row=rownum, column=6).value= round(point.xdata,2)
                    sheet.cell(row=rownum, column=7).value= round(point.ydata,2)
                    f=round(point.xdata,2)
                    d=round(point.ydata,2)
                    rownum+=1
                    x+=1
                    chart_data.append([title,f,d])
            
            #print('chart3_data=',chart_data)
            if chart_data:
                load_chart= MakeCharts(sheet,self.spec_type,chart_data)
                load_chart.chart3()
                all_charts =+ 1
                print('good chart3 data')
            else:
                print('no chart3 data')
             
            #~~~~~~~~~~~~~~~~~~~~Chart3~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
            
            #~~~~~~~~~~~~~~~~~~~~Chart4~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
            title = 'Phase Balance'
            if '90 DEGREE COUPLER' in self.spec_type or 'BALUN' in self.spec_type:
                title1='Phase Balance J3'
                title2='Phase Balance J4'
            else:
                title1='Coupled Flatness J3'
                title2='Coupled Flatness J4'
            chart_data = [] 
            trace_id1 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title1).filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()
            trace_id2 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title2).filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()
        
            #~~~~~~~~~~~~~~~~~~~~~~Phase Balance J3 & J4~~~~~~~~~~~~~~~~~~~~~~~~
            trace_points = []
            chart_data = []
            f_list = []
            d1_list=[]
            if not trace_id1:
                serialnumber = 'UUT ' + str(idx+1)
                #print('serialnumber=',serialnumber)
                trace_id1 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title1).filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()
                trace_id2 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title2).filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()  
            if not trace_id1:
                serialnumber = 'UUT  ' + str(idx+1)
                #print('serialnumber=',serialnumber)
                trace_id1 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title1).filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()
                trace_id2 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title2).filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last() 
            if not trace_id1:
                serialnumber = 'UUT' + str(idx+1)
                #print('serialnumber=',serialnumber)
                trace_id1 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title1).filter(serialnumber=serialnumber).values_list('id').last()
                trace_id2 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title2).filter(serialnumber=serialnumber).filter(artwork_rev=self.artwork_rev).values_list('id').last()  
            if not trace_id1:
                serialnumber = 'UUT ' + str(idx+1)
               # print('serialnumber=',serialnumber)
                trace_id1 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title1).filter(serialnumber=serialnumber).values_list('id').last()
                trace_id2 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title2).filter(serialnumber=serialnumber).values_list('id').last()  
            if not trace_id1:
                serialnumber = 'UUT  ' + str(idx+1)
                #print('serialnumber=',serialnumber)
                trace_id1 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title1).filter(serialnumber=serialnumber).values_list('id').last()
                trace_id2 = Trace.objects.using('TEST').filter(jobnumber=self.job_num).filter(title=title2).filter(serialnumber=serialnumber).values_list('id').last()  
            
            if trace_id1:
                if trace_id1[0] > 171666:
                    trace_points1 = Tracepoints2.objects.using('TEST').filter(traceid=trace_id1[0]).all()
                else:
                    trace_points1 = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id1[0]).all()
            if trace_id2:
                if trace_id1[0] > 171666:
                    trace_points2 = Tracepoints2.objects.using('TEST').filter(traceid=trace_id2[0]).all()
                else:
                    trace_points2 = Tracepoints.objects.using('TEST').objects.using('TEST').filter(traceid=trace_id2[0]).all()
                    
                rownum=50
                x=0
                yarray=[]
                for point in trace_points2:
                    yarray.append(point.ydata)
                
                for point in trace_points1:
                    if point.xdata==0 or x>200:
                        break    # break here
                    ydata=point.ydata-yarray[x]
                    print('ydata',ydata,'x=',x)
                    sheet.cell(row=rownum, column=8).value= round(point.xdata,2)
                    sheet.cell(row=rownum, column=9).value= round(ydata,2)
                    f=round(point.xdata,2)
                    d=round(ydata,2)
                    chart_data.append([title,f,d])
                    rownum+=1
                    x+=1
            #print('chart4_data=',chart_data)
            if chart_data:
                load_chart= MakeCharts(sheet,self.spec_type,chart_data)
                load_chart.chart4()
                all_charts =+ 1
                print('good chart4 data')
            else:
                print('no chart4 data')
               
            if all_charts ==0:
                    print ('sheet names',self.workbook.sheetnames)
                    print ('No Chart data remove active sheet',sheet.title)
                    self.workbook.remove(sheet)
            else:
                print (sheet.title,' Has good data')
  
 
  
class Statistics:  
    def __init__(self,test1,test2,test3,test4,test5):
        self.test1 = test1
        self.test2 = test2
        self.test3 = test3
        self.test4 = test4
        self.test5 = test5
       
        
    
    def get_stats(self):
        #~~~~~~~~~~~~~~~~Statics and Summary ~~~~~~~~~~~~~~~~~~~~
        s1_list=[]
        s2_list=[]
        s3_list=[]
        s4_list=[]
        s5_list=[]
        if len(self.test1) > 1 and  len(self.test2) > 1 and  len(self.test3) > 1 and  len(self.test4) > 1 and  len(self.test5) > 1:# must have at least two tests
            try:
                #print('insertion_loss=',self.test1)
                s1_stdev = round(statistics.stdev(self.test1),2) #Standard deviation
                s1_var = round(statistics.variance(self.test1),2) #Variance
                s1_avg = round(statistics.mean(self.test1),2) #Mean Average
                s1_min = round(min(self.test1),2) #Min
                s1_max = round(max(self.test1),2) #Max
                s1_list = [s1_min,s1_max,s1_avg,s1_stdev]
                #print('il_list=',il_list)
            except BaseException as err:
                print(f"Unexpected {err=}, {type(err)=}")

            try:
                print('return_loss=',self.test2)
                s2_stdev = round(statistics.stdev(self.test2),2) #Standard deviation
                s2_var = round(statistics.variance(self.test2),2) #Variance
                s2_avg = round(statistics.mean(self.test2),2) #Mean Average
                s2_min = round(min(self.test2),2) #Min
                s2_max = round(max(self.test2),2) #Max
                s2_list = [s2_min,s2_max,s2_avg,s2_stdev]
            except BaseException as err:
                print(f"Unexpected {err=}, {type(err)=}")
            
            try:
                s3_stdev = round(statistics.stdev(self.test3),2) #Standard deviation
                s3_var = round(statistics.variance(self.test3),2) #Variance
                s3_avg = round(statistics.mean(self.test3),2) #Mean Average
                s3_min = round(min(self.test3),2) #Min
                s3_max = round(max(self.test3),2) #Max
                s3_list = [s3_min,s3_max,s3_avg,s3_stdev]
            except BaseException as err:
                print(f"Unexpected {err=}, {type(err)=}")
           
            try:
                s4_stdev = round(statistics.stdev(self.test4),2) #Standard deviation
                s4_var = round(statistics.variance(self.test4),2) #Variance
                s4_avg = round(statistics.mean(self.test4),2) #Mean Average
                s4_min = round(min(self.test4),2) #Min
                s4_max = round(max(self.test4),2) #Max
                s4_list = [s4_min,s4_max,s4_avg,s4_stdev]
            except BaseException as err:
                print(f"Unexpected {err=}, {type(err)=}")
           
            try:
                s5_stdev = round(statistics.stdev(self.test5),2) #Standard deviation
                s5_var = round(statistics.variance(self.test5),2) #Variance
                s5_avg = round(statistics.mean(self.test5),2) #Mean Average
                s5_min = round(min(self.test5),2) #Min
                s5_max = round(max(self.test5),2) #Max
                s5_list = [s5_min,s5_max,s5_avg,s5_stdev]
            except BaseException as err:
                print(f"Unexpected {err=}, {type(err)=}")
            
            stat_list = [s1_list,s2_list,s3_list,s4_list,s5_list]
           
            return stat_list
           
class XY_Chart:  
    def __init__(self,tests,specs,test):
        self.test1 = tests[0]
        self.test2 = tests[1]
        self.test3 = tests[2]
        self.test4 = tests[3]
        self.test5 = tests[4]
        self.spec1 = specs[0]
        self.spec2 = specs[1]
        self.spec3 = 0-specs[2]
        self.spec4 = specs[3]
        self.spec5 = specs[4]
        self.test = test
        
    def Chart_data(self):
        chart = []
        if self.test == 'test1':
            testing = self.test1
            spec = self.spec1
        if self.test == 'test2':
            testing = self.test2
            spec = self.spec2
        if self.test == 'test3':
            testing = self.test3
            spec = self.spec3    
        if self.test == 'test4':
            testing = self.test4
            spec = self.spec4
        if self.test == 'test5':
            testing = self.test5
            spec = self.spec5
        
        x=0
        for test in testing:
            chart.append((x,test))
            x+=1
        
        return chart

class XY_Hist:  
    def __init__(self,sd_list,sd_x_range):
        self.sd_list = sd_list
        self.sd_x_range = sd_x_range
            
    def data(self):
        xy = []
        x=0
        for test in self.sd_x_range:
            xy.append((self.sd_x_range[x],test))
            x+=1
        
        return xy        
            

class SDEV_Dist:
    def __init__(self,spec,data,sdev,Min,Max,mean):
        #print('##########################################spec=',spec)
        try:
            self.Min = Min 
            self.Max = Max
            self.data = data
            self.sdev = sdev
            self.mean = mean
            span = self.Max - self.Min
            step = int(len(data))
            self.stepsize = (span/step)
            self.height = max(data)
        except BaseException as err:
                print(f"Unexpected {err=}, {type(err)=}")
        
        
    def matlab(self): # pur matlab
        from math import exp, pow
        temp_list = []
        try:
            import numpy as np
            variance = pow(self.sdev, 2)
            x = np.linspace(-3 * self.sdev +  self.mean, 3 * self.sdev + self.mean , 100)
            x = np.arange(0.4,0.8,self.stepsize)
            temp_list = np.exp(-np.square(x-self.mean)/2*variance)/(np.sqrt(2*np.pi*variance))
        except ZeroDivisionError as e:
           print('matlab list error=',e)
        return temp_list
    
    def linspace(self): # includes linspace
        import numpy as np
        temp_list = np.linspace(-3 * self.sdev +  self.mean, 3 * self.sdev + self.mean , 100)
        return temp_list
        
    
    def gauss(self): # includes height
        chart = []
        try:      
            #gaussian distribution
            import numpy as np
            variance = np.square(self.sdev)
            x = np.arange(self.Min,self.Max,self.stepsize)
            f = (-np.square(x-self.mean)/2*variance)/(np.sqrt(2*np.pi*variance))
            y=0
            for sd in f:
                chart.append((x[y],sd))
                y+=1
        except ZeroDivisionError as e:
           print('gaussian dist error=',e)
        return chart
    
    def gauss_min_max(self): # includes height
        chart = []
        answer = [0,0]      
        try:
            #gaussian distribution
            import numpy as np
            variance = np.square(self.sdev)
            x = np.arange(self.Min,self.Max,self.stepsize)
            f = (-np.square(x-self.mean)/2*variance)/(np.sqrt(2*np.pi*variance))
            y=0
            for sd in f:
                chart.append(sd)
                
            Min = min(chart)
            Max = max(chart)
            answer = [Min,Max]
        except ZeroDivisionError as e:
           print('gaussian dist error=',e)
        return answer   
     


class X_Range:
    def __init__(self,data,spec,Min,Max,Mean):
        self.data = data
        if spec < Min:
            self.Min = spec
        else:
            self.Min = Min 
            
        if spec > Max:
            self.Max = spec
        else:
            self.Max = Max
            
        span = self.Max - self.Min
        step = int(len(data))
        self.stepsize = (span/step)
        #print('self.stepsize=',self.stepsize)
        self.Mean = Mean
        #print('self.Min=',self.Min)
        #print('self.Max=',self.Max)
        #print('self.Mean=',self.Mean)
        self.step = int(step) + 2

    def list(self):
        #print('stepsize =',self.stepsize)
        import numpy as np
        lower_list = np.linspace(self.Min, self.Mean, self.step)
        upper_list = np.linspace(self.Mean+self.stepsize,self.Max, self.step)
        #print('lower_lis=',lower_list)
        #print('upper_list=',upper_list)
        x_range_list = np.concatenate((lower_list,upper_list),axis=None)
        temp_list = []
        for temp in x_range_list:
            temp_list.append(round(temp,3))
        x_range_list = temp_list    
        return x_range_list

 
class Histogram_data:  
    def __init__(self,tests,specs,test):
        self.test1 = tests[0]
        self.test2 = tests[1]
        self.test3 = tests[2]
        self.test4 = tests[3]
        self.test5 = tests[4]
        self.spec1 = specs[0]
        self.spec2 = specs[1]
        self.spec3 = specs[2]
        self.spec4 = specs[3]
        self.spec5 = specs[4]
        self.test = test
        
    def Hist_data(self):
        bin1 = []
        bin2 = []
        bin3 = []
        bin4 = []
        bin5 = []
        bin6 = []
        bin7 = []
        bin8 = []
        bin9 = []
        bin10 = []
        bin11 = []
        bin12 = []
        bin13 = []
        bin14 = []
        bin15 = []
        bin16 = []
        bin17 = []
        bin18 = []
        bin19 = []
        bin20 = []
        bin21 = []
        bin22 = []
        bin23 = []
        bin24 = []
        bin25 = []
        bin26 = []
        bin27 = []
        bin28 = []
        bin29 = []
        bin30 = []
        bin31 = []
        bin32 = []
        bin33 = []
        bin34 = []
        bin35 = []
        bin36 = []
        bin37 = []
        bin38 = []
        bin39 = []
        bin40 = []
        bin41 = []
        bin42 = []
        bin43 = []
        bin44 = []
        bin45 = []
        bin46 = []
        bin47 = []
        bin48 = []
        bin49 = []
        bin50 = []
        bin51 = []
        bin52 = []
        bin53 = []
        bin54 = []
        bin55 = []
        bin56 = []
        bin57 = []
        bin58 = []
        bin59 = []
        bin60 = []
        bin61 = []
        bin62 = []
        bin63 = []
        bin64 = []
        bin65 = []
        bin66 = []
        bin67 = []
        bin68 = []
        bin69 = []
        bin70 = []
        bin71 = []
        bin72 = []
        bin73 = []
        bin74 = []
        bin75 = []
        bin76 = []
        bin77 = []
        bin78 = []
        bin79 = []
        bin80 = []
        bin81 = []
        bin82 = []
        bin83 = []
        bin84 = []
        bin85 = []
        bin86 = []
        bin87 = []
        bin88 = []
        bin89 = []
       
        
        if self.test == 'test1':
            testing = self.test1
            spec = self.spec1
        if self.test == 'test2':
            testing = self.test2
            spec = self.spec2
        if self.test == 'test3':
            testing = self.test3
            spec = self.spec3    
        if self.test == 'test4':
            testing = self.test4
            spec = self.spec4
        if self.test == 'test5':
            testing = self.test5
            spec = self.spec5
            
        try:
            for tst in testing:
                if abs(tst) < (spec*0.01) - spec:
                    bin1.append(abs(tst))
                elif abs(tst) < (spec*0.02) - spec:
                    bin2.append(tst)
                elif abs(tst) < (spec*0.03) - spec:
                    bin3.append(tst)
                elif abs(tst) < (spec*0.04) - spec:
                    bin4.append(tst)
                elif abs(tst) < (spec*0.05) - spec:
                    bin5.append(tst)
                elif abs(tst) < (spec*0.06) - spec:
                    bin6.append(tst)
                elif abs(tst) < (spec*0.07) - spec:
                    bin7.append(tst)
                elif abs(tst) < (spec*0.08) - spec:
                    bin8.append(tst)
                elif abs(tst) < (spec*0.09) - spec:
                    bin9.append(tst)
                elif abs(tst) < (spec*0.10) - spec:
                    bin10.append(tst)
                elif abs(tst) < (spec*0.11) - spec:
                    bin11.append(tst)
                elif abs(tst) < (spec*0.12) - spec:
                    bin12.append(tst)
                elif abs(tst) < (spec*0.13) - spec:
                    bin13.append(tst)
                elif abs(tst) < (spec*0.14) - spec:
                    bin14.append(tst)
                elif abs(tst) < (spec*0.15) - spec:
                    bin15.append(tst)
                elif abs(tst) < (spec*0.16) - spec:
                    bin15.append(tst)
                elif abs(tst) < (spec*0.17) - spec:
                    bin16.append(tst)
                elif abs(tst) < (spec*0.18) - spec:
                    bin17.append(tst)
                elif abs(tst) < (spec*0.19) - spec:
                    bin18.append(tst)
                elif abs(tst) < (spec*0.2) - spec:
                    bin19.append(tst)
                elif abs(tst) < (spec*0.3) - spec:
                    bin20.append(tst)
                elif abs(tst) < (spec*0.4) - spec:
                    bin21.append(tst)
                elif abs(tst) < (spec*0.5) - spec:
                    bin22.append(tst)
                elif abs(tst) < (spec*0.6) - spec:
                    bin23.append(tst)
                elif abs(tst) < (spec*0.7) - spec:
                    bin24.append(tst)
                elif abs(tst) < (spec*0.8) - spec:
                    bin25.append(tst)
                elif abs(tst) < (spec*0.9) - spec:
                    bin26.append(tst)
                elif abs(tst) < (spec*1) - spec:
                    bin27.append(tst)
                elif abs(tst) < (spec*1.1) - spec:
                    bin28.append(tst)
                elif abs(tst) < (spec*1.2) - spec:
                    bin29.append(tst)
                elif abs(tst) < (spec*1.3) - spec:
                    bin30.append(tst)
                elif abs(tst) < (spec*1.4) - spec:
                    bin31.append(tst)
                elif abs(tst) < (spec*1.5) - spec:
                    bin32.append(tst)
                elif abs(tst) < (spec*1.6) - spec:
                    bin33.append(tst)
                elif abs(tst) < (spec*1.7) - spec:
                    bin34.append(tst)
                elif abs(tst) < (spec*1.8) - spec:
                    bin35.append(tst)
                elif abs(tst) < (spec*1.9) - spec:
                    bin36.append(tst)
                elif abs(tst) < (spec*2) - spec:
                    bin37.append(tst)
                elif abs(tst) < (spec*2.1) - spec:
                    bin38.append(tst)
                elif abs(tst) < (spec*2.2) - spec:
                    bin39.append(tst)
                elif abs(tst) < (spec*2.3) - spec:
                    bin40.append(tst)
                elif abs(tst) < (spec*2.4) - spec:
                    bin41.append(tst)
                elif abs(tst) < (spec*2.5) - spec:
                    bin42.append(tst)
                elif abs(tst) < (spec*2.6) - spec:
                    bin43.append(tst)
                elif abs(tst) < (spec*2.7) - spec:
                    bin44.append(tst)
                elif abs(tst) < (spec*2.8) - spec:
                    bin45.append(tst)
                elif abs(tst) < (spec*2.9) - spec:
                    bin46.append(tst)
                elif abs(tst) < (spec*3) - spec:
                    bin47.append(tst)
                elif abs(tst) > (spec*3) + spec:
                    bin48.append(tst)
                elif abs(tst) > (spec*2.9) + spec:
                    bin49.append(tst)
                elif abs(tst) > (spec*2.8) + spec:
                    bin50.append(tst)
                elif abs(tst) > (spec*2.7) + spec:
                    bin51.append(tst)
                elif abs(tst) > (spec*2.6) + spec:
                    bin52.append(tst)
                elif abs(tst) > (spec*2.5) + spec:
                    bin53.append(tst)
                elif abs(tst) > (spec*2.4) + spec:
                    bin54.append(tst)
                elif abs(tst) > (spec*2.3) + spec:
                    bin55.append(tst)
                elif abs(tst) > (spec*2.2) + spec:
                    bin56.append(tst)
                elif abs(tst) > (spec*2.1) + spec:
                    bin57.append(tst)
                elif abs(tst) > (spec*2) + spec:
                    bin58.append(tst)
                elif abs(tst) > (spec*1.9) + spec:
                    bin59.append(tst)
                elif abs(tst) > (spec*1.8) + spec:
                    bin60.append(tst)
                elif abs(tst) > (spec*1.7) + spec:
                    bin61.append(tst)
                elif abs(tst) > (spec*1.6) + spec:
                    bin62.append(tst)
                elif abs(tst) > (spec*1.5) + spec:
                    bin63.append(tst)
                elif abs(tst) > (spec*2.4) + spec:
                    bin64.append(tst)
                elif abs(tst) > (spec*1.3) + spec:
                    bin65.append(tst)
                elif abs(tst) > (spec*1.2) + spec:
                    bin66.append(tst)
                elif abs(tst) > (spec*1.1) + spec:
                    bin67.append(tst)
                elif abs(tst) > (spec*1) + spec:
                    bin68.append(tst)
                elif abs(tst) > (spec*0.9) + spec:
                    bin69.append(tst)
                elif abs(tst) > (spec*0.8) + spec:
                    bin70.append(tst)
                elif abs(tst) > (spec*0.7) + spec:
                    bin71.append(tst)
                elif abs(tst) > (spec*0.6) + spec:
                    bin72.append(tst)
                elif abs(tst) > (spec*0.5) + spec:
                    bin73.append(tst)
                elif abs(tst) > (spec*0.4) + spec:
                    bin74.append(tst)
                elif abs(tst) > (spec*0.3) + spec:
                    bin75.append(tst)
                elif abs(tst) > (spec*0.2) + spec:
                    bin76.append(tst)
                elif abs(tst) > (spec*0.19) + spec:
                    bin77.append(tst)
                elif abs(tst) > (spec*0.18) + spec:
                    bin78.append(tst)
                elif abs(tst) > (spec*0.17) + spec:
                    bin79.append(tst)
                elif abs(tst) > (spec*0.16) + spec:
                    bin80.append(tst)
                elif abs(tst) > (spec*0.15) + spec:
                    bin81.append(tst)
                elif abs(tst) > (spec*0.14) + spec:
                    bin82.append(tst)
                elif abs(tst) > (spec*0.13) + spec:
                    bin83.append(tst)
                elif abs(tst) > (spec*0.12) + spec:
                    bin84.append(tst)
                elif abs(tst) > (spec*0.11) + spec:
                    bin85.append(tst)
                elif abs(tst) > (spec*0.1) + spec:
                    bin86.append(tst)
                elif abs(tst) > (spec*0.05) + spec:
                    bin87.append(tst)
                elif abs(tst) > (spec*0.02) + spec:
                    bin88.append(tst)
                elif abs(tst) > (spec*0.01) + spec:
                    bin89.append(tst)
        except BaseException as err:
            print(f"Unexpected {err=}, {type(err)=}")
                    
        
        
        hist = []
        if len(bin1)>0:
            hist.append((len(bin1),min(bin1),max(bin1)))
        if len(bin2)>0:
            hist.append((len(bin2),min(bin2),max(bin2)))
        if len(bin3)>0:
            hist.append((len(bin3),min(bin3),max(bin3)))
        if len(bin4)>0:
            hist.append((len(bin4),min(bin4),max(bin4)))
        if len(bin5)>0:
            hist.append((len(bin5),min(bin5),max(bin5)))
        if len(bin6)>0:
            hist.append((len(bin6),min(bin6),max(bin6)))
        if len(bin7)>0:
            hist.append((len(bin7),min(bin7),max(bin7)))
        if len(bin8)>0:
            hist.append((len(bin8),min(bin8),max(bin8)))
        if len(bin9)>0:
            hist.append((len(bin9),min(bin9),max(bin9)))
        if len(bin10)>0:
            hist.append((len(bin10),min(bin10),max(bin10)))
        if len(bin11)>0:
            hist.append((len(bin11),min(bin11),max(bin11)))
        if len(bin12)>0:
            hist.append((len(bin12),min(bin12),max(bin12)))
        if len(bin13)>0:
            hist.append((len(bin13),min(bin13),max(bin13)))
        if len(bin14)>0:
            hist.append((len(bin14),min(bin14),max(bin14)))
        if len(bin15)>0:
            hist.append((len(bin15),min(bin15),max(bin15)))
        if len(bin16)>0:
            hist.append((len(bin16),min(bin16),max(bin16)))
        if len(bin17)>0:
            hist.append((len(bin17),min(bin17),max(bin17)))
        if len(bin18)>0:
            hist.append((len(bin18),min(bin18),max(bin18)))
        if len(bin19)>0:
            hist.append((len(bin19),min(bin19),max(bin19)))
        if len(bin20)>0:
            hist.append((len(bin20),min(bin20),max(bin20)))
        if len(bin21)>0:
            hist.append((len(bin21),min(bin21),max(bin21)))
        if len(bin22)>0:
            hist.append((len(bin22),min(bin22),max(bin22)))
        if len(bin23)>0:
            hist.append((len(bin23),min(bin23),max(bin23)))
        if len(bin24)>0:
            hist.append((len(bin24),min(bin24),max(bin24)))
        if len(bin25)>0:
            hist.append((len(bin25),min(bin25),max(bin25)))
        if len(bin26)>0:
            hist.append((len(bin26),min(bin26),max(bin26)))
        if len(bin27)>0:
            hist.append((len(bin27),min(bin27),max(bin27)))
        if len(bin28)>0:
            hist.append((len(bin28),min(bin28),max(bin28)))
        if len(bin29)>0:
            hist.append((len(bin29),min(bin29),max(bin29)))
        if len(bin30)>0:
            hist.append((len(bin30),min(bin30),max(bin30)))
        if len(bin31)>0:
            hist.append((len(bin31),min(bin31),max(bin31)))
        if len(bin32)>0:
            hist.append((len(bin32),min(bin32),max(bin32)))
        if len(bin33)>0:
            hist.append((len(bin33),min(bin33),max(bin33)))
        if len(bin34)>0:
            hist.append((len(bin34),min(bin34),max(bin34)))
        if len(bin35)>0:
            hist.append((len(bin35),min(bin35),max(bin35)))
        if len(bin36)>0:
            hist.append((len(bin36),min(bin36),max(bin36)))
        if len(bin37)>0:    
            hist.append((len(bin37),min(bin37),max(bin37)))
        if len(bin38)>0:
            hist.append((len(bin38),min(bin38),max(bin38)))
        if len(bin39)>0:
            hist.append((len(bin39),min(bin39),max(bin39)))
        if len(bin40)>0:
            hist.append((len(bin40),min(bin40),max(bin40)))
        if len(bin41)>0:
            hist.append((len(bin41),min(bin41),max(bin41)))
        if len(bin42)>0:
            hist.append((len(bin42),min(bin42),max(bin42)))
        if len(bin43)>0:
            hist.append((len(bin43),min(bin43),max(bin43)))
        if len(bin44)>0:
            hist.append((len(bin44),min(bin44),max(bin44)))
        if len(bin45)>0:
            hist.append((len(bin45),min(bin45),max(bin45)))
        if len(bin46)>0:
            hist.append((len(bin46),min(bin46),max(bin46)))
        if len(bin47)>0:    
            hist.append((len(bin47),min(bin47),max(bin47)))
        if len(bin48)>0:
            hist.append((len(bin48),min(bin48),max(bin48)))
        if len(bin49)>0:
            hist.append((len(bin49),min(bin49),max(bin49)))
        if len(bin50)>0:
            hist.append((len(bin50),min(bin50),max(bin50)))
        if len(bin51)>0:
            hist.append((len(bin51),min(bin51),max(bin51)))
        if len(bin42)>0:
            hist.append((len(bin42),min(bin42),max(bin42)))
        if len(bin43)>0:
            hist.append((len(bin43),min(bin43),max(bin43)))
        if len(bin44)>0:
            hist.append((len(bin44),min(bin44),max(bin44)))
        if len(bin45)>0:
            hist.append((len(bin45),min(bin45),max(bin45)))
        if len(bin46)>0:
            hist.append((len(bin46),min(bin46),max(bin46)))
        if len(bin47)>0:    
            hist.append((len(bin47),min(bin47),max(bin47)))
        if len(bin48)>0:
            hist.append((len(bin48),min(bin48),max(bin48)))
        if len(bin49)>0:
            hist.append((len(bin49),min(bin49),max(bin49)))
        if len(bin50)>0:
            hist.append((len(bin50),min(bin50),max(bin50)))
        if len(bin51)>0:
            hist.append((len(bin51),min(bin51),max(bin51)))
        if len(bin42)>0:
            hist.append((len(bin42),min(bin42),max(bin42)))
        if len(bin43)>0:
            hist.append((len(bin43),min(bin43),max(bin43)))
        if len(bin44)>0:
            hist.append((len(bin44),min(bin44),max(bin44)))
        if len(bin45)>0:
            hist.append((len(bin45),min(bin45),max(bin45)))
        if len(bin46)>0:
            hist.append((len(bin46),min(bin46),max(bin46)))
        if len(bin47)>0:    
            hist.append((len(bin47),min(bin47),max(bin47)))
        if len(bin48)>0:
            hist.append((len(bin48),min(bin48),max(bin48)))
        if len(bin49)>0:
            hist.append((len(bin49),min(bin49),max(bin49)))
        if len(bin50)>0:
            hist.append((len(bin50),min(bin50),max(bin50)))
        if len(bin51)>0:
            hist.append((len(bin51),min(bin51),max(bin51)))
        if len(bin52)>0:
            hist.append((len(bin52),min(bin52),max(bin52)))
        if len(bin53)>0:
            hist.append((len(bin53),min(bin53),max(bin53)))
        if len(bin54)>0:
            hist.append((len(bin54),min(bin54),max(bin54)))
        if len(bin55)>0:
            hist.append((len(bin55),min(bin55),max(bin55)))
        if len(bin56)>0:
            hist.append((len(bin56),min(bin56),max(bin56)))
        if len(bin57)>0:    
            hist.append((len(bin57),min(bin57),max(bin57)))
        if len(bin58)>0:
            hist.append((len(bin58),min(bin58),max(bin58)))
        if len(bin59)>0:
            hist.append((len(bin59),min(bin59),max(bin59)))
        if len(bin60)>0:
            hist.append((len(bin60),min(bin60),max(bin60)))
        if len(bin61)>0:
            hist.append((len(bin61),min(bin61),max(bin61)))
        if len(bin62)>0:
            hist.append((len(bin42),min(bin42),max(bin62)))
        if len(bin63)>0:
            hist.append((len(bin63),min(bin63),max(bin63)))
        if len(bin64)>0:
            hist.append((len(bin64),min(bin64),max(bin64)))
        if len(bin65)>0:
            hist.append((len(bin65),min(bin65),max(bin65)))
        if len(bin66)>0:
            hist.append((len(bin66),min(bin66),max(bin66)))
        if len(bin67)>0:    
            hist.append((len(bin67),min(bin67),max(bin67)))
        if len(bin68)>0:
            hist.append((len(bin68),min(bin68),max(bin68)))
        if len(bin69)>0:
            hist.append((len(bin69),min(bin69),max(bin69)))
        if len(bin70)>0:
            hist.append((len(bin70),min(bin70),max(bin70)))
        if len(bin57)>0:
            hist.append((len(bin71),min(bin71),max(bin71)))
        if len(bin72)>0:
            hist.append((len(bin72),min(bin72),max(bin72)))
        if len(bin73)>0:
            hist.append((len(bin73),min(bin73),max(bin73)))
        if len(bin74)>0:
            hist.append((len(bin74),min(bin74),max(bin74)))
        if len(bin75)>0:
            hist.append((len(bin75),min(bin75),max(bin75)))
        if len(bin76)>0:
            hist.append((len(bin76),min(bin76),max(bin76)))
        if len(bin77)>0:    
            hist.append((len(bin77),min(bin77),max(bin77)))
        if len(bin78)>0:
            hist.append((len(bin78),min(bin78),max(bin78)))
        if len(bin79)>0:
            hist.append((len(bin79),min(bin79),max(bin79)))
        if len(bin80)>0:
            hist.append((len(bin80),min(bin80),max(bin80)))
        if len(bin81)>0:
            hist.append((len(bin81),min(bin81),max(bin81)))
        if len(bin82)>0:
            hist.append((len(bin82),min(bin82),max(bin82)))
        if len(bin83)>0:
            hist.append((len(bin83),min(bin83),max(bin83)))
        if len(bin84)>0:
            hist.append((len(bin84),min(bin84),max(bin84)))
        if len(bin85)>0:
            hist.append((len(bin85),min(bin85),max(bin85)))
        if len(bin86)>0:
            hist.append((len(bin86),min(bin86),max(bin86)))
        if len(bin87)>0:    
            hist.append((len(bin87),min(bin87),max(bin87)))
        if len(bin88)>0:
            hist.append((len(bin88),min(bin88),max(bin88)))
        if len(bin49)>0:
            hist.append((len(bin89),min(bin89),max(bin89)))
       
        return hist
        